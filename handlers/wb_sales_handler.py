# handlers/wb_sales_handler.py

import sys
import os
import pandas as pd
import logging
import re
import requests
import time
from datetime import datetime
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import CallbackContext, ConversationHandler
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Настройка путей
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)
utils_dir = os.path.join(root_dir, 'utils')

if root_dir not in sys.path:
    sys.path.append(root_dir)
if utils_dir not in sys.path:
    sys.path.append(utils_dir)

logger = logging.getLogger(__name__)

# Состояния
from states import WB_SALES_CABINET_CHOICE, WB_SALES_DATE_START, WB_SALES_DATE_END

# Импорт функции для работы с шаблоном
from utils.template_loader import get_cabinet_articles_by_template_id


def extract_period_from_filename(filename):
    """
    Извлекает период из имени файла WB.
    Формат: supplier-goods-1214047-2025-08-01-2025-08-31-szciqutca.XLSX

    Returns:
        (date_from, date_to) в формате строк 'YYYY-MM-DD'
    """
    # Паттерн для извлечения дат из имени файла
    pattern = r'(\d{4}-\d{2}-\d{2})-(\d{4}-\d{2}-\d{2})'
    match = re.search(pattern, filename)

    if match:
        date_from = match.group(1)
        date_to = match.group(2)
        logger.info(f"Извлечён период из файла: {date_from} - {date_to}")
        return date_from, date_to
    else:
        logger.warning(f"Не удалось извлечь период из имени файла: {filename}")
        return None, None


def get_wb_api_token(cabinet_id):
    """Получает API токен для указанного кабинета WB"""
    from dotenv import load_dotenv
    load_dotenv()

    if cabinet_id == 1:
        token = os.getenv('WB_API_TOKEN_1')
    elif cabinet_id == 2:
        token = os.getenv('WB_API_TOKEN_2')
    else:
        raise ValueError(f"Неподдерживаемый cabinet_id: {cabinet_id}")

    if not token:
        raise ValueError(f"WB_API_TOKEN_{cabinet_id} не найден в .env")

    return token


def fetch_wb_detailed_report(cabinet_id, date_from, date_to):
    """
    Получает детальный отчёт WB через API v5.

    Args:
        cabinet_id: ID кабинета (1 или 2)
        date_from: дата начала в формате 'YYYY-MM-DD'
        date_to: дата окончания в формате 'YYYY-MM-DD'

    Returns:
        list: список строк отчёта
    """
    token = get_wb_api_token(cabinet_id)

    url = "https://statistics-api.wildberries.ru/api/v5/supplier/reportDetailByPeriod"
    headers = {
        'Authorization': token
    }

    all_data = []
    rrdid = 0
    limit = 100000

    logger.info(f"Начинаю загрузку детального отчёта WB за период {date_from} - {date_to}")

    while True:
        params = {
            'dateFrom': date_from,
            'dateTo': date_to,
            'limit': limit,
            'rrdid': rrdid
        }

        try:
            response = requests.get(url, headers=headers, params=params, timeout=30)

            if response.status_code == 204:
                logger.info(f"Загрузка завершена. Получено {len(all_data)} записей")
                break

            response.raise_for_status()
            data = response.json()

            if not data:
                logger.info("Получен пустой ответ, загрузка завершена")
                break

            all_data.extend(data)
            logger.info(f"Загружено {len(data)} записей, всего: {len(all_data)}")

            # Получаем rrd_id из последней строки для следующего запроса
            if len(data) > 0:
                rrdid = data[-1].get('rrd_id', 0)
            else:
                break

            # Пауза между запросами
            time.sleep(0.5)

        except requests.exceptions.RequestException as e:
            logger.error(f"Ошибка при запросе к API WB: {e}")
            raise

    return all_data


def process_wb_api_data(api_data):
    """
    Обрабатывает данные из API WB и возвращает ТОЛЬКО финансовые расходы.

    Returns:
        expenses_data: {артикул: {logistics, storage, penalty, acceptance, damage_comp, return_comp, additional_payment}}
        art_original_case: {артикул_lowercase: Артикул_OriginalCase}
    """
    expenses_data = {}
    art_original_case = {}

    # Общие расходы (без привязки к конкретному артикулу)
    general_expenses = {
        'logistics': 0,
        'storage': 0,
        'penalty': 0,
        'acceptance': 0,
        'additional_payment': 0
    }

    logger.info(f"Обработка {len(api_data)} записей из API WB...")

    # Безопасная конвертация в float
    def safe_float(value):
        """Безопасная конвертация в float"""
        if value is None or value == '':
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    for row in api_data:
        # Получаем артикул продавца
        sa_name = row.get('sa_name')

        # Финансовые данные
        delivery_rub_raw = row.get('delivery_rub')
        storage_fee_raw = row.get('storage_fee')
        penalty_raw = row.get('penalty')
        acceptance_raw = row.get('acceptance')
        additional_payment_raw = row.get('additional_payment')

        # Конвертируем в числа
        delivery_rub = abs(safe_float(delivery_rub_raw))
        storage_fee = safe_float(storage_fee_raw)
        penalty = abs(safe_float(penalty_raw))
        acceptance = safe_float(acceptance_raw)
        additional_payment = safe_float(additional_payment_raw)

        # Если артикула НЕТ - это общие расходы (хранение, приёмка и т.д.)
        if not sa_name:
            general_expenses['logistics'] += delivery_rub
            general_expenses['storage'] += storage_fee
            general_expenses['penalty'] += penalty
            general_expenses['acceptance'] += acceptance
            general_expenses['additional_payment'] += additional_payment
            continue

        art_original = str(sa_name).strip()
        art = art_original.lower()

        # Сохраняем оригинальный регистр
        if art not in art_original_case:
            art_original_case[art] = art_original

        # Инициализируем expenses_data для артикула
        if art not in expenses_data:
            expenses_data[art] = {
                'logistics': 0,
                'storage': 0,
                'penalty': 0,
                'acceptance': 0,
                'additional_payment': 0
            }

        # Суммируем расходы по артикулам
        expenses_data[art]['logistics'] += delivery_rub
        expenses_data[art]['storage'] += storage_fee
        expenses_data[art]['penalty'] += penalty
        expenses_data[art]['acceptance'] += acceptance
        expenses_data[art]['additional_payment'] += additional_payment


    # Логируем итоговые суммы
    total_logistics = sum(exp['logistics'] for exp in expenses_data.values()) + general_expenses['logistics']
    total_storage = sum(exp['storage'] for exp in expenses_data.values()) + general_expenses['storage']
    total_penalty = sum(exp['penalty'] for exp in expenses_data.values()) + general_expenses['penalty']
    total_acceptance = sum(exp['acceptance'] for exp in expenses_data.values()) + general_expenses['acceptance']

    logger.info(f"Обработано {len(expenses_data)} артикулов из API")
    logger.info(f"Логистика: {total_logistics:.2f}, Хранение: {total_storage:.2f}, Штрафы: {total_penalty:.2f}, Приёмка: {total_acceptance:.2f}")

    return expenses_data, general_expenses, art_original_case



async def start_wb_sales(update: Update, context: CallbackContext) -> int:
    """Начало процесса выгрузки продаж WB (через файлы)"""
    context.user_data['current_flow'] = 'wb_sales'

    keyboard = [
        [InlineKeyboardButton("🏪 WB_1 Nimba", callback_data='wb_cabinet_1')],
        [InlineKeyboardButton("🏬 WB_2 Galioni", callback_data='wb_cabinet_2')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    sent_message = await update.message.reply_text(
        "🏢 Выберите кабинет Wildberries для выгрузки продаж:\n\n"
        "⚠️ После выбора кабинета загрузите файл отчёта о продажах WB",
        reply_markup=reply_markup
    )
    context.user_data['wb_sales_initial_message_id'] = sent_message.message_id

    return WB_SALES_CABINET_CHOICE


async def handle_wb_sales_cabinet_choice(update: Update, context: CallbackContext) -> int:
    """Обработка выбора кабинета WB"""
    query = update.callback_query
    await query.answer()

    cabinet_data = query.data
    if cabinet_data not in ('wb_cabinet_1', 'wb_cabinet_2'):
        await query.message.reply_text("❌ Неизвестный кабинет.")
        return ConversationHandler.END

    cabinet_id = 1 if cabinet_data == 'wb_cabinet_1' else 2
    context.user_data['wb_sales_cabinet_id'] = cabinet_id

    await query.message.edit_reply_markup(reply_markup=None)
    await query.message.reply_text(
        f"✅ Выбран кабинет: WB {cabinet_id}\n\n"
        "📤 Теперь отправьте файл отчёта о продажах Wildberries (.xlsx)\n\n"
        "Файл должен содержать столбцы:\n"
        "• Артикул продавца\n"
        "• шт. (заказы)\n"
        "• Выкупили, шт.\n"
        "• К перечислению за товар, руб."
    )
    return WB_SALES_DATE_START


async def handle_wb_sales_date_start(update: Update, context: CallbackContext) -> int:
    """Обработка загруженного файла WB"""
    if not update.message.document:
        await update.message.reply_text(
            "❌ Пожалуйста, отправьте файл Excel (.xlsx)"
        )
        return WB_SALES_DATE_START

    document = update.message.document
    file_name = document.file_name

    # Проверка формата файла
    if not file_name.lower().endswith('.xlsx'):
        await update.message.reply_text(
            "❌ Файл должен быть в формате Excel (.xlsx)"
        )
        return WB_SALES_DATE_START

    try:
        # Скачивание файла
        file = await context.bot.get_file(document)
        file_path = f"temp_wb_sales_{update.effective_user.id}.xlsx"
        await file.download_to_drive(file_path)

        # Сохраняем путь к файлу
        context.user_data['wb_sales_file_path'] = file_path
        context.user_data['wb_sales_file_name'] = file_name


        # Переходим к генерации отчёта
        return await handle_wb_sales_date_end(update, context)

    except Exception as e:
        logger.error(f"Ошибка при загрузке файла WB: {e}", exc_info=True)
        await update.message.reply_text(
            f"❌ Ошибка при загрузке файла: {str(e)}"
        )
        return ConversationHandler.END


async def handle_wb_sales_date_end(update: Update, context: CallbackContext) -> int:
    """Генерация отчёта WB: количественные данные из файла + расходы из API"""
    cabinet_id = context.user_data.get('wb_sales_cabinet_id', 1)
    file_path = context.user_data.get('wb_sales_file_path')
    file_name = context.user_data.get('wb_sales_file_name', '')

    if not file_path or not os.path.exists(file_path):
        await update.message.reply_text("❌ Файл не найден")
        return ConversationHandler.END

    try:
        # Отправляем временное сообщение о прогрессе
        progress_msg = await update.message.reply_text(
            f"✅ <b>Файл получен:</b> {file_name}\n\n"
            f"⏳ <b>Генерирую отчёт...</b>",
            parse_mode="HTML"
        )

        # === 1. Обрабатываем файл Excel (количественные данные + "К перечислению за товар") ===
        orders_data, purchases_data, cancels_data, income_data, art_original_case_file = process_wb_sales_file(file_path)

        logger.info(f"Из файла: заказы={sum(orders_data.values())}, выкупы={sum(purchases_data.values())}")

        # === 2. Загружаем расходы из API ===
        date_from, date_to = extract_period_from_filename(file_name)

        if date_from and date_to:
            # Обновляем сообщение о прогрессе
            await progress_msg.edit_text(
                f"✅ <b>Файл получен:</b> {file_name}\n"
                f"✅ <b>Файл обработан</b>\n\n"
                f"📅 <b>Период:</b> {date_from} — {date_to}\n"
                f"⏳ <b>Загружаю финансовые данные из API...</b>",
                parse_mode="HTML"
            )

            try:
                # Загружаем данные из API
                api_data = fetch_wb_detailed_report(cabinet_id, date_from, date_to)

                if api_data:
                    # Обновляем сообщение о прогрессе
                    await progress_msg.edit_text(
                        f"✅ <b>Файл получен:</b> {file_name}\n"
                        f"✅ <b>Файл обработан</b>\n"
                        f"✅ <b>Загружено {len(api_data)} записей из API</b>\n\n"
                        f"⏳ <b>Формирую отчёт...</b>",
                        parse_mode="HTML"
                    )

                    # Обрабатываем только расходы из API
                    expenses_data, general_expenses, art_original_case_api = process_wb_api_data(api_data)

                    # Объединяем оригинальные регистры из файла и API
                    art_original_case = {**art_original_case_file, **art_original_case_api}
                else:
                    logger.warning("API вернул пустой ответ, расходы будут нулевыми")
                    expenses_data = {}
                    general_expenses = {
                        'logistics': 0, 'storage': 0, 'penalty': 0, 'acceptance': 0,
                        'additional_payment': 0
                    }
                    art_original_case = art_original_case_file

            except Exception as api_error:
                logger.error(f"Ошибка при загрузке данных из API: {api_error}", exc_info=True)
                expenses_data = {}
                general_expenses = {
                    'logistics': 0, 'storage': 0, 'penalty': 0, 'acceptance': 0,
                    'additional_payment': 0
                }
                art_original_case = art_original_case_file
        else:
            logger.warning("Не удалось извлечь период из имени файла, расходы будут нулевыми")
            expenses_data = {}
            general_expenses = {
                'logistics': 0, 'storage': 0, 'penalty': 0, 'acceptance': 0,
                'additional_payment': 0
            }
            art_original_case = art_original_case_file

            # Обновляем сообщение о прогрессе
            await progress_msg.edit_text(
                f"✅ <b>Файл получен:</b> {file_name}\n"
                f"✅ <b>Файл обработан</b>\n\n"
                f"⏳ <b>Формирую отчёт...</b>",
                parse_mode="HTML"
            )

        # === 3. Загружаем шаблон артикулов ===
        template_id_to_name, template_id_to_cabinet_arts = get_cabinet_articles_by_template_id("Шаблон_WB")

        # === 4. Распределяем общие расходы пропорционально доходу ===
        total_income_sum = sum(income_data.values())

        # Распределяем general_expenses по артикулам пропорционально их доходу
        for art in income_data:
            if art not in expenses_data:
                expenses_data[art] = {
                    'logistics': 0,
                    'storage': 0,
                    'penalty': 0,
                    'acceptance': 0,
                    'additional_payment': 0
                }

            # Доля этого артикула в общем доходе
            income_share = income_data[art] / total_income_sum if total_income_sum > 0 else 0

            # Распределяем общие расходы пропорционально
            expenses_data[art]['logistics'] += general_expenses['logistics'] * income_share
            expenses_data[art]['storage'] += general_expenses['storage'] * income_share
            expenses_data[art]['penalty'] += general_expenses['penalty'] * income_share
            expenses_data[art]['acceptance'] += general_expenses['acceptance'] * income_share
            expenses_data[art]['additional_payment'] += general_expenses['additional_payment'] * income_share

        # === 5. Рассчитываем "Итого к перечислению" для каждого артикула ===
        total_payout_data = {}
        for art in income_data:
            payout = income_data.get(art, 0)
            exp = expenses_data.get(art, {})

            # Вычитаем расходы (уже включая распределённые общие расходы)
            payout -= exp.get('logistics', 0)
            payout -= exp.get('storage', 0)
            payout -= exp.get('penalty', 0)
            payout -= exp.get('acceptance', 0)

            # Добавляем доплаты
            payout += exp.get('additional_payment', 0)

            total_payout_data[art] = payout

        # === 6. Группируем данные ===
        grouped, unmatched, raw_art_data = group_wb_sales_data_v2(
            orders_data, purchases_data, cancels_data, income_data,
            expenses_data, total_payout_data,
            template_id_to_name, template_id_to_cabinet_arts, art_original_case
        )

        # === 6. Считаем общие показатели ===
        total_orders = int(sum(orders_data.values()))
        total_purchases = int(sum(purchases_data.values()))
        total_cancels = int(sum(cancels_data.values()))
        total_income = sum(income_data.values())
        total_payout = sum(total_payout_data.values())

        # Считаем общие расходы (уже включают распределённые general_expenses)
        total_logistics = sum(exp.get('logistics', 0) for exp in expenses_data.values())
        total_storage = sum(exp.get('storage', 0) for exp in expenses_data.values())
        total_penalty = sum(exp.get('penalty', 0) for exp in expenses_data.values())
        total_acceptance = sum(exp.get('acceptance', 0) for exp in expenses_data.values())
        total_additional = sum(exp.get('additional_payment', 0) for exp in expenses_data.values())

        # Рассчитываем итого к перечислению
        total_payout_final = (total_income
                            - total_logistics
                            - total_storage
                            - total_penalty
                            - total_acceptance
                            + total_additional)

        # === 7. Создаём отчёт ===
        period_str = f"{date_from}_{date_to}" if date_from and date_to else "unknown"
        output_path = f"WB_Sales_Report_Cabinet_{cabinet_id}_{period_str}.xlsx"

        create_wb_excel_report_v2(
            grouped, unmatched, template_id_to_name,
            sorted(template_id_to_name.keys()),
            output_path,
            total_orders, total_purchases, total_cancels,
            total_income, total_payout_final,
            total_logistics, total_storage, total_penalty, total_acceptance,
            total_additional,
            raw_art_data
        )

        # === 8. Формируем текстовую сводку ===
        top_5 = raw_art_data[:5] if raw_art_data else []

        def fmt_num(x):
            """Форматирование чисел с пробелами"""
            if isinstance(x, float):
                return f"{x:,.2f}".replace(",", " ")
            elif isinstance(x, int):
                return f"{x:,}".replace(",", " ")
            return str(x)

        total_shipments = total_purchases + total_cancels
        purchase_percent = (total_purchases / total_shipments * 100) if total_shipments > 0 else 0
        avg_payout_per_unit = total_payout_final / total_purchases if total_purchases > 0 else 0

        text_summary = (
            f"📊 <b>Сводка по продажам Wildberries</b>\n"
            f"Кабинет: <b>WB {cabinet_id}</b>\n"
        )

        if date_from and date_to:
            text_summary += f"Период: <b>{date_from} - {date_to}</b>\n\n"
        else:
            text_summary += "\n"

        text_summary += (
            f"📦 <b>Заказы:</b> {fmt_num(total_orders)} шт\n"
            f"✅ <b>Выкупы:</b> {fmt_num(total_purchases)} шт\n"
            f"❌ <b>Отмены:</b> {fmt_num(total_cancels)} шт\n"
            f"💵 <b>Итого к перечислению:</b> {fmt_num(total_payout_final)} ₽\n"
            f"📈 <b>К перечислению на 1 ед:</b> {fmt_num(avg_payout_per_unit)} ₽\n"
            f"🔄 <b>Процент выкупов:</b> {purchase_percent:.2f}%\n"
        )

        if total_logistics > 0 or total_storage > 0 or total_penalty > 0 or total_acceptance > 0:
            text_summary += f"\n📉 <b>Расходы:</b>\n"
            if total_logistics > 0:
                text_summary += f"  • Стоимость логистики: {fmt_num(total_logistics)} ₽\n"
            if total_storage > 0:
                text_summary += f"  • Стоимость хранения: {fmt_num(total_storage)} ₽\n"
            if total_penalty > 0:
                text_summary += f"  • Штрафы: {fmt_num(total_penalty)} ₽\n"
            if total_acceptance > 0:
                text_summary += f"  • Операции при приёмке: {fmt_num(total_acceptance)} ₽\n"

        text_summary += (
            f"\n💰 <b>Прибыль до вычета расходов:</b> {fmt_num(total_income)} ₽\n"
            f"\n🏆 <b>Топ-5 артикулов по выкупам:</b>\n"
        )

        if top_5:
            for i, item in enumerate(top_5, 1):
                art = item["art"]
                purchases = item["purchases"]
                payout = item["payout"]
                text_summary += (
                    f"🔹 {i}. <b>{art}</b>\n"
                    f"   ✅ Выкупы: {fmt_num(int(purchases))} шт\n"
                    f"   💰 К перечислению: {fmt_num(payout)} ₽\n\n"
                )
        else:
            text_summary += "   — Нет данных по выкупам\n"

        # === 9. Отправка отчёта ===
        caption = f"📊 Подробный отчёт в Excel по продажам Wildberries (кабинет {cabinet_id})"
        if date_from and date_to:
            caption += f"\nПериод: {date_from} - {date_to}"

        await update.message.reply_document(
            document=open(output_path, 'rb'),
            caption=caption
        )

        await update.message.reply_text(
            text_summary,
            parse_mode="HTML"
        )

        # Удаляем временное сообщение о прогрессе
        try:
            await progress_msg.delete()
        except Exception:
            pass  # Игнорируем ошибку если сообщение уже удалено

        # Удаляем временные файлы
        if os.path.exists(file_path):
            os.remove(file_path)
        if os.path.exists(output_path):
            os.remove(output_path)

        logger.info(f"✅ Отчёт WB успешно создан для кабинета {cabinet_id}")

    except Exception as e:
        logger.error(f"Ошибка при создании отчёта WB: {e}", exc_info=True)
        await update.message.reply_text(
            f"❌ Ошибка при создании отчёта:\n{str(e)}"
        )
        # Удаляем временный файл в случае ошибки
        if file_path and os.path.exists(file_path):
            os.remove(file_path)

    return ConversationHandler.END


def process_wb_sales_file(file_path):
    """
    Обрабатывает файл продаж WB и возвращает данные по артикулам.

    Returns:
        orders_data: {артикул: количество_заказов}
        purchases_data: {артикул: количество_выкупов}
        cancels_data: {артикул: количество_отмен}
        income_data: {артикул: сумма_к_перечислению}
    """
    logger.info(f"Обработка файла WB: {file_path}")

    # Поиск правильной строки заголовков
    df = None
    for header_row in range(10):
        try:
            temp_df = pd.read_excel(file_path, header=header_row)
            required_columns = [
                'Артикул продавца',
                'шт.',
                'Выкупили, шт.',
                'К перечислению за товар, руб.'
            ]

            if all(col in temp_df.columns for col in required_columns):
                df = temp_df
                logger.info(f"Найдены заголовки в строке {header_row}")
                break
        except Exception as e:
            continue

    if df is None:
        raise ValueError("Не найдены требуемые столбцы в файле WB")

    orders_data = {}
    purchases_data = {}
    cancels_data = {}
    income_data = {}
    art_original_case = {}  # Сохраняем оригинальный регистр

    for _, row in df.iterrows():
        # Получаем артикул
        art_raw = row.get('Артикул продавца')
        if pd.isna(art_raw):
            continue

        art_original = str(art_raw).strip()
        art = art_original.lower()

        if not art or art == 'nan':
            continue

        # Сохраняем оригинальный регистр (при первой встрече)
        if art not in art_original_case:
            art_original_case[art] = art_original

        # Получаем значения
        ordered = row.get('шт.', 0)
        purchased = row.get('Выкупили, шт.', 0)
        amount = row.get('К перечислению за товар, руб.', 0)

        # Проверяем, что это числа
        try:
            ordered = float(ordered) if not pd.isna(ordered) else 0
            purchased = float(purchased) if not pd.isna(purchased) else 0
            amount = float(amount) if not pd.isna(amount) else 0
        except (ValueError, TypeError):
            continue

        # Суммируем данные
        orders_data[art] = orders_data.get(art, 0) + ordered
        purchases_data[art] = purchases_data.get(art, 0) + purchased
        income_data[art] = income_data.get(art, 0) + amount

    # Рассчитываем отмены
    for art in orders_data:
        cancels = orders_data[art] - purchases_data.get(art, 0)
        cancels_data[art] = max(0, cancels)

    logger.info(f"Обработано {len(purchases_data)} артикулов из файла")

    return orders_data, purchases_data, cancels_data, income_data, art_original_case


def group_wb_sales_data(orders_data, purchases_data, cancels_data, income_data,
                        template_id_to_name, template_id_to_cabinet_arts, art_original_case):
    """
    Группирует данные продаж WB по шаблонным артикулам.

    Returns:
        grouped: {template_id: {...}}
        unmatched: {артикул: {...}}
        raw_art_data: [список артикулов с данными для листа "Исходные артикулы"]
    """
    # Создаём обратный маппинг: артикул -> template_id
    art_to_template_id = {}
    for template_id, arts in template_id_to_cabinet_arts.items():
        for art in arts:
            art_lower = art.strip().lower()
            art_to_template_id[art_lower] = template_id

    grouped = {}
    unmatched = {}
    raw_art_data = []

    # Получаем все уникальные артикулы
    all_arts = set(purchases_data.keys()) | set(orders_data.keys()) | set(income_data.keys())

    for art in all_arts:
        orders = orders_data.get(art, 0)
        purchases = purchases_data.get(art, 0)
        cancels = cancels_data.get(art, 0)
        income = income_data.get(art, 0)

        # Добавляем в raw_art_data для листа "Исходные артикулы"
        total_shipments = purchases + cancels
        purchase_percent = (purchases / total_shipments * 100) if total_shipments > 0 else 0
        profit_per_unit = income / purchases if purchases > 0 else 0

        # Используем оригинальный регистр для отображения
        art_display = art_original_case.get(art, art)

        raw_art_data.append({
            "art": art_display,  # Оригинальный регистр
            "purchases": purchases,
            "profit": income,
            "purchase_percent": purchase_percent,
            "profit_per_unit": profit_per_unit,
            "orders": orders,
            "cancels": cancels
        })

        # Ищем соответствие в шаблоне
        template_id = art_to_template_id.get(art)

        if template_id is not None:
            # Артикул найден в шаблоне
            if template_id not in grouped:
                grouped[template_id] = {
                    'name': template_id_to_name.get(template_id, f"ID {template_id}"),
                    'orders': 0,
                    'purchases': 0,
                    'cancels': 0,
                    'income': 0
                }

            grouped[template_id]['orders'] += orders
            grouped[template_id]['purchases'] += purchases
            grouped[template_id]['cancels'] += cancels
            grouped[template_id]['income'] += income
        else:
            # Неопознанный артикул
            unmatched[art_display] = {  # Оригинальный регистр
                'name': f"НЕОПОЗНАННЫЙ_АРТИКУЛ: {art_display}",
                'orders': orders,
                'purchases': purchases,
                'cancels': cancels,
                'income': income
            }

    # Сортируем raw_art_data по выкупам (убывание)
    raw_art_data.sort(key=lambda x: x["purchases"], reverse=True)

    logger.info(f"Сгруппировано шаблонов: {len(grouped)}")
    logger.info(f"Неопознанных артикулов: {len(unmatched)}")

    return grouped, unmatched, raw_art_data


def group_wb_sales_data_v2(orders_data, purchases_data, cancels_data, income_data,
                            expenses_data, total_payout_data,
                            template_id_to_name, template_id_to_cabinet_arts, art_original_case):
    """
    Группирует данные продаж WB по шаблонным артикулам (версия 2 с расходами).

    Returns:
        grouped: {template_id: {...}}
        unmatched: {артикул: {...}}
        raw_art_data: [список артикулов с данными для листа "Исходные артикулы"]
    """
    # Создаём обратный маппинг: артикул -> template_id
    art_to_template_id = {}
    for template_id, arts in template_id_to_cabinet_arts.items():
        for art in arts:
            art_lower = art.strip().lower()
            art_to_template_id[art_lower] = template_id

    grouped = {}
    unmatched = {}
    raw_art_data = []

    # Получаем все уникальные артикулы
    all_arts = set(purchases_data.keys()) | set(orders_data.keys()) | set(income_data.keys())

    for art in all_arts:
        orders = orders_data.get(art, 0)
        purchases = purchases_data.get(art, 0)
        cancels = cancels_data.get(art, 0)
        income = income_data.get(art, 0)
        payout = total_payout_data.get(art, 0)
        exp = expenses_data.get(art, {})

        # Добавляем в raw_art_data для листа "Исходные артикулы"
        total_shipments = purchases + cancels
        purchase_percent = (purchases / total_shipments * 100) if total_shipments > 0 else 0
        payout_per_unit = payout / purchases if purchases > 0 else 0

        # Используем оригинальный регистр для отображения
        art_display = art_original_case.get(art, art)

        raw_art_data.append({
            "art": art_display,
            "purchases": purchases,
            "income": income,
            "payout": payout,
            "logistics": exp.get('logistics', 0),
            "storage": exp.get('storage', 0),
            "penalty": exp.get('penalty', 0),
            "acceptance": exp.get('acceptance', 0),
            "damage_comp": exp.get('damage_comp', 0),
            "return_comp": exp.get('return_comp', 0),
            "additional_payment": exp.get('additional_payment', 0),
            "purchase_percent": purchase_percent,
            "payout_per_unit": payout_per_unit,
            "orders": orders,
            "cancels": cancels
        })

        # Ищем соответствие в шаблоне
        template_id = art_to_template_id.get(art)

        if template_id is not None:
            # Артикул найден в шаблоне
            if template_id not in grouped:
                grouped[template_id] = {
                    'name': template_id_to_name.get(template_id, f"ID {template_id}"),
                    'orders': 0,
                    'purchases': 0,
                    'cancels': 0,
                    'income': 0,
                    'payout': 0,
                    'logistics': 0,
                    'storage': 0,
                    'penalty': 0,
                    'acceptance': 0,
                    'damage_comp': 0,
                    'return_comp': 0,
                    'additional_payment': 0
                }

            grouped[template_id]['orders'] += orders
            grouped[template_id]['purchases'] += purchases
            grouped[template_id]['cancels'] += cancels
            grouped[template_id]['income'] += income
            grouped[template_id]['payout'] += payout
            grouped[template_id]['logistics'] += exp.get('logistics', 0)
            grouped[template_id]['storage'] += exp.get('storage', 0)
            grouped[template_id]['penalty'] += exp.get('penalty', 0)
            grouped[template_id]['acceptance'] += exp.get('acceptance', 0)
            grouped[template_id]['damage_comp'] += exp.get('damage_comp', 0)
            grouped[template_id]['return_comp'] += exp.get('return_comp', 0)
            grouped[template_id]['additional_payment'] += exp.get('additional_payment', 0)
        else:
            # Неопознанный артикул
            unmatched[art_display] = {
                'name': f"НЕОПОЗНАННЫЙ_АРТИКУЛ: {art_display}",
                'orders': orders,
                'purchases': purchases,
                'cancels': cancels,
                'income': income,
                'payout': payout,
                'logistics': exp.get('logistics', 0),
                'storage': exp.get('storage', 0),
                'penalty': exp.get('penalty', 0),
                'acceptance': exp.get('acceptance', 0),
                'damage_comp': exp.get('damage_comp', 0),
                'return_comp': exp.get('return_comp', 0),
                'additional_payment': exp.get('additional_payment', 0)
            }

    # Сортируем raw_art_data по выкупам (убывание)
    raw_art_data.sort(key=lambda x: x["purchases"], reverse=True)


    return grouped, unmatched, raw_art_data


def create_wb_excel_report(grouped, unmatched, id_to_name, main_ids_ordered, output_path,
                           total_orders, total_purchases, total_cancels, total_income,
                           raw_art_data=None):
    """Создаёт Excel-отчёт по продажам WB (аналогично Ozon)"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Сводный"

    # === 1. Общая сводка ===
    headers1 = ["Показатель", "Значение"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    ws1.append(["Заказы, шт", total_orders])
    ws1.append(["Выкупы, шт", total_purchases])
    ws1.append(["Отмены, шт", total_cancels])
    ws1.append(["Валовая прибыль, руб", total_income])

    avg_profit_per_unit = total_income / total_purchases if total_purchases > 0 else 0
    ws1.append(["Прибыль на 1 ед, руб", avg_profit_per_unit])

    total_shipments = total_purchases + total_cancels
    purchase_percent = (total_purchases / total_shipments * 100) if total_shipments > 0 else 0
    ws1.append(["Процент выкупов", f"{purchase_percent:.2f}%"])

    # === 2. Разделитель ===
    ws1.append([])

    # === 3. ТОП-5 артикулов по выкупам ===
    if raw_art_data and len(raw_art_data) > 0:
        top_5 = raw_art_data[:5]

        ws1.append(["🏆 ТОП-5 артикулов по выкупам"])
        header_cell = ws1.cell(row=ws1.max_row, column=1)
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal="center")

        ws1.append([])

        top_headers = ["Место", "Артикул", "Выкупы, шт", "Прибыль, ₽"]
        ws1.append(top_headers)
        for col in range(1, len(top_headers) + 1):
            cell = ws1.cell(row=ws1.max_row, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for i, item in enumerate(top_5, 1):
            ws1.append([
                i,
                item["art"],
                item["purchases"],
                item["profit"]
            ])

    # === 4. Форматирование листа "Сводный" ===
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws1.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

    # Автоподбор ширины
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column].width = adjusted_width

    # === 5. Лист "Подробный" ===
    ws2 = wb.create_sheet(title="Подробный")
    headers2 = [
        "Наименование",
        "Выкупы, шт",
        "Итого к перечислению, руб",
        "Процент выкупов",
        "К перечислению на 1 ед, руб",
        "Заказы, шт",
        "Отмены, шт"
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    row_index = 2
    for group_id in main_ids_ordered:
        data = grouped.get(group_id, {})
        name = data.get('name', f"Группа {group_id}")
        orders = int(data.get('orders', 0))
        purchases = int(data.get('purchases', 0))
        cancels = int(data.get('cancels', 0))

        # Пересчитываем итого к перечислению с учётом ВСЕХ расходов группы
        income_val = data.get('income', 0)
        logistics = data.get('logistics', 0)
        storage = data.get('storage', 0)
        penalty = data.get('penalty', 0)
        acceptance = data.get('acceptance', 0)
        additional = data.get('additional_payment', 0)

        payout_val = income_val - logistics - storage - penalty - acceptance + additional

        payout_per_unit = payout_val / purchases if purchases > 0 else 0
        total_shipments_row = purchases + cancels
        purchase_percent_val = (purchases / total_shipments_row * 100) if total_shipments_row > 0 else 0

        ws2.append([
            name,
            purchases,
            payout_val,
            f"{purchase_percent_val:.2f}%",
            payout_per_unit,
            orders,
            cancels
        ])

        percent_cell = ws2.cell(row=row_index, column=4)
        if purchase_percent_val <= 50:
            percent_cell.fill = red_fill
        elif 50 < purchase_percent_val <= 60:
            percent_cell.fill = orange_fill
        row_index += 1

    # Неопознанные артикулы
    for art, data in sorted(unmatched.items()):
        name = data['name']
        orders = int(data.get('orders', 0))
        purchases = int(data.get('purchases', 0))
        cancels = int(data.get('cancels', 0))

        # Пересчитываем итого к перечислению с учётом ВСЕХ расходов
        income_val = data.get('income', 0)
        logistics = data.get('logistics', 0)
        storage = data.get('storage', 0)
        penalty = data.get('penalty', 0)
        acceptance = data.get('acceptance', 0)
        additional = data.get('additional_payment', 0)

        payout_val = income_val - logistics - storage - penalty - acceptance + additional

        payout_per_unit = payout_val / purchases if purchases > 0 else 0
        total_shipments_row = purchases + cancels
        purchase_percent_val = (purchases / total_shipments_row * 100) if total_shipments_row > 0 else 0

        ws2.append([
            name,
            purchases,
            payout_val,
            f"{purchase_percent_val:.2f}%",
            payout_per_unit,
            orders,
            cancels
        ])

        percent_cell = ws2.cell(row=row_index, column=4)
        if purchase_percent_val <= 50:
            percent_cell.fill = red_fill
        elif 50 < purchase_percent_val <= 60:
            percent_cell.fill = orange_fill
        row_index += 1

    # === 6. Лист "Исходные артикулы" ===
    ws3 = None
    if raw_art_data:
        ws3 = wb.create_sheet(title="Исходные артикулы")
        headers3 = [
            "Артикул",
            "Выкупы, шт",
            "Итого к перечислению, руб",
            "Процент выкупов",
            "К перечислению на 1 ед, руб",
            "Заказы, шт",
            "Отмены, шт"
        ]
        ws3.append(headers3)
        for cell in ws3[1]:
            cell.font = Font(bold=True)

        row_idx = 2
        for item in raw_art_data:
            art = item["art"]
            purchases = int(item["purchases"])
            payout = item["payout"]
            purchase_percent_item = item["purchase_percent"]
            payout_per_unit = item["payout_per_unit"]
            orders = int(item["orders"])
            cancels = int(item["cancels"])

            ws3.append([
                art,
                purchases,
                payout,
                f"{purchase_percent_item:.2f}%",
                payout_per_unit,
                orders,
                cancels
            ])

            percent_cell = ws3.cell(row=row_idx, column=4)
            if purchase_percent_item <= 50:
                percent_cell.fill = red_fill
            elif 50 < purchase_percent_item <= 60:
                percent_cell.fill = orange_fill
            row_idx += 1

    # === 7. Форматирование остальных листов ===
    worksheets = [ws2]
    if ws3 is not None:
        worksheets.append(ws3)

    for ws in worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(output_path)

def create_wb_excel_report_v2(grouped, unmatched, id_to_name, main_ids_ordered, output_path,
                               total_orders, total_purchases, total_cancels,
                               total_income, total_payout,
                               total_logistics, total_storage, total_penalty, total_acceptance,
                               total_additional,
                               raw_art_data=None):
    """Создаёт Excel-отчёт по продажам WB с детальными расходами"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Сводный"

    # === 1. Общая сводка ===
    headers1 = ["Показатель", "Значение"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    ws1.append(["Заказы, шт", total_orders])
    ws1.append(["Выкупы, шт", total_purchases])
    ws1.append(["Отмены, шт", total_cancels])
    ws1.append(["Итого к перечислению, руб", total_payout])

    avg_payout_per_unit = total_payout / total_purchases if total_purchases > 0 else 0
    ws1.append(["К перечислению на 1 ед, руб", avg_payout_per_unit])

    total_shipments = total_purchases + total_cancels
    purchase_percent = (total_purchases / total_shipments * 100) if total_shipments > 0 else 0
    ws1.append(["Процент выкупов", f"{purchase_percent:.2f}%"])

    ws1.append([])
    ws1.append(["РАСХОДЫ:", ""])
    ws1.append(["Стоимость логистики, руб", total_logistics])
    ws1.append(["Стоимость хранения, руб", total_storage])
    ws1.append(["Штрафы, руб", total_penalty])
    ws1.append(["Операции при приёмке, руб", total_acceptance])
    ws1.append(["Доплаты, руб", total_additional])
    ws1.append([])
    ws1.append(["Прибыль до вычета расходов, руб", total_income])


    # === 2. Разделитель ===
    ws1.append([])

    # === 3. ТОП-5 артикулов по выкупам ===
    if raw_art_data and len(raw_art_data) > 0:
        top_5 = raw_art_data[:5]

        ws1.append(["🏆 ТОП-5 артикулов по выкупам"])
        header_cell = ws1.cell(row=ws1.max_row, column=1)
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal="center")

        ws1.append([])

        top_headers = ["Место", "Артикул", "Выкупы, шт", "К перечислению, ₽"]
        ws1.append(top_headers)
        for col in range(1, len(top_headers) + 1):
            cell = ws1.cell(row=ws1.max_row, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for i, item in enumerate(top_5, 1):
            ws1.append([
                i,
                item["art"],
                int(item["purchases"]),
                item["payout"]
            ])

    # === 4. Форматирование листа "Сводный" ===
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws1.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

    # Автоподбор ширины
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column].width = adjusted_width

    # === 5. Лист "Подробный" ===
    ws2 = wb.create_sheet(title="Подробный")
    headers2 = [
        "Наименование",
        "Выкупы, шт",
        "Итого к перечислению, руб",
        "Процент выкупов",
        "К перечислению на 1 ед, руб",
        "Заказы, шт",
        "Отмены, шт"
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    row_index = 2
    for group_id in main_ids_ordered:
        data = grouped.get(group_id, {})
        name = data.get('name', f"Группа {group_id}")
        orders = int(data.get('orders', 0))
        purchases = int(data.get('purchases', 0))
        cancels = int(data.get('cancels', 0))

        # Пересчитываем итого к перечислению с учётом ВСЕХ расходов группы
        income_val = data.get('income', 0)
        logistics = data.get('logistics', 0)
        storage = data.get('storage', 0)
        penalty = data.get('penalty', 0)
        acceptance = data.get('acceptance', 0)
        additional = data.get('additional_payment', 0)

        payout_val = income_val - logistics - storage - penalty - acceptance + additional

        payout_per_unit = payout_val / purchases if purchases > 0 else 0
        total_shipments_row = purchases + cancels
        purchase_percent_val = (purchases / total_shipments_row * 100) if total_shipments_row > 0 else 0

        ws2.append([
            name,
            purchases,
            payout_val,
            f"{purchase_percent_val:.2f}%",
            payout_per_unit,
            orders,
            cancels
        ])

        percent_cell = ws2.cell(row=row_index, column=4)
        if purchase_percent_val <= 50:
            percent_cell.fill = red_fill
        elif 50 < purchase_percent_val <= 60:
            percent_cell.fill = orange_fill
        row_index += 1

    # Неопознанные артикулы
    for art, data in sorted(unmatched.items()):
        name = data['name']
        orders = int(data.get('orders', 0))
        purchases = int(data.get('purchases', 0))
        cancels = int(data.get('cancels', 0))

        # Пересчитываем итого к перечислению с учётом ВСЕХ расходов
        income_val = data.get('income', 0)
        logistics = data.get('logistics', 0)
        storage = data.get('storage', 0)
        penalty = data.get('penalty', 0)
        acceptance = data.get('acceptance', 0)
        additional = data.get('additional_payment', 0)

        payout_val = income_val - logistics - storage - penalty - acceptance + additional

        payout_per_unit = payout_val / purchases if purchases > 0 else 0
        total_shipments_row = purchases + cancels
        purchase_percent_val = (purchases / total_shipments_row * 100) if total_shipments_row > 0 else 0

        ws2.append([
            name,
            purchases,
            payout_val,
            f"{purchase_percent_val:.2f}%",
            payout_per_unit,
            orders,
            cancels
        ])

        percent_cell = ws2.cell(row=row_index, column=4)
        if purchase_percent_val <= 50:
            percent_cell.fill = red_fill
        elif 50 < purchase_percent_val <= 60:
            percent_cell.fill = orange_fill
        row_index += 1

    # === 6. Лист "Исходные артикулы" ===
    ws3 = None
    if raw_art_data:
        ws3 = wb.create_sheet(title="Исходные артикулы")
        headers3 = [
            "Артикул",
            "Выкупы, шт",
            "Итого к перечислению, руб",
            "Процент выкупов",
            "К перечислению на 1 ед, руб",
            "Заказы, шт",
            "Отмены, шт"
        ]
        ws3.append(headers3)
        for cell in ws3[1]:
            cell.font = Font(bold=True)

        row_idx = 2
        for item in raw_art_data:
            art = item["art"]
            purchases = int(item["purchases"])
            payout = item["payout"]
            purchase_percent_item = item["purchase_percent"]
            payout_per_unit = item["payout_per_unit"]
            orders = int(item["orders"])
            cancels = int(item["cancels"])

            ws3.append([
                art,
                purchases,
                payout,
                f"{purchase_percent_item:.2f}%",
                payout_per_unit,
                orders,
                cancels
            ])

            percent_cell = ws3.cell(row=row_idx, column=4)
            if purchase_percent_item <= 50:
                percent_cell.fill = red_fill
            elif 50 < purchase_percent_item <= 60:
                percent_cell.fill = orange_fill
            row_idx += 1

    # === 7. Форматирование остальных листов ===
    worksheets = [ws2]
    if ws3 is not None:
        worksheets.append(ws3)

    for ws in worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(output_path)

