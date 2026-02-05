# handlers/all_mp_remains_handler.py

import os
import sys
import shutil
import logging
import pandas as pd
import time
import asyncio
from telegram import Update, ReplyKeyboardRemove
from telegram.ext import CallbackContext, ConversationHandler
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from requests.exceptions import Timeout, RequestException

current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)
utils_dir = os.path.join(root_dir, 'utils')

if root_dir not in sys.path:
    sys.path.append(root_dir)
if utils_dir not in sys.path:
    sys.path.append(utils_dir)

logger = logging.getLogger(__name__)

# Диагностика Ozon категорий/атрибутов: включается переменной окружения OZON_CATEGORY_DEBUG=1
OZON_CATEGORY_DEBUG = os.getenv("OZON_CATEGORY_DEBUG", "0").strip() in ("1", "true", "True", "yes", "Y")

from handlers.ozon_remains_handler import OzonAPI
from handlers.wb_remains_handler import WildberriesAPI
from handlers.ozon_remains_handler import clean_offer_id
from handlers.wb_remains_handler import clean_article
from utils.stock_control import resolve_stock_thresholds, apply_fill_to_cells
from utils.ozon_attributes import (
    flatten_description_category_tree,
    build_category_full_paths,
    extract_attribute_values_from_product_attributes,
)

MAX_RETRIES = 3
RETRY_DELAY = 2  # секунды

CACHE_DIR = os.path.join(root_dir, "cache")

# === ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ: СЫРЫЕ ДАННЫЕ ===

async def fetch_ozon_remains_raw(cabinet_id):
    """Полностью копируем логику из handle_cabinet_choice для надежности"""
    ozon = OzonAPI(cabinet_id=cabinet_id)

    os.makedirs(CACHE_DIR, exist_ok=True)
    ozon_cache_dir = os.path.join(CACHE_DIR, "ozon")
    os.makedirs(ozon_cache_dir, exist_ok=True)

    t0_total = time.time()

    def _ozon_post(path: str, payload: dict, timeout: int = 60) -> dict | None:
        """POST в Ozon Seller API через requests, т.к. не все методы обёрнуты в OzonAPI."""
        import requests
        try:
            resp = requests.post(f"{ozon.base_url}{path}", json=payload, headers=ozon.headers, timeout=timeout)
            if resp.status_code != 200:
                logger.warning(f"Ozon кабинет {cabinet_id}: {path} -> {resp.status_code}: {resp.text}")
                return None
            return resp.json() or {}
        except Exception as e:
            logger.warning(f"Ozon кабинет {cabinet_id}: ошибка запроса {path}: {e}")
            return None

    def _cache_read_json(path: str) -> dict | None:
        import json
        try:
            if not os.path.exists(path):
                return None
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None

    def _cache_write_json(path: str, data: dict) -> None:
        import json
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False)  # type: ignore[arg-type]
        except Exception:
            return None

    # Для диагностики: соберём несколько товаров, у которых категория получилась слишком общей
    debug_category_samples: list[dict] = []
    debug_category_samples_limit = 10

    # --- Получение данных (точно как в рабочей функции) ---
    t0 = time.time()
    product_list = ozon.get_product_list(limit=1000)
    logger.info(f"Ozon кабинет {cabinet_id}: get_product_list за {time.time() - t0:.2f}s")
    if not product_list:
        logger.warning(f"Ozon кабинет {cabinet_id}: не удалось получить список товаров")
        return {}, []

    items = product_list.get('result', {}).get('items', [])
    if not items:
        logger.warning(f"Ozon кабинет {cabinet_id}: товары не найдены")
        return {}, []

    offer_ids = []
    for item in items:
        offer_id = clean_offer_id(item.get('offer_id'))
        if offer_id:
            offer_ids.append(offer_id)

    all_skus = []
    offer_id_to_name = {}
    offer_id_to_product_id = {}

    # Категории лучше брать из attributes (/v4/product/info/attributes),
    # т.к. /v3/product/info/list часто не возвращает человекочитаемую категорию.
    offer_id_to_category = {}
    offer_id_to_barcodes: dict[str, list[str]] = {}

    offer_id_to_type_id: dict[str, int] = {}
    offer_id_to_description_category_id: dict[str, int] = {}

    def _extract_ozon_category_from_attributes(info_item: dict) -> str:
        # Приоритет: явная строка категории -> id категории описания
        for key in ("category", "category_name", "category_id"):
            value = info_item.get(key)
            if value is not None and str(value).strip() != "":
                return str(value).strip()

        dcid = info_item.get("description_category_id")
        if dcid is not None and str(dcid).strip() != "":
            return str(dcid).strip()

        return "—"

    from handlers.ozon_remains_handler import chunk_list

    # 1) Получаем sku/name + product_id для последующего запроса attributes
    t0 = time.time()
    for chunk in chunk_list(offer_ids, 1000):
        product_info_response = ozon.get_product_info_list(offer_ids=chunk)
        if not product_info_response:
            continue

        items_in_response = []
        if 'result' in product_info_response and 'items' in product_info_response['result']:
            items_in_response = product_info_response['result']['items']
        elif 'items' in product_info_response:
            items_in_response = product_info_response['items']
        elif isinstance(product_info_response.get('result'), list):
            items_in_response = product_info_response['result']
        else:
            continue

        for item_info in items_in_response:
            offer_id = clean_offer_id(item_info.get('offer_id'))
            sku = item_info.get('sku')
            product_id = item_info.get('id') or item_info.get('product_id')
            name = item_info.get('name', '—')
            if offer_id:
                offer_id_to_name[offer_id] = name
                if product_id is not None:
                    offer_id_to_product_id[offer_id] = product_id

                # Для материалов/категорий
                dcid = item_info.get("description_category_id")
                if dcid is not None:
                    try:
                        offer_id_to_description_category_id[offer_id] = int(dcid)
                    except Exception:
                        pass

                tpid = item_info.get("type_id")
                if tpid is not None:
                    try:
                        offer_id_to_type_id[offer_id] = int(tpid)
                    except Exception:
                        pass

            if offer_id and sku:
                all_skus.append(sku)

    logger.info(f"Ozon кабинет {cabinet_id}: get_product_info_list (chunks={max(1, (len(offer_ids)+999)//1000)}) за {time.time() - t0:.2f}s")

    # === Справочник description_category_id -> category_name ===
    category_name_by_id: dict[int, str] = {}
    category_full_path_by_id: dict[int, str] = {}
    # type_name по type_id (внутри конкретной description_category_id)
    category_type_name_by_pair: dict[tuple[int, int], str] = {}
    try:
        t0 = time.time()
        tree_cache_path = os.path.join(ozon_cache_dir, "description_category_tree_DEFAULT.json")
        tree = _cache_read_json(tree_cache_path)
        if not tree:
            tree = ozon.get_description_category_tree(language="DEFAULT")
            if tree:
                _cache_write_json(tree_cache_path, tree)
        logger.info(f"Ozon кабинет {cabinet_id}: description-category/tree за {time.time() - t0:.2f}s (cache={'hit' if os.path.exists(tree_cache_path) else 'miss'})")
        if tree and tree.get("result"):
            category_name_by_id = flatten_description_category_tree(tree.get("result"))
            category_full_path_by_id = build_category_full_paths(tree.get("result"))

            # Соберём маппинг (description_category_id, type_id) -> type_name
            def _walk(nodes, parent_dcid=None):
                for n in nodes or []:
                    dcid = n.get("description_category_id", parent_dcid)
                    # В некоторых узлах type_id/type_name есть при отсутствии dcid
                    tpid = n.get("type_id")
                    tname = n.get("type_name")
                    if dcid is not None and tpid is not None and tname:
                        try:
                            category_type_name_by_pair[(int(dcid), int(tpid))] = str(tname).strip()
                        except Exception:
                            pass

                    children = n.get("children") or []
                    _walk(children, dcid)

            _walk(tree.get("result"), None)
    except Exception as e:
        logger.warning(f"Ozon кабинет {cabinet_id}: не удалось получить дерево категорий: {e}")

    # 2) Запрашиваем product attributes (и через них заполним категорию/состав)
    offer_id_to_category = dict(offer_id_to_category)  # keep existing
    offer_id_to_composition: dict[str, str] = {}

    # Кеш метаданных атрибутов по (dcid,type_id)
    attributes_meta_cache: dict[tuple[int, int], list[dict]] = {}

    def _resolve_composition_attribute_id(dcid: int, type_id: int) -> int | None:
        """Ищем атрибут состава по МЕТАДАННЫМ атрибутов категории.

        В отличие от "материал" — состав обычно приходит как строка (dictionary_id == 0)
        и содержит проценты (например: "65% полиэстер, 35% хлопок").
        """
        key = (int(dcid), int(type_id))
        if key in attributes_meta_cache:
            attrs = attributes_meta_cache[key]
        else:
            cache_path = os.path.join(ozon_cache_dir, f"attrs_{key[0]}_{key[1]}_DEFAULT.json")
            cached = _cache_read_json(cache_path)
            if cached and isinstance(cached.get("result"), list):
                attrs = cached.get("result")
            else:
                t_req = time.time()
                resp = _ozon_post(
                    "/v1/description-category/attribute",
                    {
                        "description_category_id": key[0],
                        "type_id": key[1],
                        "language": "DEFAULT",
                    },
                    timeout=60,
                )
                logger.info(
                    f"Ozon кабинет {cabinet_id}: /v1/description-category/attribute dcid={key[0]} type_id={key[1]} за {time.time()-t_req:.2f}s"
                )
                attrs = (resp or {}).get("result") or []
                if resp:
                    _cache_write_json(cache_path, resp)
            attributes_meta_cache[key] = attrs

        # кандидаты только строковые и не collection
        candidates = []
        for a in attrs:
            try:
                if str(a.get("type") or "").lower() not in ("string", "text"):
                    continue
                if bool(a.get("is_collection")):
                    continue
            except Exception:
                pass
            candidates.append(a)

        if not candidates:
            return None

        preferred_names = [
            "Состав",
            "Состав ткани",
            "Состав материала",
            "Состав изделия",
            "Состав верха",
        ]

        def _score(a: dict) -> int:
            name = str(a.get("name") or "").strip()
            group = str(a.get("group_name") or "").strip()
            s = 0
            if name in preferred_names:
                s += 100
            if "состав" in name.lower():
                s += 30
            if group in ("Состав", "Состав и уход", "Материалы"):
                s += 10
            return s

        best = sorted(candidates, key=_score, reverse=True)[0]
        try:
            return int(best.get("id"))
        except Exception:
            return None

    # Убираем кеши material_* — они больше не нужны для "Состав".

    # NOTE: для "Состав" мы используем строковый атрибут из /v4/product/info/attributes,
    # поэтому справочник значений (/attribute/values) не нужен.

    # (метод отсутствует в OzonAPI, поэтому вызываем напрямую)
    product_ids = list({pid for pid in offer_id_to_product_id.values() if pid is not None})
    if product_ids:
        import requests
        url = f"{ozon.base_url}/v4/product/info/attributes"
        t0 = time.time()
        # 1000 иногда долго/таймаутит, дробим меньше
        batch_size = 250
        total_chunks = max(1, (len(product_ids) + batch_size - 1) // batch_size)
        for chunk in chunk_list(product_ids, batch_size):
            payload = {
                "filter": {
                    "product_id": chunk,
                    "visibility": "ALL"
                },
                "limit": len(chunk)
            }
            try:
                resp = requests.post(url, json=payload, headers=ozon.headers, timeout=60)
                if resp.status_code != 200:
                    continue
                data = resp.json() or {}
                for info_item in (data.get("result") or []):
                    offer_id = clean_offer_id(info_item.get('offer_id'))
                    if not offer_id:
                        continue

                    # Баркоды
                    try:
                        bcs = info_item.get("barcodes")
                        if isinstance(bcs, list):
                            offer_id_to_barcodes[offer_id] = [str(x).strip() for x in bcs if str(x).strip()]
                        else:
                            bc = info_item.get("barcode")
                            if bc is not None and str(bc).strip():
                                offer_id_to_barcodes[offer_id] = [str(bc).strip()]
                    except Exception:
                        pass

                    # Категория: переводим description_category_id -> name, если возможно
                    cat_name = None
                    dcid = info_item.get("description_category_id")
                    if dcid is not None:
                        try:
                            dcid_i = int(dcid)
                            # Сначала пытаемся взять type_name — это самая узкая категория (например: 'Парео', 'Туники')
                            tpid = info_item.get("type_id")
                            if tpid is not None:
                                try:
                                    type_name = category_type_name_by_pair.get((dcid_i, int(tpid)))
                                except Exception:
                                    type_name = None
                                if type_name:
                                    cat_name = type_name

                            # Если type_name не нашли, берём последний сегмент полного пути
                            if not cat_name:
                                full_path = category_full_path_by_id.get(dcid_i)
                                if full_path:
                                    cat_name = str(full_path).split(" / ")[-1].strip() or None

                            # Если нет полного пути — берём просто имя категории
                            if not cat_name:
                                cat_name = category_name_by_id.get(dcid_i)

                            offer_id_to_description_category_id[offer_id] = dcid_i
                        except Exception:
                            pass

                    if not cat_name:
                        cat_name = _extract_ozon_category_from_attributes(info_item)

                    offer_id_to_category[offer_id] = cat_name or "—"

                    # === Состав (composition) ===
                    dcid_i = offer_id_to_description_category_id.get(offer_id)
                    type_id_i = offer_id_to_type_id.get(offer_id)
                    if dcid_i and type_id_i:
                        comp_attr_id = _resolve_composition_attribute_id(dcid_i, type_id_i)
                        if comp_attr_id:
                            str_vals, dict_ids = extract_attribute_values_from_product_attributes(info_item, comp_attr_id)
                            # Для состава ожидаем строковое значение, dict_ids игнорируем
                            comp_texts = [str(v).strip() for v in (str_vals or []) if str(v).strip()]
                            if comp_texts:
                                # Берём первое непустое
                                offer_id_to_composition[offer_id] = comp_texts[0]

            except Exception as e:
                logger.warning(f"Ozon кабинет {cabinet_id}: не удалось получить категории/состав через attributes: {e}")
                continue
        logger.info(f"Ozon кабинет {cabinet_id}: /v4/product/info/attributes chunks={total_chunks} за {time.time()-t0:.2f}s")

    # Выведем диагностический лог после обработки attributes
    if OZON_CATEGORY_DEBUG and debug_category_samples:
        try:
            logger.warning(
                f"OZON_CATEGORY_DEBUG: кабинет {cabinet_id}: примеры товаров с общей категорией 'Одежда' "
                f"(показываем {len(debug_category_samples)}): {debug_category_samples}"
            )
        except Exception:
            pass

    if not all_skus:
        logger.warning(f"Ozon кабинет {cabinet_id}: не удалось получить SKU")
        return {}, []

    # === АГРЕГАЦИЯ СЫРЫХ ДАННЫХ ПО АРТИКУЛАМ ===
    raw_stock_dict = {}  # Для агрегации сырых данных

    for sku_chunk in chunk_list(all_skus, 100):
        items = ozon.get_analytics_stocks(sku_chunk)
        for item in items:
            offer_id = clean_offer_id(item.get('offer_id'))
            if not offer_id:
                continue

            name = item.get('name', offer_id_to_name.get(offer_id, '—'))
            available = item.get('available_stock_count', 0)
            returning = item.get('return_from_customer_stock_count', 0)
            prepare = item.get('valid_stock_count', 0)

            if offer_id not in raw_stock_dict:
                raw_stock_dict[offer_id] = {
                    'name': name,
                    'category': offer_id_to_category.get(offer_id, '—'),
                    'available': 0,
                    'returning': 0,
                    'prepare': 0
                }

            raw_stock_dict[offer_id]['available'] += available
            raw_stock_dict[offer_id]['returning'] += returning
            raw_stock_dict[offer_id]['prepare'] += prepare

    missing_offer_ids = list(set(offer_ids) - set(raw_stock_dict.keys()))
    if missing_offer_ids:
        for chunk in chunk_list(missing_offer_ids, 100):
            info_response = ozon.get_product_info_list(offer_ids=chunk)
            if not info_response:
                continue

            items_in_response = []
            if 'result' in info_response and 'items' in info_response['result']:
                items_in_response = info_response['result']['items']
            elif 'items' in info_response:
                items_in_response = info_response['items']
            elif isinstance(info_response.get('result'), list):
                items_in_response = info_response['result']
            else:
                continue

            for item in items_in_response:
                offer_id = clean_offer_id(item.get('offer_id'))
                if not offer_id:
                    continue

                stocks = item.get('stocks', {})
                name = item.get('name', '—')
                available = stocks.get('present', 0)
                returning = 0
                prepare = stocks.get('reserved', 0)

                if offer_id not in raw_stock_dict:
                    raw_stock_dict[offer_id] = {
                        'name': name,
                        'category': offer_id_to_category.get(offer_id, '—'),
                        'available': 0,
                        'returning': 0,
                        'prepare': 0
                    }

                raw_stock_dict[offer_id]['available'] += available
                raw_stock_dict[offer_id]['returning'] += returning
                raw_stock_dict[offer_id]['prepare'] += prepare

    # === СОЗДАНИЕ АГРЕГИРОВАННЫХ СЫРЫХ ДАННЫХ ===
    raw_data = []
    for offer_id, data in raw_stock_dict.items():
        total = data['available'] + data['returning'] + data['prepare']
        raw_data.append({
            'Категория': data.get('category', '—'),
            'Артикул': offer_id,
            'Доступно на складах': data['available'],
            'Возвращаются от покупателей': data['returning'],
            'Подготовка к продаже': data['prepare'],
            'Итого на МП': total
        })

    # Преобразуем в формат, который ожидает основная функция (для сводного отчёта)
    result_dict = {}
    for offer_id, data in raw_stock_dict.items():
        result_dict[offer_id] = {
            'avail': data['available'],
            'return': data['returning'],
            'prep': data['prepare']
        }

    return result_dict, raw_data


def normalize_wb_size(value) -> str:
    """Нормализация размера WB для отчётов.

    Если размер отсутствует/0/ONE — возвращаем "единый".
    """
    if value is None:
        return "единый"
    s = str(value).strip()
    if not s:
        return "единый"
    s_up = s.upper()
    if s_up in {"0", "ONE", "ONE SIZE", "ONESIZE", "ЕДИНЫЙ", "ЕДИНЫЙ РАЗМЕР"}:
        return "единый"
    return s


async def fetch_wb_remains_raw(cabinet_id):
    raw_stock_dict: dict[str, dict] = {}
    raw_data: list[dict] = []

    for attempt in range(MAX_RETRIES):
        try:
            wb = WildberriesAPI(cabinet_id=cabinet_id)

            t0 = time.time()
            stocks = wb.get_fbo_stocks_v1()  # синхронный вызов
            logger.info(
                f"WB кабинет {cabinet_id}: get_fbo_stocks_v1 строк={len(stocks or [])} за {time.time() - t0:.2f}s"
            )

            for item in (stocks or []):
                art = clean_article(item.get("supplierArticle"))
                if not art:
                    continue

                category = item.get("subject") or item.get("category") or "—"

                quantity = item.get('quantity', 0) or 0
                in_way_to_client = item.get('inWayToClient', 0) or 0
                in_way_from_client = item.get('inWayFromClient', 0) or 0

                if art not in raw_stock_dict:
                    raw_stock_dict[art] = {
                        'category': str(category).strip() if str(category).strip() else "—",
                        'quantity': 0,
                        'in_way_to_client': 0,
                        'in_way_from_client': 0
                    }

                raw_stock_dict[art]['quantity'] += quantity
                raw_stock_dict[art]['in_way_to_client'] += in_way_to_client
                raw_stock_dict[art]['in_way_from_client'] += in_way_from_client

            for art, data in raw_stock_dict.items():
                total = data['quantity'] + data['in_way_to_client'] + data['in_way_from_client']
                raw_data.append({
                    'Категория': data.get('category', '—'),
                    'Артикул': art,
                    'Доступно на складах': data['quantity'],
                    'Возвращаются от покупателей': data['in_way_from_client'],
                    'В пути до покупателей': data['in_way_to_client'],
                    'Итого на МП': total
                })

            result_dict = {}
            for art, data in raw_stock_dict.items():
                result_dict[art] = {
                    'avail': data['quantity'],
                    'return': data['in_way_from_client'],
                    'inway': data['in_way_to_client']
                }

            logger.info(f"✅ Успешно получены остатки WB кабинет {cabinet_id} (попытка {attempt + 1})")
            return result_dict, raw_data

        except (Timeout, RequestException) as e:
            logger.warning(f"⚠️ Попытка {attempt + 1}/{MAX_RETRIES} не удалась для WB кабинет {cabinet_id}: {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)
            else:
                logger.error(f"❌ Ошибка получения остатков WB кабинет {cabinet_id} после {MAX_RETRIES} попыток")
                return {}, []
        except Exception as e:
            logger.error(f"❌ Ошибка получения остатков WB кабинет {cabinet_id}: {e}", exc_info=True)
            return {}, []

# === ФУНКЦИЯ НОРМАЛИЗАЦИИ ===

def normalize_art(art_str):
    """Нормализует строку: приводит к нижнему регистру, удаляет лишние пробелы, очищает от невидимых символов"""
    if not art_str:
        return ""
    s = str(art_str)
    s = ''.join(c for c in s if c.isprintable())
    s = s.strip().lower()
    return s


def _get_rows_to_color(df, art_column, cabinet_arts_set):
    rows = []
    for idx, art in enumerate(df[art_column], start=3):
        if normalize_art(art) in cabinet_arts_set:
            rows.append(idx)
    return rows


# === ФУНКЦИИ ДЛЯ СОЗДАНИЯ EXCEL ЛИСТОВ ===

def _write_sheet(ws, df, headers, has_name=False):
    """Вспомогательная функция для записи одного листа с форматированием"""
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Объединение ячеек в заголовке
    ws.merge_cells('A1:A2')
    if has_name:
        ws.merge_cells('B1:B2')

    data_start_row = 3
    sum_row = 2

    # Данные
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), data_start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border

    # Суммы
    num_rows = len(df)
    if num_rows > 0:
        start_col_index = 3 if has_name else 2
        for col in range(start_col_index, len(headers) + 1):
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_start_row + num_rows - 1})"
            cell = ws.cell(row=sum_row, column=col, value=formula)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border

    # Автоподбор ширины
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


# === ОСНОВНОЙ ОБРАБОТЧИК ===

async def start_all_mp_remains(update: Update, context: CallbackContext) -> int:
    # Сохраняем message_id для последующего удаления
    context.user_data['all_mp_status_messages'] = []

    status_msg = await update.message.reply_text("⏳ Начинаю выгрузку остатков со всех маркетплейсов...",
                                                 reply_markup=ReplyKeyboardRemove())
    context.user_data['all_mp_status_messages'].append(status_msg.message_id)

    await generate_all_mp_report(update, context)
    return ConversationHandler.END


async def generate_all_mp_report(update: Update, context: CallbackContext):
    try:
        # Список для хранения ID сообщений о статусе
        status_message_ids = context.user_data.get('all_mp_status_messages', [])

        # === 1. Получаем сырые данные ===
        status_msg = await update.message.reply_text("📊 Запрашиваю остатки Ozon Кабинет 1 (Nimba)...")
        status_message_ids.append(status_msg.message_id)
        ozon1_raw_dict, ozon1_raw_data = await fetch_ozon_remains_raw(1)

        status_msg = await update.message.reply_text("📊 Запрашиваю остатки Ozon Кабинет 2 (Galioni)...")
        status_message_ids.append(status_msg.message_id)
        ozon2_raw_dict, ozon2_raw_data = await fetch_ozon_remains_raw(2)

        status_msg = await update.message.reply_text("📊 Запрашиваю остатки Ozon Кабинет 3 (AGNIA)...")
        status_message_ids.append(status_msg.message_id)
        ozon3_raw_dict, ozon3_raw_data = await fetch_ozon_remains_raw(3)

        status_msg = await update.message.reply_text("📊 Запрашиваю остатки Wildberries Кабинет 1 (Nimba)...")
        status_message_ids.append(status_msg.message_id)
        wb1_raw_dict, wb1_raw_data = await fetch_wb_remains_raw(1)

        status_msg = await update.message.reply_text("📊 Запрашиваю остатки Wildberries Кабинет 2 (Galioni)...")
        status_message_ids.append(status_msg.message_id)
        wb2_raw_dict, wb2_raw_data = await fetch_wb_remains_raw(2)

        status_msg = await update.message.reply_text("📊 Запрашиваю остатки Wildberries Кабинет 3 (AGNIA)...")
        status_message_ids.append(status_msg.message_id)
        wb3_raw_dict, wb3_raw_data = await fetch_wb_remains_raw(3)

        # === 2. Загружаем маппинги ===
        from utils.template_loader import get_cabinet_articles_by_template_id

        ozon1_id_to_name, ozon1_id_to_arts = get_cabinet_articles_by_template_id("Отдельно Озон Nimba")
        ozon2_id_to_name, ozon2_id_to_arts = get_cabinet_articles_by_template_id("Отдельно Озон Galioni")
        ozon3_id_to_name, ozon3_id_to_arts = get_cabinet_articles_by_template_id("Отдельно Озон AGNIA")
        wb1_id_to_name, wb1_id_to_arts = get_cabinet_articles_by_template_id("Отдельно ВБ Nimba")
        wb2_id_to_name, wb2_id_to_arts = get_cabinet_articles_by_template_id("Отдельно ВБ Galioni")
        wb3_id_to_name, wb3_id_to_arts = get_cabinet_articles_by_template_id("Отдельно ВБ AGNIA")

        ozon1_linked_ids = set(ozon1_id_to_arts.keys())
        ozon2_linked_ids = set(ozon2_id_to_arts.keys())
        ozon3_linked_ids = set(ozon3_id_to_arts.keys())
        wb1_linked_ids = set(wb1_id_to_arts.keys())
        wb2_linked_ids = set(wb2_id_to_arts.keys())
        wb3_linked_ids = set(wb3_id_to_arts.keys())

        ozon1_arts_set = {normalize_art(art) for arts in ozon1_id_to_arts.values() for art in arts}
        ozon2_arts_set = {normalize_art(art) for arts in ozon2_id_to_arts.values() for art in arts}
        ozon3_arts_set = {normalize_art(art) for arts in ozon3_id_to_arts.values() for art in arts}
        wb1_arts_set = {normalize_art(art) for arts in wb1_id_to_arts.values() for art in arts}
        wb2_arts_set = {normalize_art(art) for arts in wb2_id_to_arts.values() for art in arts}
        wb3_arts_set = {normalize_art(art) for arts in wb3_id_to_arts.values() for art in arts}

        # === 3. Построим обратные маппинги ===
        def build_reverse(id_to_arts):
            rev = {}
            for tid, arts in id_to_arts.items():
                for art in arts:
                    clean_art = normalize_art(art)
                    rev[clean_art] = tid
            return rev

        ozon1_rev = build_reverse(ozon1_id_to_arts)
        ozon2_rev = build_reverse(ozon2_id_to_arts)
        ozon3_rev = build_reverse(ozon3_id_to_arts)
        wb1_rev = build_reverse(wb1_id_to_arts)
        wb2_rev = build_reverse(wb2_id_to_arts)
        wb3_rev = build_reverse(wb3_id_to_arts)

        # === 4. Агрегация данных ===
        ozon1_agg = {}
        for art, data in ozon1_raw_dict.items():
            clean_art = normalize_art(art)
            tid = ozon1_rev.get(clean_art)
            if tid is not None:
                if tid not in ozon1_agg:
                    ozon1_agg[tid] = {'avail': 0, 'return': 0, 'prep': 0}
                ozon1_agg[tid]['avail'] += data['avail']
                ozon1_agg[tid]['return'] += data['return']
                ozon1_agg[tid]['prep'] += data['prep']

        ozon2_agg = {}
        for art, data in ozon2_raw_dict.items():
            clean_art = normalize_art(art)
            tid = ozon2_rev.get(clean_art)
            if tid is not None:
                if tid not in ozon2_agg:
                    ozon2_agg[tid] = {'avail': 0, 'return': 0, 'prep': 0}
                ozon2_agg[tid]['avail'] += data['avail']
                ozon2_agg[tid]['return'] += data['return']
                ozon2_agg[tid]['prep'] += data['prep']

        ozon3_agg = {}
        for art, data in ozon3_raw_dict.items():
            clean_art = normalize_art(art)
            tid = ozon3_rev.get(clean_art)
            if tid is not None:
                if tid not in ozon3_agg:
                    ozon3_agg[tid] = {'avail': 0, 'return': 0, 'prep': 0}
                ozon3_agg[tid]['avail'] += data['avail']
                ozon3_agg[tid]['return'] += data['return']
                ozon3_agg[tid]['prep'] += data['prep']

        wb1_agg = {}
        for art, data in wb1_raw_dict.items():
            clean_art = normalize_art(art)
            tid = wb1_rev.get(clean_art)
            if tid is not None:
                if tid not in wb1_agg:
                    wb1_agg[tid] = {'avail': 0, 'return': 0, 'inway': 0}
                wb1_agg[tid]['avail'] += data['avail']
                wb1_agg[tid]['return'] += data['return']
                wb1_agg[tid]['inway'] += data['inway']

        wb2_agg = {}
        for art, data in wb2_raw_dict.items():
            clean_art = normalize_art(art)
            tid = wb2_rev.get(clean_art)
            if tid is not None:
                if tid not in wb2_agg:
                    wb2_agg[tid] = {'avail': 0, 'return': 0, 'inway': 0}
                wb2_agg[tid]['avail'] += data['avail']
                wb2_agg[tid]['return'] += data['return']
                wb2_agg[tid]['inway'] += data['inway']

        wb3_agg = {}
        for art, data in wb3_raw_dict.items():
            clean_art = normalize_art(art)
            tid = wb3_rev.get(clean_art)
            if tid is not None:
                if tid not in wb3_agg:
                    wb3_agg[tid] = {'avail': 0, 'return': 0, 'inway': 0}
                wb3_agg[tid]['avail'] += data['avail']
                wb3_agg[tid]['return'] += data['return']
                wb3_agg[tid]['inway'] += data['inway']

        # === 5. РАБОТА С ШАБЛОНОМ - ПОЛНОЕ КОПИРОВАНИЕ ===
        template_report_path = os.path.join(root_dir, "Шаблон выгрузки остатков всех МП.xlsx")
        if not os.path.exists(template_report_path):
            raise FileNotFoundError("Файл 'Шаблон выгрузки остатков всех МП.xlsx' не найден!")

        report_copy = os.path.join(root_dir, "Остатки_все_МП_отчёт.xlsx")

        # ПОЛНОСТЬЮ КОПИРУЕМ ФАЙЛ ШАБЛОНА
        shutil.copy(template_report_path, report_copy)

        # Загружаем скопированный файл
        wb = load_workbook(report_copy)
        ws = wb.active  # Это уже готовый лист "Остатки на МП" с правильным оформлением

        thresholds = resolve_stock_thresholds(context, update.effective_chat.id)

        # Заполняем данными (только значения, оформление остаётся как в шаблоне)
        row = 7
        while True:
            cell_value = ws[f"A{row}"].value
            if not cell_value or str(cell_value).strip().upper() == "ИТОГО":
                break

            art_name = str(cell_value).strip()

            # Ищем template_id по имени во ВСЕХ кабинетах
            template_id = None
            all_id_to_name = [
                ozon1_id_to_name,
                ozon2_id_to_name,
                ozon3_id_to_name,
                wb1_id_to_name,
                wb2_id_to_name,
                wb3_id_to_name
            ]

            for id_to_name in all_id_to_name:
                for tid, name in id_to_name.items():
                    if str(name).strip().lower() == art_name.lower():
                        template_id = tid
                        break
                if template_id is not None:
                    break

            if template_id is not None:
                # --- Ozon 1 ---
                o1 = ozon1_agg.get(template_id, {'avail': 0, 'return': 0, 'prep': 0})
                ws[f"B{row}"] = o1['avail']
                ws[f"C{row}"] = o1['return']
                ws[f"D{row}"] = o1['prep']
                ws[f"E{row}"] = o1['avail'] + o1['return'] + o1['prep']
                if thresholds and template_id in ozon1_linked_ids:
                    apply_fill_to_cells(ws, [row], [5], thresholds)

                # --- Ozon 2 ---
                o2 = ozon2_agg.get(template_id, {'avail': 0, 'return': 0, 'prep': 0})
                ws[f"G{row}"] = o2['avail']
                ws[f"H{row}"] = o2['return']
                ws[f"I{row}"] = o2['prep']
                ws[f"J{row}"] = o2['avail'] + o2['return'] + o2['prep']
                if thresholds and template_id in ozon2_linked_ids:
                    apply_fill_to_cells(ws, [row], [10], thresholds)

                # --- Ozon 3 ---
                o3 = ozon3_agg.get(template_id, {'avail': 0, 'return': 0, 'prep': 0})
                ws[f"L{row}"] = o3['avail']
                ws[f"M{row}"] = o3['return']
                ws[f"N{row}"] = o3['prep']
                ws[f"O{row}"] = o3['avail'] + o3['return'] + o3['prep']
                if thresholds and template_id in ozon3_linked_ids:
                    apply_fill_to_cells(ws, [row], [15], thresholds)

                # --- WB 1 ---
                w1 = wb1_agg.get(template_id, {'avail': 0, 'return': 0, 'inway': 0})
                ws[f"Q{row}"] = w1['avail']
                ws[f"R{row}"] = w1['return']
                ws[f"S{row}"] = w1['inway']
                ws[f"T{row}"] = w1['avail'] + w1['return'] + w1['inway']
                if thresholds and template_id in wb1_linked_ids:
                    apply_fill_to_cells(ws, [row], [20], thresholds)

                # --- WB 2 ---
                w2 = wb2_agg.get(template_id, {'avail': 0, 'return': 0, 'inway': 0})
                ws[f"V{row}"] = w2['avail']
                ws[f"W{row}"] = w2['return']
                ws[f"X{row}"] = w2['inway']
                ws[f"Y{row}"] = w2['avail'] + w2['return'] + w2['inway']
                if thresholds and template_id in wb2_linked_ids:
                    apply_fill_to_cells(ws, [row], [25], thresholds)

                # --- WB 3 ---
                w3 = wb3_agg.get(template_id, {'avail': 0, 'return': 0, 'inway': 0})
                ws[f"AA{row}"] = w3['avail']
                ws[f"AB{row}"] = w3['return']
                ws[f"AC{row}"] = w3['inway']
                ws[f"AD{row}"] = w3['avail'] + w3['return'] + w3['inway']
                if thresholds and template_id in wb3_linked_ids:
                    apply_fill_to_cells(ws, [row], [30], thresholds)

            row += 1

        # === ДОПОЛНИТЕЛЬНЫЕ ЛИСТЫ ===

        # Ozon1 исходные артикулы
        if ozon1_raw_data:
            df_ozon1_raw = pd.DataFrame(ozon1_raw_data).sort_values(by='Категория',
                                                                    key=lambda x: x.str.lower()).reset_index(drop=True)
            headers_ozon1 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей",
                              "Подготовка к продаже", "Итого на МП"]
            ws_ozon1 = wb.create_sheet(title="Ozon1 исходные артикулы")
            _write_sheet(ws_ozon1, df_ozon1_raw, headers_ozon1, has_name=True)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_ozon1_raw, "Артикул", ozon1_arts_set)
                total_col = headers_ozon1.index("Итого на МП") + 1
                apply_fill_to_cells(ws_ozon1, rows_to_color, [total_col], thresholds)
        else:
            ws_ozon1 = wb.create_sheet(title="Ozon1 исходные артикулы")
            ws_ozon1.append(["Нет данных"])

        # Ozon2 исходные артикулы
        if ozon2_raw_data:
            df_ozon2_raw = pd.DataFrame(ozon2_raw_data).sort_values(by='Категория',
                                                                    key=lambda x: x.str.lower()).reset_index(drop=True)
            headers_ozon2 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей",
                              "Подготовка к продаже", "Итого на МП"]
            ws_ozon2 = wb.create_sheet(title="Ozon2 исходные артикулы")
            _write_sheet(ws_ozon2, df_ozon2_raw, headers_ozon2, has_name=True)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_ozon2_raw, "Артикул", ozon2_arts_set)
                total_col = headers_ozon2.index("Итого на МП") + 1
                apply_fill_to_cells(ws_ozon2, rows_to_color, [total_col], thresholds)
        else:
            ws_ozon2 = wb.create_sheet(title="Ozon2 исходные артикулы")
            ws_ozon2.append(["Нет данных"])

        # Ozon3 исходные артикулы
        if ozon3_raw_data:
            df_ozon3_raw = pd.DataFrame(ozon3_raw_data).sort_values(by='Категория',
                                                                    key=lambda x: x.str.lower()).reset_index(drop=True)
            headers_ozon3 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей",
                              "Подготовка к продаже", "Итого на МП"]
            ws_ozon3 = wb.create_sheet(title="Ozon AGNIA исходные артикулы")
            _write_sheet(ws_ozon3, df_ozon3_raw, headers_ozon3, has_name=True)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_ozon3_raw, "Артикул", ozon3_arts_set)
                total_col = headers_ozon3.index("Итого на МП") + 1
                apply_fill_to_cells(ws_ozon3, rows_to_color, [total_col], thresholds)
        else:
            ws_ozon3 = wb.create_sheet(title="Ozon AGNIA исходные артикулы")
            ws_ozon3.append(["Нет данных"])

        # WB1 исходные артикулы
        if wb1_raw_data:
            df_wb1_raw = pd.DataFrame(wb1_raw_data).sort_values(by='Артикул').reset_index(drop=True)
            headers_wb1 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей", "В пути до покупателей",
                            "Итого на МП"]
            ws_wb1 = wb.create_sheet(title="WB1 исходные артикулы")
            _write_sheet(ws_wb1, df_wb1_raw, headers_wb1, has_name=False)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_wb1_raw, "Артикул", wb1_arts_set)
                total_col = headers_wb1.index("Итого на МП") + 1
                apply_fill_to_cells(ws_wb1, rows_to_color, [total_col], thresholds)
        else:
            ws_wb1 = wb.create_sheet(title="WB1 исходные артикулы")
            ws_wb1.append(["Нет данных"])

        # WB2 исходные артикулы
        if wb2_raw_data:
            df_wb2_raw = pd.DataFrame(wb2_raw_data).sort_values(by='Артикул').reset_index(drop=True)
            headers_wb2 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей", "В пути до покупателей",
                            "Итого на МП"]
            ws_wb2 = wb.create_sheet(title="WB2 исходные артикулы")
            _write_sheet(ws_wb2, df_wb2_raw, headers_wb2, has_name=False)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_wb2_raw, "Артикул", wb2_arts_set)
                total_col = headers_wb2.index("Итого на МП") + 1
                apply_fill_to_cells(ws_wb2, rows_to_color, [total_col], thresholds)
        else:
            ws_wb2 = wb.create_sheet(title="WB2 исходные артикулы")
            ws_wb2.append(["Нет данных"])

        # WB3 исходные артикулы
        if wb3_raw_data:
            df_wb3_raw = pd.DataFrame(wb3_raw_data).sort_values(by='Артикул').reset_index(drop=True)
            headers_wb3 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей", "В пути до покупателей",
                            "Итого на МП"]
            ws_wb3 = wb.create_sheet(title="WB AGNIA исходные артикулы")
            _write_sheet(ws_wb3, df_wb3_raw, headers_wb3, has_name=False)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_wb3_raw, "Артикул", wb3_arts_set)
                total_col = headers_wb3.index("Итого на МП") + 1
                apply_fill_to_cells(ws_wb3, rows_to_color, [total_col], thresholds)
        else:
            ws_wb3 = wb.create_sheet(title="WB AGNIA исходные артикулы")
            ws_wb3.append(["Нет данных"])

        wb.save(report_copy)

        # === РАСЧЁТ СВОДНЫХ ДАННЫХ ПО ВСЕМ КАБИНЕТАМ ===

        # Ozon 1
        ozon1_total_avail = sum(data['avail'] for data in ozon1_raw_dict.values())
        ozon1_total_return = sum(data['return'] for data in ozon1_raw_dict.values())
        ozon1_total_prep = sum(data['prep'] for data in ozon1_raw_dict.values())
        ozon1_total_mp = ozon1_total_avail + ozon1_total_return + ozon1_total_prep

        # Ozon 2
        ozon2_total_avail = sum(data['avail'] for data in ozon2_raw_dict.values())
        ozon2_total_return = sum(data['return'] for data in ozon2_raw_dict.values())
        ozon2_total_prep = sum(data['prep'] for data in ozon2_raw_dict.values())
        ozon2_total_mp = ozon2_total_avail + ozon2_total_return + ozon2_total_prep

        # Ozon 3
        ozon3_total_avail = sum(data['avail'] for data in ozon3_raw_dict.values())
        ozon3_total_return = sum(data['return'] for data in ozon3_raw_dict.values())
        ozon3_total_prep = sum(data['prep'] for data in ozon3_raw_dict.values())
        ozon3_total_mp = ozon3_total_avail + ozon3_total_return + ozon3_total_prep

        # WB 1
        wb1_total_avail = sum(data['avail'] for data in wb1_raw_dict.values())
        wb1_total_return = sum(data['return'] for data in wb1_raw_dict.values())
        wb1_total_inway = sum(data['inway'] for data in wb1_raw_dict.values())
        wb1_total_mp = wb1_total_avail + wb1_total_return + wb1_total_inway

        # WB 2
        wb2_total_avail = sum(data['avail'] for data in wb2_raw_dict.values())
        wb2_total_return = sum(data['return'] for data in wb2_raw_dict.values())
        wb2_total_inway = sum(data['inway'] for data in wb2_raw_dict.values())
        wb2_total_mp = wb2_total_avail + wb2_total_return + wb2_total_inway

        # WB 3
        wb3_total_avail = sum(data['avail'] for data in wb3_raw_dict.values())
        wb3_total_return = sum(data['return'] for data in wb3_raw_dict.values())
        wb3_total_inway = sum(data['inway'] for data in wb3_raw_dict.values())
        wb3_total_mp = wb3_total_avail + wb3_total_return + wb3_total_inway

        # Общая сумма по всем маркетплейсам
        total_all_mp = (
            ozon1_total_mp + ozon2_total_mp + ozon3_total_mp +
            wb1_total_mp + wb2_total_mp + wb3_total_mp
        )

        def fmt(x):
            return f"{x:,}".replace(",", " ")

        # === ФОРМИРОВАНИЕ КРАСИВОГО СООБЩЕНИЯ ===
        summary_text = (
            "📊 <b>Сводка по остаткам на всех маркетплейсах</b>\n\n"

            "🏪 <b>Ozon Кабинет 1 (Nimba)</b>\n"
            f"   📦 Доступно на складах: {fmt(ozon1_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(ozon1_total_return)} шт\n"
            f"   🔄 Подготовка к продаже: {fmt(ozon1_total_prep)} шт\n"
            f"   ✅ Итого на МП: {fmt(ozon1_total_mp)} шт\n\n"

            "🏬 <b>Ozon Кабинет 2 (Galioni)</b>\n"
            f"   📦 Доступно на складах: {fmt(ozon2_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(ozon2_total_return)} шт\n"
            f"   🔄 Подготовка к продаже: {fmt(ozon2_total_prep)} шт\n"
            f"   ✅ Итого на МП: {fmt(ozon2_total_mp)} шт\n\n"

            "🏢 <b>Ozon Кабинет 3 (AGNIA)</b>\n"
            f"   📦 Доступно на складах: {fmt(ozon3_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(ozon3_total_return)} шт\n"
            f"   🔄 Подготовка к продаже: {fmt(ozon3_total_prep)} шт\n"
            f"   ✅ Итого на МП: {fmt(ozon3_total_mp)} шт\n\n"

            "🏪 <b>Wildberries Кабинет 1 (Nimba)</b>\n"
            f"   📦 Доступно на складах: {fmt(wb1_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(wb1_total_return)} шт\n"
            f"   🚚 В пути до покупателей: {fmt(wb1_total_mp - wb1_total_avail - wb1_total_return)} шт\n"
            f"   ✅ Итого на МП: {fmt(wb1_total_mp)} шт\n\n"

            "🏬 <b>Wildberries Кабинет 2 (Galioni)</b>\n"
            f"   📦 Доступно на складах: {fmt(wb2_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(wb2_total_return)} шт\n"
            f"   🚚 В пути до покупателей: {fmt(wb2_total_mp - wb2_total_avail - wb2_total_return)} шт\n"
            f"   ✅ Итого на МП: {fmt(wb2_total_mp)} шт\n\n"

            "🏢 <b>Wildberries Кабинет 3 (AGNIA)</b>\n"
            f"   📦 Доступно на складах: {fmt(wb3_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(wb3_total_return)} шт\n"
            f"   🚚 В пути до покупателей: {fmt(wb3_total_mp - wb3_total_avail - wb3_total_return)} шт\n"
            f"   ✅ Итого на МП: {fmt(wb3_total_mp)} шт\n\n"

            f"🔹 <b>ВСЕГО на всех маркетплейсах:</b> {fmt(total_all_mp)} шт"
        )

        # === Удаляем все промежуточные сообщения ===
        for msg_id in status_message_ids:
            try:
                await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=msg_id)
            except Exception:
                pass

        # === Отправляем финальный отчёт ===
        await update.message.reply_document(
            document=open(report_copy, 'rb'),
            caption="📊 Объединённый отчёт по остаткам на всех маркетплейсах\n\n"
                    "📄 Листы:\n"
                    "• Остатки на МП — сводный отчёт\n"
                    "• Ozon1/Ozon2 исходные артикулы — сырые данные Ozon\n"
                    "• WB1/WB2 исходные артикулы — сырые данные Wildberries",
            reply_markup=ReplyKeyboardRemove()
        )

        # === Отправляем сводку текстом ===
        await update.message.reply_text(summary_text, parse_mode="HTML")

        if os.path.exists(report_copy):
            os.remove(report_copy)

    except Exception as e:
        # Удаляем промежуточные сообщения даже при ошибке
        status_message_ids = context.user_data.get('all_mp_status_messages', [])
        for msg_id in status_message_ids:
            try:
                await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=msg_id)
            except Exception:
                pass

        logger.error(f"Ошибка в объединённом отчёте: {e}", exc_info=True)
        await update.message.reply_text(f"❌ Ошибка: {str(e)}", reply_markup=ReplyKeyboardRemove())


# ======================
# Автоматическая отправка отчёта по всем маркетплейсам (для job_queue)
# ======================

def get_frequency_label_for_report(config: dict) -> str:
    """Возвращает человекочитаемую метку частоты для заголовка автоотчёта."""
    schedule = config.get('schedule', {})
    sched_type = schedule.get('type')

    if sched_type == 'interval_hours':
        hours = schedule.get('hours', 1)
        if hours == 24:
            return "Ежедневный"
        elif hours == 1:
            return "Почасовой"
        else:
            return f"Каждые {hours} ч"

    elif sched_type == 'interval_days':
        days = schedule.get('days', 1)
        if days == 1:
            return "Ежедневный"
        elif days == 7 and 'day_of_week' in schedule:
            return "Еженедельный"
        else:
            return f"Каждые {days} дн"

    return "Авто"


async def send_all_mp_remains_automatic(context: CallbackContext):
    """Автоматическая отправка объединённого отчёта по остаткам на всех маркетплейсах"""
    chat_id = context.job.data.get('chat_id')
    if not chat_id:
        logger.error("Автоматический отчёт: chat_id не указан в job.data")
        return

    # === ЗАГРУЖАЕМ КОНФИГ АВТООТЧЁТА ДЛЯ ОПРЕДЕЛЕНИЯ ЧАСТОТЫ ===
    from utils.auto_report_manager import load_auto_reports
    reports = load_auto_reports()
    config = reports.get(str(chat_id), {})
    frequency_label = get_frequency_label_for_report(config)

    try:
        # === 1. Получаем сырые данные ===
        ozon1_raw_dict, ozon1_raw_data = await fetch_ozon_remains_raw(1)
        ozon2_raw_dict, ozon2_raw_data = await fetch_ozon_remains_raw(2)
        ozon3_raw_dict, ozon3_raw_data = await fetch_ozon_remains_raw(3)
        wb1_raw_dict, wb1_raw_data = await fetch_wb_remains_raw(1)
        wb2_raw_dict, wb2_raw_data = await fetch_wb_remains_raw(2)
        wb3_raw_dict, wb3_raw_data = await fetch_wb_remains_raw(3)

        # === 2–5. (всё остальное без изменений — копируем твой существующий код) ===
        from utils.template_loader import get_cabinet_articles_by_template_id

        ozon1_id_to_name, ozon1_id_to_arts = get_cabinet_articles_by_template_id("Отдельно Озон Nimba")
        ozon2_id_to_name, ozon2_id_to_arts = get_cabinet_articles_by_template_id("Отдельно Озон Galioni")
        ozon3_id_to_name, ozon3_id_to_arts = get_cabinet_articles_by_template_id("Отдельно Озон AGNIA")
        wb1_id_to_name, wb1_id_to_arts = get_cabinet_articles_by_template_id("Отдельно ВБ Nimba")
        wb2_id_to_name, wb2_id_to_arts = get_cabinet_articles_by_template_id("Отдельно ВБ Galioni")
        wb3_id_to_name, wb3_id_to_arts = get_cabinet_articles_by_template_id("Отдельно ВБ AGNIA")

        ozon1_linked_ids = set(ozon1_id_to_arts.keys())
        ozon2_linked_ids = set(ozon2_id_to_arts.keys())
        ozon3_linked_ids = set(ozon3_id_to_arts.keys())
        wb1_linked_ids = set(wb1_id_to_arts.keys())
        wb2_linked_ids = set(wb2_id_to_arts.keys())
        wb3_linked_ids = set(wb3_id_to_arts.keys())

        ozon1_arts_set = {normalize_art(art) for arts in ozon1_id_to_arts.values() for art in arts}
        ozon2_arts_set = {normalize_art(art) for arts in ozon2_id_to_arts.values() for art in arts}
        ozon3_arts_set = {normalize_art(art) for arts in ozon3_id_to_arts.values() for art in arts}
        wb1_arts_set = {normalize_art(art) for arts in wb1_id_to_arts.values() for art in arts}
        wb2_arts_set = {normalize_art(art) for arts in wb2_id_to_arts.values() for art in arts}
        wb3_arts_set = {normalize_art(art) for arts in wb3_id_to_arts.values() for art in arts}

        # === 3. Построим обратные маппинги ===
        def build_reverse(id_to_arts):
            rev = {}
            for tid, arts in id_to_arts.items():
                for art in arts:
                    clean_art = normalize_art(art)
                    rev[clean_art] = tid
            return rev

        ozon1_rev = build_reverse(ozon1_id_to_arts)
        ozon2_rev = build_reverse(ozon2_id_to_arts)
        ozon3_rev = build_reverse(ozon3_id_to_arts)
        wb1_rev = build_reverse(wb1_id_to_arts)
        wb2_rev = build_reverse(wb2_id_to_arts)
        wb3_rev = build_reverse(wb3_id_to_arts)

        # === 4. Агрегация данных ===
        ozon1_agg = {}
        for art, data in ozon1_raw_dict.items():
            clean_art = normalize_art(art)
            tid = ozon1_rev.get(clean_art)
            if tid is not None:
                if tid not in ozon1_agg:
                    ozon1_agg[tid] = {'avail': 0, 'return': 0, 'prep': 0}
                ozon1_agg[tid]['avail'] += data['avail']
                ozon1_agg[tid]['return'] += data['return']
                ozon1_agg[tid]['prep'] += data['prep']

        ozon2_agg = {}
        for art, data in ozon2_raw_dict.items():
            clean_art = normalize_art(art)
            tid = ozon2_rev.get(clean_art)
            if tid is not None:
                if tid not in ozon2_agg:
                    ozon2_agg[tid] = {'avail': 0, 'return': 0, 'prep': 0}
                ozon2_agg[tid]['avail'] += data['avail']
                ozon2_agg[tid]['return'] += data['return']
                ozon2_agg[tid]['prep'] += data['prep']

        ozon3_agg = {}
        for art, data in ozon3_raw_dict.items():
            clean_art = normalize_art(art)
            tid = ozon3_rev.get(clean_art)
            if tid is not None:
                if tid not in ozon3_agg:
                    ozon3_agg[tid] = {'avail': 0, 'return': 0, 'prep': 0}
                ozon3_agg[tid]['avail'] += data['avail']
                ozon3_agg[tid]['return'] += data['return']
                ozon3_agg[tid]['prep'] += data['prep']

        wb1_agg = {}
        for art, data in wb1_raw_dict.items():
            clean_art = normalize_art(art)
            tid = wb1_rev.get(clean_art)
            if tid is not None:
                if tid not in wb1_agg:
                    wb1_agg[tid] = {'avail': 0, 'return': 0, 'inway': 0}
                wb1_agg[tid]['avail'] += data['avail']
                wb1_agg[tid]['return'] += data['return']
                wb1_agg[tid]['inway'] += data['inway']

        wb2_agg = {}
        for art, data in wb2_raw_dict.items():
            clean_art = normalize_art(art)
            tid = wb2_rev.get(clean_art)
            if tid is not None:
                if tid not in wb2_agg:
                    wb2_agg[tid] = {'avail': 0, 'return': 0, 'inway': 0}
                wb2_agg[tid]['avail'] += data['avail']
                wb2_agg[tid]['return'] += data['return']
                wb2_agg[tid]['inway'] += data['inway']

        wb3_agg = {}
        for art, data in wb3_raw_dict.items():
            clean_art = normalize_art(art)
            tid = wb3_rev.get(clean_art)
            if tid is not None:
                if tid not in wb3_agg:
                    wb3_agg[tid] = {'avail': 0, 'return': 0, 'inway': 0}
                wb3_agg[tid]['avail'] += data['avail']
                wb3_agg[tid]['return'] += data['return']
                wb3_agg[tid]['inway'] += data['inway']

        # === 5. РАБОТА С ШАБЛОНОМ - ПОЛНОЕ КОПИРОВАНИЕ ===
        template_report_path = os.path.join(root_dir, "Шаблон выгрузки остатков всех МП.xlsx")
        if not os.path.exists(template_report_path):
            raise FileNotFoundError("Файл 'Шаблон выгрузки остатков всех МП.xlsx' не найден!")

        report_copy = os.path.join(root_dir, "Остатки_все_МП_отчёт.xlsx")

        # ПОЛНОСТЬЮ КОПИРУЕМ ФАЙЛ ШАБЛОНА
        shutil.copy(template_report_path, report_copy)

        # Загружаем скопированный файл
        wb = load_workbook(report_copy)
        ws = wb.active  # Это уже готовый лист "Остатки на МП" с правильным оформлением

        thresholds = resolve_stock_thresholds(context, chat_id)

        # Заполняем данными (только значения, оформление остаётся как в шаблоне)
        row = 7
        while True:
            cell_value = ws[f"A{row}"].value
            if not cell_value or str(cell_value).strip().upper() == "ИТОГО":
                break

            art_name = str(cell_value).strip()

            # Ищем template_id по имени во ВСЕХ кабинетах
            template_id = None
            all_id_to_name = [
                ozon1_id_to_name,
                ozon2_id_to_name,
                ozon3_id_to_name,
                wb1_id_to_name,
                wb2_id_to_name,
                wb3_id_to_name
            ]

            for id_to_name in all_id_to_name:
                for tid, name in id_to_name.items():
                    if str(name).strip().lower() == art_name.lower():
                        template_id = tid
                        break
                if template_id is not None:
                    break

            if template_id is not None:
                # --- Ozon 1 ---
                o1 = ozon1_agg.get(template_id, {'avail': 0, 'return': 0, 'prep': 0})
                ws[f"B{row}"] = o1['avail']
                ws[f"C{row}"] = o1['return']
                ws[f"D{row}"] = o1['prep']
                ws[f"E{row}"] = o1['avail'] + o1['return'] + o1['prep']
                if thresholds and template_id in ozon1_linked_ids:
                    apply_fill_to_cells(ws, [row], [5], thresholds)

                # --- Ozon 2 ---
                o2 = ozon2_agg.get(template_id, {'avail': 0, 'return': 0, 'prep': 0})
                ws[f"G{row}"] = o2['avail']
                ws[f"H{row}"] = o2['return']
                ws[f"I{row}"] = o2['prep']
                ws[f"J{row}"] = o2['avail'] + o2['return'] + o2['prep']
                if thresholds and template_id in ozon2_linked_ids:
                    apply_fill_to_cells(ws, [row], [10], thresholds)

                # --- Ozon 3 ---
                o3 = ozon3_agg.get(template_id, {'avail': 0, 'return': 0, 'prep': 0})
                ws[f"L{row}"] = o3['avail']
                ws[f"M{row}"] = o3['return']
                ws[f"N{row}"] = o3['prep']
                ws[f"O{row}"] = o3['avail'] + o3['return'] + o3['prep']
                if thresholds and template_id in ozon3_linked_ids:
                    apply_fill_to_cells(ws, [row], [15], thresholds)

                # --- WB 1 ---
                w1 = wb1_agg.get(template_id, {'avail': 0, 'return': 0, 'inway': 0})
                ws[f"Q{row}"] = w1['avail']
                ws[f"R{row}"] = w1['return']
                ws[f"S{row}"] = w1['inway']
                ws[f"T{row}"] = w1['avail'] + w1['return'] + w1['inway']
                if thresholds and template_id in wb1_linked_ids:
                    apply_fill_to_cells(ws, [row], [20], thresholds)

                # --- WB 2 ---
                w2 = wb2_agg.get(template_id, {'avail': 0, 'return': 0, 'inway': 0})
                ws[f"V{row}"] = w2['avail']
                ws[f"W{row}"] = w2['return']
                ws[f"X{row}"] = w2['inway']
                ws[f"Y{row}"] = w2['avail'] + w2['return'] + w2['inway']
                if thresholds and template_id in wb2_linked_ids:
                    apply_fill_to_cells(ws, [row], [25], thresholds)

                # --- WB 3 ---
                w3 = wb3_agg.get(template_id, {'avail': 0, 'return': 0, 'inway': 0})
                ws[f"AA{row}"] = w3['avail']
                ws[f"AB{row}"] = w3['return']
                ws[f"AC{row}"] = w3['inway']
                ws[f"AD{row}"] = w3['avail'] + w3['return'] + w3['inway']
                if thresholds and template_id in wb3_linked_ids:
                    apply_fill_to_cells(ws, [row], [30], thresholds)

            row += 1

        # === ДОПОЛНИТЕЛЬНЫЕ ЛИСТЫ ===

        # Ozon1 исходные артикулы
        if ozon1_raw_data:
            df_ozon1_raw = pd.DataFrame(ozon1_raw_data).sort_values(by='Категория',
                                                                    key=lambda x: x.str.lower()).reset_index(drop=True)
            headers_ozon1 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей",
                              "Подготовка к продаже", "Итого на МП"]
            ws_ozon1 = wb.create_sheet(title="Ozon1 исходные артикулы")
            _write_sheet(ws_ozon1, df_ozon1_raw, headers_ozon1, has_name=True)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_ozon1_raw, "Артикул", ozon1_arts_set)
                total_col = headers_ozon1.index("Итого на МП") + 1
                apply_fill_to_cells(ws_ozon1, rows_to_color, [total_col], thresholds)
        else:
            ws_ozon1 = wb.create_sheet(title="Ozon1 исходные артикулы")
            ws_ozon1.append(["Нет данных"])

        # Ozon2 исходные артикулы
        if ozon2_raw_data:
            df_ozon2_raw = pd.DataFrame(ozon2_raw_data).sort_values(by='Категория',
                                                                    key=lambda x: x.str.lower()).reset_index(drop=True)
            headers_ozon2 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей",
                              "Подготовка к продаже", "Итого на МП"]
            ws_ozon2 = wb.create_sheet(title="Ozon2 исходные артикулы")
            _write_sheet(ws_ozon2, df_ozon2_raw, headers_ozon2, has_name=True)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_ozon2_raw, "Артикул", ozon2_arts_set)
                total_col = headers_ozon2.index("Итого на МП") + 1
                apply_fill_to_cells(ws_ozon2, rows_to_color, [total_col], thresholds)
        else:
            ws_ozon2 = wb.create_sheet(title="Ozon2 исходные артикулы")
            ws_ozon2.append(["Нет данных"])

        # Ozon3 исходные артикулы
        if ozon3_raw_data:
            df_ozon3_raw = pd.DataFrame(ozon3_raw_data).sort_values(by='Категория',
                                                                    key=lambda x: x.str.lower()).reset_index(drop=True)
            headers_ozon3 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей",
                              "Подготовка к продаже", "Итого на МП"]
            ws_ozon3 = wb.create_sheet(title="Ozon AGNIA исходные артикулы")
            _write_sheet(ws_ozon3, df_ozon3_raw, headers_ozon3, has_name=True)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_ozon3_raw, "Артикул", ozon3_arts_set)
                total_col = headers_ozon3.index("Итого на МП") + 1
                apply_fill_to_cells(ws_ozon3, rows_to_color, [total_col], thresholds)
        else:
            ws_ozon3 = wb.create_sheet(title="Ozon AGNIA исходные артикулы")
            ws_ozon3.append(["Нет данных"])

        # WB1 исходные артикулы
        if wb1_raw_data:
            df_wb1_raw = pd.DataFrame(wb1_raw_data).sort_values(by='Артикул').reset_index(drop=True)
            headers_wb1 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей", "В пути до покупателей",
                            "Итого на МП"]
            ws_wb1 = wb.create_sheet(title="WB1 исходные артикулы")
            _write_sheet(ws_wb1, df_wb1_raw, headers_wb1, has_name=False)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_wb1_raw, "Артикул", wb1_arts_set)
                total_col = headers_wb1.index("Итого на МП") + 1
                apply_fill_to_cells(ws_wb1, rows_to_color, [total_col], thresholds)
        else:
            ws_wb1 = wb.create_sheet(title="WB1 исходные артикулы")
            ws_wb1.append(["Нет данных"])

        # WB2 исходные артикулы
        if wb2_raw_data:
            df_wb2_raw = pd.DataFrame(wb2_raw_data).sort_values(by='Артикул').reset_index(drop=True)
            headers_wb2 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей", "В пути до покупателей",
                            "Итого на МП"]
            ws_wb2 = wb.create_sheet(title="WB2 исходные артикулы")
            _write_sheet(ws_wb2, df_wb2_raw, headers_wb2, has_name=False)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_wb2_raw, "Артикул", wb2_arts_set)
                total_col = headers_wb2.index("Итого на МП") + 1
                apply_fill_to_cells(ws_wb2, rows_to_color, [total_col], thresholds)
        else:
            ws_wb2 = wb.create_sheet(title="WB2 исходные артикулы")
            ws_wb2.append(["Нет данных"])

        # WB3 исходные артикулы
        if wb3_raw_data:
            df_wb3_raw = pd.DataFrame(wb3_raw_data).sort_values(by='Артикул').reset_index(drop=True)
            headers_wb3 = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей", "В пути до покупателей",
                            "Итого на МП"]
            ws_wb3 = wb.create_sheet(title="WB AGNIA исходные артикулы")
            _write_sheet(ws_wb3, df_wb3_raw, headers_wb3, has_name=False)
            if thresholds:
                rows_to_color = _get_rows_to_color(df_wb3_raw, "Артикул", wb3_arts_set)
                total_col = headers_wb3.index("Итого на МП") + 1
                apply_fill_to_cells(ws_wb3, rows_to_color, [total_col], thresholds)
        else:
            ws_wb3 = wb.create_sheet(title="WB AGNIA исходные артикулы")
            ws_wb3.append(["Нет данных"])

        wb.save(report_copy)

        # === СВОДКА ===
        ozon1_total_avail = sum(data['avail'] for data in ozon1_raw_dict.values())
        ozon1_total_return = sum(data['return'] for data in ozon1_raw_dict.values())
        ozon1_total_prep = sum(data['prep'] for data in ozon1_raw_dict.values())
        ozon1_total_mp = ozon1_total_avail + ozon1_total_return + ozon1_total_prep

        ozon2_total_avail = sum(data['avail'] for data in ozon2_raw_dict.values())
        ozon2_total_return = sum(data['return'] for data in ozon2_raw_dict.values())
        ozon2_total_prep = sum(data['prep'] for data in ozon2_raw_dict.values())
        ozon2_total_mp = ozon2_total_avail + ozon2_total_return + ozon2_total_prep

        ozon3_total_avail = sum(data['avail'] for data in ozon3_raw_dict.values())
        ozon3_total_return = sum(data['return'] for data in ozon3_raw_dict.values())
        ozon3_total_prep = sum(data['prep'] for data in ozon3_raw_dict.values())
        ozon3_total_mp = ozon3_total_avail + ozon3_total_return + ozon3_total_prep

        wb1_total_avail = sum(data['avail'] for data in wb1_raw_dict.values())
        wb1_total_return = sum(data['return'] for data in wb1_raw_dict.values())
        wb1_total_inway = sum(data['inway'] for data in wb1_raw_dict.values())
        wb1_total_mp = wb1_total_avail + wb1_total_return + wb1_total_inway

        wb2_total_avail = sum(data['avail'] for data in wb2_raw_dict.values())
        wb2_total_return = sum(data['return'] for data in wb2_raw_dict.values())
        wb2_total_inway = sum(data['inway'] for data in wb2_raw_dict.values())
        wb2_total_mp = wb2_total_avail + wb2_total_return + wb2_total_inway

        wb3_total_avail = sum(data['avail'] for data in wb3_raw_dict.values())
        wb3_total_return = sum(data['return'] for data in wb3_raw_dict.values())
        wb3_total_inway = sum(data['inway'] for data in wb3_raw_dict.values())
        wb3_total_mp = wb3_total_avail + wb3_total_return + wb3_total_inway

        total_all_mp = (
            ozon1_total_mp + ozon2_total_mp + ozon3_total_mp +
            wb1_total_mp + wb2_total_mp + wb3_total_mp
        )

        def fmt(x):
            return f"{x:,}".replace(",", " ")

        # === ДИНАМИЧЕСКИЙ ЗАГОЛОВОК ===
        summary_text = (
            f"📊 <b>{frequency_label} отчёт по остаткам на всех маркетплейсах</b>\n"
            f"📅 Дата: {time.strftime('%Y-%m-%d %H:%M')}\n\n"

            f"🏪 <b>Ozon Кабинет 1 (Nimba)</b>\n"
            f"   📦 Доступно на складах: {fmt(ozon1_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(ozon1_total_return)} шт\n"
            f"   🔄 Подготовка к продаже: {fmt(ozon1_total_prep)} шт\n"
            f"   ✅ Итого на МП: {fmt(ozon1_total_mp)} шт\n\n"

            f"🏬 <b>Ozon Кабинет 2 (Galioni)</b>\n"
            f"   📦 Доступно на складах: {fmt(ozon2_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(ozon2_total_return)} шт\n"
            f"   🔄 Подготовка к продаже: {fmt(ozon2_total_prep)} шт\n"
            f"   ✅ Итого на МП: {fmt(ozon2_total_mp)} шт\n\n"

            f"🏢 <b>Ozon Кабинет 3 (AGNIA)</b>\n"
            f"   📦 Доступно на складах: {fmt(ozon3_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(ozon3_total_return)} шт\n"
            f"   🔄 Подготовка к продаже: {fmt(ozon3_total_prep)} шт\n"
            f"   ✅ Итого на МП: {fmt(ozon3_total_mp)} шт\n\n"

            f"🏪 <b>Wildberries Кабинет 1 (Nimba)</b>\n"
            f"   📦 Доступно на складах: {fmt(wb1_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(wb1_total_return)} шт\n"
            f"   🚚 В пути до покупателей: {fmt(wb1_total_mp - wb1_total_avail - wb1_total_return)} шт\n"
            f"   ✅ Итого на МП: {fmt(wb1_total_mp)} шт\n\n"

            f"🏬 <b>Wildberries Кабинет 2 (Galioni)</b>\n"
            f"   📦 Доступно на складах: {fmt(wb2_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(wb2_total_return)} шт\n"
            f"   🚚 В пути до покупателей: {fmt(wb2_total_mp - wb2_total_avail - wb2_total_return)} шт\n"
            f"   ✅ Итого на МП: {fmt(wb2_total_mp)} шт\n\n"

            f"🏢 <b>Wildberries Кабинет 3 (AGNIA)</b>\n"
            f"   📦 Доступно на складах: {fmt(wb3_total_avail)} шт\n"
            f"   ↩️ Возвращаются от покупателей: {fmt(wb3_total_return)} шт\n"
            f"   🚚 В пути до покупателей: {fmt(wb3_total_mp - wb3_total_avail - wb3_total_return)} шт\n"
            f"   ✅ Итого на МП: {fmt(wb3_total_mp)} шт\n\n"

            f"🔹 <b>ВСЕГО на всех маркетплейсах:</b> {fmt(total_all_mp)} шт"
        )

        # === ОТПРАВКА С ДИНАМИЧЕСКИМ CAPTION ===
        await context.bot.send_document(
            chat_id=chat_id,
            document=open(report_copy, 'rb'),
            caption=f"📊 {frequency_label} отчёт: остатки на всех маркетплейсах"
        )
        await context.bot.send_message(chat_id=chat_id, text=summary_text, parse_mode="HTML")

        if os.path.exists(report_copy):
            os.remove(report_copy)

    except Exception as e:
        logger.error(f"Ошибка в автоматическом отчёте по всем МП: {e}", exc_info=True)
        await context.bot.send_message(
            chat_id=chat_id,
            text=f"❌ Ошибка при генерации {frequency_label.lower()} отчёта по всем маркетплейсам: {str(e)}"
        )

