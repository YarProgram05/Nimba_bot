import os
import sys
import logging
import asyncio
import pandas as pd
from datetime import datetime, timezone, timedelta
import requests
import time
from telegram import Update, ReplyKeyboardRemove, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import CallbackContext, ConversationHandler
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Настройка путей
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)
utils_dir = os.path.join(root_dir, 'utils')

if root_dir not in sys.path:
    sys.path.append(root_dir)
if utils_dir not in sys.path:
    sys.path.append(utils_dir)

logger = logging.getLogger(__name__)

# Состояния
from states import OZON_SALES_CABINET_CHOICE, OZON_SALES_DATE_START, OZON_SALES_DATE_END

# Импорт новой функции из template_loader
from utils.template_loader import get_cabinet_articles_by_template_id


def split_by_max_period(start: datetime, end: datetime, max_days: int):
    """Разбивает диапазон на чанки не длиннее max_days дней."""
    chunks = []
    current = start.date()
    end_date = end.date()
    while current <= end_date:
        chunk_end = min(current + timedelta(days=max_days - 1), end_date)
        chunks.append((
            datetime.combine(current, datetime.min.time()).replace(tzinfo=timezone.utc),
            datetime.combine(chunk_end, datetime.max.time()).replace(tzinfo=timezone.utc)
        ))
        current = chunk_end + timedelta(days=1)
    return chunks


class OzonAPI:
    def __init__(self, cabinet_id=1):
        from dotenv import load_dotenv
        load_dotenv()

        if cabinet_id == 1:
            self.client_id = os.getenv('OZON_CLIENT_ID_1')
            self.api_key = os.getenv('OZON_API_KEY_1')
        elif cabinet_id == 2:
            self.client_id = os.getenv('OZON_CLIENT_ID_2')
            self.api_key = os.getenv('OZON_API_KEY_2')
        elif cabinet_id == 3:
            self.client_id = os.getenv('OZON_CLIENT_ID_3')
            self.api_key = os.getenv('OZON_API_KEY_3')
        else:
            raise ValueError("Поддерживаются только cabinet_id 1, 2 или 3")

        if not self.client_id or not self.api_key:
            raise ValueError(f"❌ OZON_CLIENT_ID или OZON_API_KEY не заданы в .env для кабинета {cabinet_id}")

        self.headers = {
            "Client-Id": self.client_id,
            "Api-Key": self.api_key,
            "Content-Type": "application/json"
        }

    def _post_with_retry(self, url: str, payload: dict, max_retries: int = 3):
        for attempt in range(max_retries):
            try:
                response = requests.post(url, json=payload, headers=self.headers, timeout=30)
                if response.status_code == 200:
                    return response.json()
                elif response.status_code in (502, 503, 504):
                    logger.warning(f"⚠️ Ozon API error {response.status_code} on {url}, attempt {attempt + 1}")
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)
                        continue
                else:
                    raise Exception(f"API error {response.status_code}: {response.text}")
            except requests.exceptions.Timeout:
                logger.warning(f"⚠️ Timeout on {url}, attempt {attempt + 1}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
                    continue
        raise Exception(f"Failed to call {url} after {max_retries} attempts")

    def get_fbo_postings(self, since: str, to: str):
        all_postings = []
        offset = 0
        limit = 1000
        while True:
            payload = {
                "dir": "ASC",
                "filter": {"since": since, "to": to},
                "limit": limit,
                "offset": offset,
                "with": {"analytics_data": False, "financial_data": False}
            }
            data = self._post_with_retry("https://api-seller.ozon.ru/v2/posting/fbo/list", payload)
            postings = data.get("result", [])
            if not postings:
                break
            all_postings.extend(postings)
            if len(postings) < limit:
                break
            offset += limit
        return all_postings

    def get_financial_operations(self, date_from: str, date_to: str):
        all_ops = []
        page = 1
        while True:
            payload = {
                "filter": {"date": {"from": date_from, "to": date_to}},
                "page": page,
                "page_size": 1000
            }
            data = self._post_with_retry("https://api-seller.ozon.ru/v3/finance/transaction/list", payload)
            ops = data.get("result", {}).get("operations", [])
            if not ops:
                break
            all_ops.extend(ops)
            if page > 100:
                break
            page += 1
        return all_ops


def parse_date_input(date_str: str) -> datetime:
    return datetime.strptime(date_str.strip(), "%d.%m.%Y").replace(tzinfo=timezone.utc)


def validate_date_format(text: str) -> bool:
    import re
    return bool(re.fullmatch(r'\d{2}\.\d{2}\.\d{4}', text.strip()))


def split_by_calendar_months(start_dt: datetime, end_dt: datetime):
    """
    Разбивает диапазон на чанки по календарным месяцам.
    Пример: 10.03.2025 – 26.06.2025 →
        [10.03–31.03], [01.04–30.04], [01.05–31.05], [01.06–26.06]
    """
    chunks = []
    current_start = start_dt.date()
    end_date = end_dt.date()

    while current_start <= end_date:
        if current_start.month == 12:
            next_month = current_start.replace(year=current_start.year + 1, month=1)
        else:
            next_month = current_start.replace(month=current_start.month + 1)
        month_end = next_month - timedelta(days=1)
        chunk_end = min(month_end, end_date)

        chunks.append((
            datetime.combine(current_start, datetime.min.time()).replace(tzinfo=timezone.utc),
            datetime.combine(chunk_end, datetime.max.time()).replace(tzinfo=timezone.utc)
        ))
        current_start = next_month

    return chunks


async def start_ozon_sales(update: Update, context: CallbackContext) -> int:
    context.user_data['current_flow'] = 'sales'

    keyboard = [
        [InlineKeyboardButton("🏪 Озон_1 Nimba", callback_data='cabinet_1')],
        [InlineKeyboardButton("🏬 Озон_2 Galioni", callback_data='cabinet_2')],
        [InlineKeyboardButton("🏢 Озон_3 AGNIA", callback_data='cabinet_3')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    sent_message = await update.message.reply_text(
        "🏢 Выберите кабинет Ozon для выгрузки продаж:",
        reply_markup=reply_markup
    )
    # Сохраняем ID сообщения для последующего удаления
    context.user_data['ozon_sales_initial_message_id'] = sent_message.message_id

    return OZON_SALES_CABINET_CHOICE


async def handle_sales_cabinet_choice(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()

    cabinet_data = query.data
    cabinet_map = {
        'cabinet_1': 1,
        'cabinet_2': 2,
        'cabinet_3': 3
    }
    if cabinet_data not in cabinet_map:
        await query.message.reply_text("❌ Неизвестный кабинет.")
        return ConversationHandler.END

    cabinet_id = cabinet_map[cabinet_data]
    context.user_data['ozon_sales_cabinet_id'] = cabinet_id

    await query.message.edit_reply_markup(reply_markup=None)
    await query.message.reply_text(
        f"✅ Выбран кабинет: Озон {cabinet_id}\n\n"
        "📅 Введите дату начала периода в формате ДД.ММ.ГГГГ:"
    )
    return OZON_SALES_DATE_START


async def handle_sales_date_start(update: Update, context: CallbackContext) -> int:
    logger.info(
        f"[OZON SALES] Получена дата начала: '{update.message.text}' от пользователя {update.effective_user.id}")

    text = update.message.text.strip()
    if not validate_date_format(text):
        await update.message.reply_text("❌ Неверный формат даты. Введите в формате ДД.ММ.ГГГГ:")
        return OZON_SALES_DATE_START

    try:
        start_dt = parse_date_input(text)
        today = datetime.now(timezone.utc).date()
        if start_dt.date() > today:
            await update.message.reply_text("❌ Дата начала не может быть в будущем.")
            return OZON_SALES_DATE_START
    except ValueError:
        await update.message.reply_text("❌ Некорректная дата. Введите в формате ДД.ММ.ГГГГ:")
        return OZON_SALES_DATE_START
    except Exception as e:
        logger.error(f"Неожиданная ошибка при парсинге даты: {e}")
        await update.message.reply_text("❌ Произошла ошибка. Попробуйте снова.")
        return OZON_SALES_DATE_START

    context.user_data['ozon_sales_start_date'] = text
    await update.message.reply_text("📅 Введите дату окончания периода в формате ДД.ММ.ГГГГ:")
    return OZON_SALES_DATE_END


async def handle_sales_date_end(update: Update, context: CallbackContext) -> int:
    text = update.message.text.strip()
    if not validate_date_format(text):
        await update.message.reply_text("❌ Неверный формат даты. Введите в формате ДД.ММ.ГГГГ:")
        return OZON_SALES_DATE_END

    try:
        start_str = context.user_data['ozon_sales_start_date']
        start_dt = parse_date_input(start_str)
        end_dt = parse_date_input(text)

        if end_dt < start_dt:
            await update.message.reply_text("❌ Дата окончания не может быть раньше начала.")
            return OZON_SALES_DATE_END

        if end_dt.date() > datetime.now(timezone.utc).date():
            await update.message.reply_text("❌ Дата окончания не может быть в будущем.")
            return OZON_SALES_DATE_END

    except Exception as e:
        await update.message.reply_text("❌ Ошибка при обработке дат. Введите в формате ДД.ММ.ГГГГ:")
        return OZON_SALES_DATE_END

    context.user_data['ozon_sales_end_date'] = text
    loading_message = await update.message.reply_text(
        "⏳ Загружаю данные с Ozon API... Это может занять несколько минут.")
    context.user_data['ozon_sales_loading_message_id'] = loading_message.message_id

    start_time = time.time()
    try:
        cabinet_id = context.user_data['ozon_sales_cabinet_id']
        start_str = context.user_data['ozon_sales_start_date']
        end_str = context.user_data['ozon_sales_end_date']

        start_dt = parse_date_input(start_str)
        end_dt = parse_date_input(end_str)

        ozon = OzonAPI(cabinet_id=cabinet_id)

        # === FBO: до 365 дней ===
        fbo_chunks = split_by_max_period(start_dt, end_dt, max_days=365)
        all_postings = []
        for i, (chunk_start, chunk_end) in enumerate(fbo_chunks, 1):
            logger.info(f"Запрос FBO {i}/{len(fbo_chunks)}: {chunk_start.date()} – {chunk_end.date()}")
            start_iso = chunk_start.strftime("%Y-%m-%dT00:00:00Z")
            end_iso = chunk_end.strftime("%Y-%m-%dT23:59:59Z")
            postings = ozon.get_fbo_postings(start_iso, end_iso)
            all_postings.extend(postings)
            await asyncio.sleep(0.1)

        # === Финансы: по календарным месяцам (как раньше) ===
        finance_chunks = split_by_calendar_months(start_dt, end_dt)
        all_operations = []
        for i, (chunk_start, chunk_end) in enumerate(finance_chunks, 1):
            logger.info(f"Запрос финансов {i}/{len(finance_chunks)}: {chunk_start.date()} – {chunk_end.date()}")
            start_iso = chunk_start.strftime("%Y-%m-%dT00:00:00.000Z")
            end_iso = chunk_end.strftime("%Y-%m-%dT23:59:59.999Z")

            # Retry-логика (оставляем для надёжности)
            ops = None
            for attempt in range(3):
                try:
                    ops = ozon.get_financial_operations(start_iso, end_iso)
                    break
                except Exception as e:
                    if any(code in str(e) for code in ["504", "502", "timeout"]):
                        logger.warning(f"⚠️ Финансы: попытка {attempt + 1}/3 провалена для {chunk_start.date()}")
                        if attempt < 2:
                            await asyncio.sleep(2 ** attempt)
                            continue
                    raise

            all_operations.extend(ops)
            await asyncio.sleep(0.1)

        logger.info(f"Данные загружены за {time.time() - start_time:.1f} сек")

        # === Обработка FBO ===
        art_data = {}
        for p in all_postings:
            posting_number = p.get("posting_number")
            status = p.get("status")
            for prod in p.get("products", []):
                offer_id = str(prod.get("offer_id", "")).strip().lower()
                if not offer_id:
                    continue
                qty = prod.get("quantity", 0)

                if offer_id not in art_data:
                    art_data[offer_id] = {"orders": set(), "purchases": 0, "cancels": 0}

                art_data[offer_id]["orders"].add(posting_number)
                if status == "delivered":
                    art_data[offer_id]["purchases"] += qty
                elif status == "cancelled":
                    art_data[offer_id]["cancels"] += qty

        for art in art_data:
            art_data[art]["orders"] = len(art_data[art]["orders"])

        total_purchases = sum(data["purchases"] for data in art_data.values())
        total_cancels = sum(data["cancels"] for data in art_data.values())
        total_orders = sum(data["orders"] for data in art_data.values())

        # === Обработка финансов ===
        operations = all_operations
        skus = set()
        for op in operations:
            for item in op.get("items", []):
                sku = item.get("sku")
                if sku is not None:
                    skus.add(sku)

        sku_to_offer = {}
        if skus:
            valid_skus = []
            for s in skus:
                try:
                    if isinstance(s, float) and s.is_integer():
                        valid_skus.append(str(int(s)))
                    else:
                        valid_skus.append(str(s))
                except (ValueError, TypeError, OverflowError):
                    continue

            for chunk in [valid_skus[i:i + 1000] for i in range(0, len(valid_skus), 1000)]:
                payload = {"sku": chunk}
                response = requests.post(
                    "https://api-seller.ozon.ru/v3/product/info/list",
                    headers=ozon.headers,
                    json=payload,
                    timeout=30
                )
                if response.status_code == 200:
                    items = response.json().get("items", [])
                    for item in items:
                        sku = item.get("sku")
                        offer_id = item.get("offer_id")
                        if sku is not None and offer_id:
                            sku_to_offer[str(sku)] = str(offer_id).strip().lower()

        # === Доход по артикулам ===
        income = {}
        for op in operations:
            amount = op.get("amount", 0)
            if amount == 0:
                continue

            items = op.get("items", [])
            operation_type_name = op.get("operation_type_name", "").strip()

            if items:
                offer_ids_found = []
                for item in items:
                    sku = item.get("sku")
                    if sku is not None:
                        offer_id = sku_to_offer.get(str(sku))
                        if offer_id:
                            offer_ids_found.append(offer_id)
                if offer_ids_found:
                    split_amount = amount / len(offer_ids_found)
                    for offer_id in offer_ids_found:
                        income[offer_id] = income.get(offer_id, 0) + split_amount
                else:
                    art = f"тип_начисления: {operation_type_name or op.get('type', 'other')}"
                    income[art] = income.get(art, 0) + amount
            else:
                art = f"тип_начисления: {operation_type_name or op.get('type', 'other')}"
                income[art] = income.get(art, 0) + amount

        total_income = sum(income.values())

        # === Загрузка шаблона ===
        sheet_map = {
            1: "Отдельно Озон Nimba",
            2: "Отдельно Озон Galioni",
            3: "Отдельно Озон AGNIA"
        }
        sheet_name = sheet_map.get(cabinet_id)
        if not sheet_name:
            raise ValueError(f"Неподдерживаемый кабинет Ozon: {cabinet_id}")

        template_id_to_name, template_id_to_cabinet_arts = get_cabinet_articles_by_template_id(sheet_name)

        # Получаем main_ids_ordered — ID в порядке появления в Excel (без дубликатов)
        template_path = os.path.join(root_dir, "База данных артикулов для выкупов и начислений.xlsx")
        if not os.path.exists(template_path):
            template_path = "База данных артикулов для выкупов и начислений.xlsx"
        df_order = pd.read_excel(template_path, sheet_name=sheet_name)
        main_ids_ordered = []
        seen = set()
        for _, row in df_order.iterrows():
            if not pd.isna(row.get('ID')):
                tid = int(row['ID'])
                if tid not in seen:
                    main_ids_ordered.append(tid)
                    seen.add(tid)

        # Построение art_to_id из template_id_to_cabinet_arts
        art_to_id = {}
        for template_id, cabinet_arts in template_id_to_cabinet_arts.items():
            for art in cabinet_arts:
                clean_art = str(art).strip().lower()
                art_to_id[clean_art] = template_id

        id_to_name = template_id_to_name

        # === Группировка по шаблону ===
        grouped = {}
        for group_id in main_ids_ordered:
            grouped[group_id] = {
                'name': id_to_name.get(group_id, f"Группа {group_id}"),
                'orders': 0,
                'purchases': 0,
                'cancels': 0,
                'income': 0
            }

        unmatched = {}
        all_arts = set(art_data.keys()) | set(income.keys())

        for art in all_arts:
            if art.lower().startswith("тип_начисления:"):
                unmatched[art] = {
                    'name': art,
                    'orders': art_data.get(art, {}).get('orders', 0),
                    'purchases': art_data.get(art, {}).get('purchases', 0),
                    'cancels': art_data.get(art, {}).get('cancels', 0),
                    'income': income.get(art, 0)
                }
                continue

            group_id = art_to_id.get(art)
            if group_id is not None:
                grouped[group_id]['orders'] += art_data.get(art, {}).get('orders', 0)
                grouped[group_id]['purchases'] += art_data.get(art, {}).get('purchases', 0)
                grouped[group_id]['cancels'] += art_data.get(art, {}).get('cancels', 0)
                grouped[group_id]['income'] += income.get(art, 0)
            else:
                unmatched[art] = {
                    'name': f"НЕОПОЗНАННЫЙ_АРТИКУЛ: {art}",
                    'orders': art_data.get(art, {}).get('orders', 0),
                    'purchases': art_data.get(art, {}).get('purchases', 0),
                    'cancels': art_data.get(art, {}).get('cancels', 0),
                    'income': income.get(art, 0)
                }

        # === Подготовка исходных артикулов для Excel ===
        raw_art_data = []
        for art in art_data:
            if art.lower().startswith("тип_начисления:"):
                continue
            data = art_data[art]
            purchases = data["purchases"]
            cancels = data["cancels"]
            orders = data["orders"]
            profit = income.get(art, 0)

            total_shipments = purchases + cancels
            purchase_percent_val = (purchases / total_shipments * 100) if total_shipments > 0 else 0
            profit_per_unit = profit / purchases if purchases > 0 else 0

            raw_art_data.append({
                "art": art,
                "orders": orders,
                "purchases": purchases,
                "cancels": cancels,
                "profit": profit,
                "purchase_percent": purchase_percent_val,
                "profit_per_unit": profit_per_unit
            })

        raw_art_data.sort(key=lambda x: x["purchases"], reverse=True)

        # === Создание Excel ===
        report_path = f"Ozon_Sales_{start_dt.strftime('%d%m%Y')}-{end_dt.strftime('%d%m%Y')}.xlsx"
        create_excel_report(
            grouped, unmatched, id_to_name, main_ids_ordered, report_path,
            total_orders, total_purchases, total_cancels, total_income,
            raw_art_data=raw_art_data
        )

        # === Топ-5 для текста ===
        top_5 = raw_art_data[:5]

        def fmt_num(x):
            if isinstance(x, float):
                return f"{x:,.2f}".replace(",", " ")
            return f"{x:,}".replace(",", " ")

        total_shipments = total_purchases + total_cancels
        purchase_percent = (total_purchases / total_shipments * 100) if total_shipments > 0 else 0
        avg_profit_per_unit = total_income / total_purchases if total_purchases > 0 else 0

        text_summary = (
            f"📊 <b>Сводка по продажам Ozon</b>\n"
            f"Кабинет: <b>Озон {cabinet_id}</b>\n"
            f"Период: <b>{start_str} – {end_str}</b>\n\n"
            f"📦 <b>Заказы:</b> {fmt_num(total_orders)} шт\n"
            f"✅ <b>Выкупы:</b> {fmt_num(total_purchases)} шт\n"
            f"❌ <b>Отмены:</b> {fmt_num(total_cancels)} шт\n"
            f"💰 <b>Валовая прибыль:</b> {fmt_num(total_income)} ₽\n"
            f"📈 <b>Прибыль на 1 ед:</b> {fmt_num(avg_profit_per_unit)} ₽\n"
            f"🔄 <b>Процент выкупов:</b> {purchase_percent:.2f}%"
            f"\n\n🏆 <b>Топ-5 артикулов по выкупам:</b>\n"
        )

        if top_5:
            for i, item in enumerate(top_5, 1):
                art = item["art"]
                purchases = item["purchases"]
                profit = item["profit"]
                text_summary += (
                    f"🔹 {i}. <b>{art}</b>\n"
                    f"   ✅ Выкупы: {fmt_num(purchases)} шт\n"
                    f"   💰 Прибыль: {fmt_num(profit)} ₽\n\n"
                )
        else:
            text_summary += "   — Нет данных по выкупам\n"

        # === Отправка ===
        await update.message.reply_document(
            document=open(report_path, 'rb'),
            caption=f"📊 Подробный отчёт в Excel по продажам Ozon (кабинет {cabinet_id})\nПериод: {start_str} – {end_str}"
        )

        await update.message.reply_text(
            text_summary,
            parse_mode="HTML",
            reply_markup=ReplyKeyboardRemove()
        )

        if os.path.exists(report_path):
            os.remove(report_path)

            # Удаляем служебные сообщения
        chat_id = update.effective_chat.id
        try:
            initial_msg_id = context.user_data.get('ozon_sales_initial_message_id')
            if initial_msg_id:
                await context.bot.delete_message(chat_id=chat_id, message_id=initial_msg_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить начальное сообщение: {e}")

        try:
            loading_msg_id = context.user_data.get('ozon_sales_loading_message_id')
            if loading_msg_id:
                await context.bot.delete_message(chat_id=chat_id, message_id=loading_msg_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить сообщение о загрузке: {e}")

    except Exception as e:
        logger.error(f"Ошибка при генерации отчёта продаж: {e}", exc_info=True)
        await update.message.reply_text(
            f"❌ Ошибка: {str(e)}",
            reply_markup=ReplyKeyboardRemove()
        )
        # Удаляем служебные сообщения и при ошибке
        chat_id = update.effective_chat.id
        try:
            initial_msg_id = context.user_data.get('ozon_sales_initial_message_id')
            if initial_msg_id:
                await context.bot.delete_message(chat_id=chat_id, message_id=initial_msg_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить начальное сообщение при ошибке: {e}")

        try:
            loading_msg_id = context.user_data.get('ozon_sales_loading_message_id')
            if loading_msg_id:
                await context.bot.delete_message(chat_id=chat_id, message_id=loading_msg_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить сообщение о загрузке при ошибке: {e}")

    return ConversationHandler.END

def create_excel_report(grouped, unmatched, id_to_name, main_ids_ordered, output_path,
                        total_orders, total_purchases, total_cancels, total_income,
                        raw_art_data=None):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Сводный"

    # === 1. Общая сводка ===
    headers1 = ["Показатель", "Значение"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    ws1.append(["Заказы, шт", total_orders])
    ws1.append(["Выкупы, шт", total_purchases])
    ws1.append(["Отмены, шт", total_cancels])
    ws1.append(["Валовая прибыль, руб", total_income])

    avg_profit_per_unit = total_income / total_purchases if total_purchases > 0 else 0
    ws1.append(["Прибыль на 1 ед, руб", avg_profit_per_unit])

    total_shipments = total_purchases + total_cancels
    purchase_percent = (total_purchases / total_shipments * 100) if total_shipments > 0 else 0
    ws1.append(["Процент выкупов", f"{purchase_percent:.2f}%"])

    # === 2. Разделитель ===
    ws1.append([])  # пустая строка

    # === 3. ТОП-5 артикулов по выкупам ===
    if raw_art_data and len(raw_art_data) > 0:
        top_5 = raw_art_data[:5]

        # Заголовок
        ws1.append(["🏆 ТОП-5 артикулов по выкупам"])
        header_cell = ws1.cell(row=ws1.max_row, column=1)
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal="center")

        # Пустая строка
        ws1.append([])

        # Заголовки таблицы
        top_headers = ["Место", "Артикул", "Выкупы, шт", "Прибыль, ₽"]
        ws1.append(top_headers)
        for col in range(1, len(top_headers) + 1):
            cell = ws1.cell(row=ws1.max_row, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Данные
        for i, item in enumerate(top_5, 1):
            ws1.append([
                i,
                item["art"],
                item["purchases"],
                item["profit"]
            ])

    # === 4. Форматирование всего листа ===
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Применяем форматирование ко всем ячейкам с данными
    for row in ws1.iter_rows():
        for cell in row:
            if cell.value is not None:
                # Выравнивание по центру
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

    # Автоподбор ширины для листа "Сводный"
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column].width = adjusted_width

    # === 5. Остальные листы (без изменений) ===
    # Подробный
    ws2 = wb.create_sheet(title="Подробный")
    headers2 = [
        "Наименование",
        "Выкупы, шт",
        "Валовая прибыль, руб",
        "Процент выкупов",
        "Прибыль на 1 ед, руб",
        "Заказы, шт",
        "Отмены, шт"
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    row_index = 2
    for group_id in main_ids_ordered:
        data = grouped.get(group_id, {})
        name = data.get('name', f"Группа {group_id}")
        orders = data.get('orders', 0)
        purchases = data.get('purchases', 0)
        cancels = data.get('cancels', 0)
        income_val = data.get('income', 0)

        profit_per_unit = income_val / purchases if purchases > 0 else 0
        total_shipments = purchases + cancels
        purchase_percent_val = (purchases / total_shipments * 100) if total_shipments > 0 else 0

        ws2.append([
            name,
            purchases,
            income_val,
            f"{purchase_percent_val:.2f}%",
            profit_per_unit,
            orders,
            cancels
        ])

        percent_cell = ws2.cell(row=row_index, column=4)
        if purchase_percent_val <= 50:
            percent_cell.fill = red_fill
        elif 50 < purchase_percent_val <= 60:
            percent_cell.fill = orange_fill
        row_index += 1

    # Неопознанные
    unknown_articles = []
    service_types = []

    for art, data in unmatched.items():
        name = data['name']
        if name.startswith("НЕОПОЗНАННЫЙ_АРТИКУЛ:"):
            unknown_articles.append((name, data))
        elif name.lower().startswith("тип_начисления:"):
            clean_name = name.split(":", 1)[-1].strip()
            new_name = f"ТИП_НАЧИСЛЕНИЯ: {clean_name}"
            service_types.append((new_name, data))
        else:
            unknown_articles.append((name, data))

    unknown_articles.sort(key=lambda x: x[0])
    service_types.sort(key=lambda x: x[0])

    for name, data in unknown_articles:
        orders = data.get('orders', 0)
        purchases = data.get('purchases', 0)
        cancels = data.get('cancels', 0)
        income_val = data.get('income', 0)
        profit_per_unit = income_val / purchases if purchases > 0 else 0
        total_shipments = purchases + cancels
        purchase_percent_val = (purchases / total_shipments * 100) if total_shipments > 0 else 0

        ws2.append([
            name,
            purchases,
            income_val,
            f"{purchase_percent_val:.2f}%",
            profit_per_unit,
            orders,
            cancels
        ])

        percent_cell = ws2.cell(row=row_index, column=4)
        if purchase_percent_val <= 50:
            percent_cell.fill = red_fill
        elif 50 < purchase_percent_val <= 60:
            percent_cell.fill = orange_fill
        row_index += 1

    for name, data in service_types:
        income_val = data.get('income', 0)
        ws2.append([
            name,
            0,
            income_val,
            "—",
            0,
            0,
            0
        ])
        row_index += 1

    # === Исходные артикулы ===
    if raw_art_data:
        ws3 = wb.create_sheet(title="Исходные артикулы")
        headers3 = [
            "Артикул (offer_id)",
            "Выкупы, шт",
            "Валовая прибыль, руб",
            "Процент выкупов",
            "Прибыль на 1 ед, руб",
            "Заказы, шт",
            "Отмены, шт"
        ]
        ws3.append(headers3)
        for cell in ws3[1]:
            cell.font = Font(bold=True)

        row_idx = 2
        for item in raw_art_data:
            art = item["art"]
            purchases = item["purchases"]
            profit = item["profit"]
            purchase_percent = item["purchase_percent"]
            profit_per_unit = item["profit_per_unit"]
            orders = item["orders"]
            cancels = item["cancels"]

            ws3.append([
                art,
                purchases,
                profit,
                f"{purchase_percent:.2f}%",
                profit_per_unit,
                orders,
                cancels
            ])

            percent_cell = ws3.cell(row=row_idx, column=4)
            if purchase_percent <= 50:
                percent_cell.fill = red_fill
            elif 50 < purchase_percent <= 60:
                percent_cell.fill = orange_fill
            row_idx += 1

    # === Форматирование остальных листов ===
    for ws in [ws2, ws3] if raw_art_data else [ws2]:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(output_path)