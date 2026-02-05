# handlers/wb_remains_handler.py

import sys
import os
import pandas as pd
import logging
import time
import requests
from telegram import Update, ReplyKeyboardRemove, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import CallbackContext, ConversationHandler
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Настройка путей
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)
utils_dir = os.path.join(root_dir, 'utils')

if root_dir not in sys.path:
    sys.path.append(root_dir)
if utils_dir not in sys.path:
    sys.path.append(utils_dir)

logger = logging.getLogger(__name__)

from states import WB_REMAINS_CABINET_CHOICE  # ← ДОЛЖЕН БЫТЬ В states.py

# Импорт новой функции из template_loader
from utils.template_loader import get_cabinet_articles_by_template_id
from utils.stock_control import resolve_stock_thresholds, apply_fill_to_cells

# Content API endpoints
WB_CONTENT_BASE_URL = "https://content-api.wildberries.ru"


def clean_article(article):
    """Очистка артикула от лишних символов"""
    try:
        if not article:
            return None
        s = str(article)
        s = ''.join(c for c in s if c.isprintable())
        s = s.strip()
        return s if s else None
    except Exception:
        return None


def normalize_wb_size(value) -> str:
    """Нормализует размер WB для отчётов.

    Требование: если размера нет (0) или ONE — писать "единый".
    """
    if value is None:
        return "единый"
    s = str(value).strip()
    if not s:
        return "единый"
    s_up = s.upper()
    if s_up in {"0", "ONE", "ONE SIZE", "ONESIZE", "ЕДИНЫЙ", "ЕДИНЫЙ РАЗМЕР"}:
        return "единый"
    return s


class WildberriesAPI:
    def __init__(self, cabinet_id=1):
        from dotenv import load_dotenv
        load_dotenv()

        if cabinet_id == 1:
            self.api_token = os.getenv('WB_API_TOKEN_1')
        elif cabinet_id == 2:
            self.api_token = os.getenv('WB_API_TOKEN_2')
        elif cabinet_id == 3:
            self.api_token = os.getenv('WB_API_TOKEN_3')
        else:
            raise ValueError("Поддерживаются только cabinet_id 1, 2 или 3")

        if not self.api_token:
            raise ValueError(f"❌ WB_API_TOKEN не задан в .env для кабинета {cabinet_id}")

        self.headers = {
            'Authorization': self.api_token,
            'Content-Type': 'application/json'
        }

    def _content_post(self, path: str, payload: dict, timeout: int = 30) -> dict | None:
        url = f"{WB_CONTENT_BASE_URL}{path}"
        try:
            resp = requests.post(url, headers=self.headers, json=payload, timeout=timeout)
            if resp.status_code != 200:
                logger.warning(f"WB content-api POST {path}: {resp.status_code} - {resp.text}")
                return None
            return resp.json() or {}
        except Exception as e:
            logger.warning(f"WB content-api POST {path} error: {e}")
            return None

    def _content_get(self, path: str, timeout: int = 30) -> dict | None:
        url = f"{WB_CONTENT_BASE_URL}{path}"
        try:
            resp = requests.get(url, headers=self.headers, timeout=timeout)
            if resp.status_code != 200:
                logger.warning(f"WB content-api GET {path}: {resp.status_code} - {resp.text}")
                return None
            return resp.json() or {}
        except Exception as e:
            logger.warning(f"WB content-api GET {path} error: {e}")
            return None

    def _http_get_json(self, url: str, timeout: int = 30, headers: dict | None = None) -> tuple[int, str, dict | None]:
        """Утилита: GET и попытка распарсить JSON. Возвращает (status, text_snippet, json_or_none)."""
        try:
            resp = requests.get(url, headers=headers, timeout=timeout)
            txt = resp.text or ""
            data = None
            try:
                data = resp.json() if txt else None
            except Exception:
                data = None
            return resp.status_code, txt[:500], data
        except Exception as e:
            return -1, f"EXC: {e}", None

    def content_health_check(self) -> dict:
        """Проверка доступности content-api и формата ответов."""
        payload = {"settings": {"cursor": {"limit": 1}}}
        data = self._content_post("/content/v2/get/cards/list", payload, timeout=30)
        if data is None:
            return {"ok": False, "reason": "no_response"}
        return {
            "ok": True,
            "keys": list(data.keys()),
            "error": data.get("error"),
            "errorText": data.get("errorText"),
            "additionalErrors": data.get("additionalErrors"),
            "cards_len": len(data.get("cards") or []),
        }

    def content_probe(self, nm_id: int | None = None) -> None:
        """Пробует несколько вариантов запросов content-api и пишет подробные логи.

        Задача: понять, доступен ли вообще контент продавца данным токеном и какие endpoints работают.
        """
        probes: list[tuple[str, dict]] = [
            ("/content/v2/get/cards/list", {"settings": {"cursor": {"limit": 1}}}),
            ("/content/v2/get/cards/list", {"settings": {"cursor": {"limit": 1}, "filter": {}}}),
            ("/content/v2/get/cards/list", {"settings": {"cursor": {"limit": 1}, "filter": {"withPhoto": -1}}}),
        ]
        if nm_id is not None:
            # Популярные вариации фильтра по nmId (в разных версиях API встречаются разные ключи)
            probes.extend([
                ("/content/v2/get/cards/list", {"settings": {"cursor": {"limit": 10}, "filter": {"nmIDs": [int(nm_id)]}}}),
                ("/content/v2/get/cards/list", {"settings": {"cursor": {"limit": 10}, "filter": {"nmId": int(nm_id)}}}),
                ("/content/v2/get/cards/list", {"settings": {"cursor": {"limit": 10}, "filter": {"nmIds": [int(nm_id)]}}}),
            ])

        for path, payload in probes:
            try:
                data = self._content_post(path, payload, timeout=30)
                if data is None:
                    logger.warning(f"WB content-probe POST {path}: None payload={payload}")
                    continue
                keys = list(data.keys())
                cards = data.get("cards") or data.get("data") or []
                cursor = data.get("cursor")
                logger.warning(
                    f"WB content-probe POST {path}: keys={keys} error={data.get('error')} errorText={data.get('errorText')} "
                    f"additionalErrors={data.get('additionalErrors')} cards_len={len(cards) if isinstance(cards, list) else type(cards)} "
                    f"cursorKeys={list(cursor.keys()) if isinstance(cursor, dict) else cursor} payload={payload}"
                )
            except Exception as e:
                logger.warning(f"WB content-probe POST {path} exception: {e} payload={payload}")
                continue

    def get_all_cards(self, limit: int = 50) -> list[dict]:
        """Получить все карточки продавца через content-api с пагинацией cursor.

        Это надёжнее, чем vendorCodes/textSearch, когда vendorCode не совпадает со supplierArticle
        или фильтры возвращают пусто.
        """
        limit = max(1, min(int(limit), 50))
        all_cards: list[dict] = []
        cursor: dict = {"limit": limit}

        # Защита от бесконечного цикла
        for page in range(1, 2000):
            # Вариант 1: без filter
            payload = {"settings": {"cursor": cursor, "filter": {"withPhoto": -1}}}
            data = self._content_post("/content/v2/get/cards/list", payload, timeout=60)

            # Вариант 2: некоторые кабинеты требуют явный filter (пустой) / сортировку
            if data is None:
                payload2 = {"settings": {"cursor": cursor, "filter": {"withPhoto": -1}}}
                data = self._content_post("/content/v2/get/cards/list", payload2, timeout=60)

            if data is None:
                logger.warning(f"WB content-api get/cards/list(all): ответ None (cursor.limit={cursor.get('limit')})")
                break

            cards = (data.get("cards") or data.get("data") or [])
            if not isinstance(cards, list):
                logger.warning(
                    f"WB content-api get/cards/list(all): неожиданный формат cards/data={type(cards)} keys={list(data.keys())}"
                )
                break

            if not cards:
                # Это нормальный признак окончания пагинации
                break

            all_cards.extend(cards)

            # cursor может быть в разных местах
            next_cursor = data.get("cursor")
            if not isinstance(next_cursor, dict):
                # Некоторые версии возвращают cursor в data['cursor'] всегда, но если нет — прекращаем
                break

            # Если API возвращает те же значения — прекращаем
            if next_cursor == cursor:
                break
            cursor = next_cursor

            time.sleep(0.2)

        logger.info(f"WB content-api get_all_cards: получено карточек={len(all_cards)}")
        return all_cards

    def build_vendor_code_index(self, cards: list[dict]) -> dict[str, dict]:
        """Индекс по vendorCode для быстрого поиска карточки."""
        idx: dict[str, dict] = {}
        for c in cards or []:
            if not isinstance(c, dict):
                continue
            vc = str(c.get("vendorCode") or c.get("vendor_code") or "").strip()
            if vc:
                idx[vc] = c
        return idx

    def search_cards_by_text(self, query: str, limit: int = 100) -> list[dict]:
        """Fallback: поиск карточек через content-api по тексту.

        Когда vendorCodes из statistics-api не совпадают с vendorCode в content,
        можно попытаться найти карточку через textSearch.
        """
        q = str(query or "").strip()
        if not q:
            return []
        payload = {"settings": {"cursor": {"limit": int(limit)}, "filter": {"textSearch": q, "withPhoto": -1}}}
        data = self._content_post("/content/v2/get/cards/list", payload, timeout=60)
        cards = (((data or {}).get("cards")) or ((data or {}).get("data")) or [])
        return cards if isinstance(cards, list) else []

    def get_cards_by_vendor_codes(self, vendor_codes: list[str]) -> list[dict]:
        """Получает карточки по vendorCode.

        Используем /content/v2/get/cards/list. Возвращает список карточек.
        """
        vendor_codes = [str(x).strip() for x in (vendor_codes or []) if str(x).strip()]
        if not vendor_codes:
            return []

        # API ограничивает размер filter.vendorCodes; держим небольшой батч
        all_cards: list[dict] = []
        for i in range(0, len(vendor_codes), 100):
            part = vendor_codes[i:i + 100]
            payload = {
                "settings": {
                    "cursor": {"limit": 100},
                    "filter": {"vendorCodes": part, "withPhoto": -1},
                }
            }
            data = self._content_post("/content/v2/get/cards/list", payload, timeout=60)
            if data is None:
                logger.warning(
                    f"WB content-api get/cards/list: пустой ответ (None). "
                    f"vendorCodes(part)={len(part)} sample={part[:5]}"
                )
                continue

            cards = (((data or {}).get("cards")) or ((data or {}).get("data")) or [])
            if not isinstance(cards, list):
                logger.warning(
                    f"WB content-api get/cards/list: неожиданный формат cards/data: {type(cards)}. "
                    f"keys={list((data or {}).keys())} vendorCodes(part)={len(part)} sample={part[:5]}"
                )
                cards = []

            if not cards:
                # Если карточек нет, очень важно понять что вернул API
                logger.warning(
                    "WB content-api get/cards/list: карточки не найдены. "
                    f"keys={list((data or {}).keys())} vendorCodes(part)={len(part)} sample={part[:10]} "
                    f"error={data.get('error')} errorText={data.get('errorText') or data.get('message') or ''}"
                )
            else:
                all_cards.extend(cards)

            time.sleep(0.2)

        if not all_cards:
            logger.warning(
                f"WB content-api get/cards/list: итого 0 карточек по {len(vendor_codes)} vendorCodes. "
                f"Пробую fallback: загрузить все карточки и сматчить vendorCodes локально..."
            )
            all_seller_cards = self.get_all_cards(limit=50)
            if not all_seller_cards:
                logger.warning(
                    "WB content-api fallback get_all_cards: 0 карточек. "
                    "Похоже, токен не имеет доступа к карточкам (content-api) или у кабинета нет карточек."
                )
                return []

            index = self.build_vendor_code_index(all_seller_cards)
            matched = []
            for vc in vendor_codes:
                c = index.get(vc)
                if c:
                    matched.append(c)

            logger.warning(
                f"WB content-api fallback get_all_cards: всего карточек={len(all_seller_cards)}, "
                f"совпадений по vendorCodes={len(matched)}"
            )
            return matched

        return all_cards

    def get_object_charcs(self, subject_id: int) -> list[dict]:
        """/content/v2/object/charcs/{subjectId}"""
        data = self._content_get(f"/content/v2/object/charcs/{int(subject_id)}", timeout=60)
        return (data or {}).get("data") or []

    def get_fbo_stocks_v1(self):
        """Получает ВСЕ FBO-остатки через statistics-api"""
        all_stocks = []
        last_change_date = "2010-01-01T00:00:00"

        while True:
            # 🔥 ИСПРАВЛЕНО: удалены лишние пробелы в конце URL!
            url = "https://statistics-api.wildberries.ru/api/v1/supplier/stocks"
            params = {"dateFrom": last_change_date}

            try:
                response = requests.get(url, headers=self.headers, params=params, timeout=10)
                response.raise_for_status()
                logger.info(f"Запрос FBO остатков v1, статус={response.status_code}, dateFrom={last_change_date}")

                if response.status_code == 200:
                    data = response.json()
                    if not isinstance(data, list):
                        logger.error(f"Некорректный ответ (не список): {data}")
                        break

                    if not data:
                        logger.info("Получен пустой ответ — выгрузка завершена")
                        break

                    all_stocks.extend(data)
                    logger.info(f"Получено {len(data)} строк, всего: {len(all_stocks)}")

                    last_change_date = data[-1].get("lastChangeDate")
                    if not last_change_date:
                        break

                    time.sleep(1)
                else:
                    logger.error(f"Ошибка v1 stocks: {response.status_code} - {response.text}")
                    break

            except requests.exceptions.Timeout:
                logger.error(f"Таймаут при запросе FBO остатков (dateFrom={last_change_date})")
                break
            except requests.exceptions.RequestException as e:
                logger.error(f"Ошибка сети при запросе FBO остатков: {e}")
                break
            except Exception as e:
                logger.error(f"Неожиданная ошибка в v1 stocks: {e}", exc_info=True)
                break

        return all_stocks

    def get_card_by_nm_id(self, nm_id: int) -> dict | None:
        """Пытаемся получить карточку по nmId.

        В разных версиях API WB используются разные домены/пути.
        Пробуем несколько самых распространённых.
        """
        nm_id = int(nm_id)
        urls = [
            f"https://card.wb.ru/cards/v1/detail?appType=1&curr=rub&dest=-1257786&nm={nm_id}",
            f"https://card.wb.ru/cards/v2/detail?appType=1&curr=rub&dest=-1257786&nm={nm_id}",
        ]
        headers = {"User-Agent": "Mozilla/5.0"}
        for url in urls:
            try:
                # card.wb.ru не требует токен, но иногда режет; пробуем как есть
                resp = requests.get(url, timeout=30, headers=headers)
                if resp.status_code != 200:
                    continue
                data = resp.json() or {}
                return data
            except Exception:
                continue
        return None

    @staticmethod
    def extract_composition_from_card_api(payload: dict) -> str | None:
        """Извлекает состав из ответа card.wb.ru.

        Обычно данные лежат в data.products[0].properties / options / характеристиках.
        """
        if not isinstance(payload, dict):
            return None

        data = payload.get("data") or payload.get("Data") or {}
        products = data.get("products") or data.get("Products") or []
        if not products or not isinstance(products, list):
            return None
        p0 = products[0] if isinstance(products[0], dict) else None
        if not p0:
            return None

        # 1) properties: [{name:..., value:...}]
        for key in ("properties", "Properties"):
            props = p0.get(key)
            if isinstance(props, list):
                for pr in props:
                    if not isinstance(pr, dict):
                        continue
                    name = str(pr.get("name") or pr.get("Name") or "").strip().lower()
                    if "состав" in name:
                        val = pr.get("value") or pr.get("Value")
                        if isinstance(val, list):
                            txt = ", ".join([str(x).strip() for x in val if str(x).strip()])
                        else:
                            txt = str(val).strip() if val is not None else ""
                        if txt:
                            return txt

        # 2) options: [{name:..., value:...}]
        for key in ("options", "Options"):
            opts = p0.get(key)
            if isinstance(opts, list):
                for op in opts:
                    if not isinstance(op, dict):
                        continue
                    name = str(op.get("name") or op.get("Name") or "").strip().lower()
                    if "состав" in name:
                        val = op.get("value") or op.get("Value")
                        txt = str(val).strip() if val is not None else ""
                        if txt:
                            return txt

        # 3) всё подряд в характеристиках
        ch = p0.get("characteristics") or p0.get("Characteristics")
        if isinstance(ch, list):
            for op in ch:
                if not isinstance(op, dict):
                    continue
                name = str(op.get("name") or op.get("Name") or "").strip().lower()
                if "состав" in name:
                    val = op.get("value") or op.get("Value")
                    txt = str(val).strip() if val is not None else ""
                    if txt:
                        return txt
        return None

    @staticmethod
    def extract_color_from_content_card(card: dict) -> str | None:
        """Извлекает цвет из карточки content-api (/content/v2/get/cards/list).

        Ожидаем, что цвет лежит в списке characteristics как характеристика,
        у которой name содержит 'цвет'.
        """
        if not isinstance(card, dict):
            return None
        chars = card.get("characteristics")
        if isinstance(chars, list):
            for ch in chars:
                if not isinstance(ch, dict):
                    continue
                name = str(ch.get("name") or "").strip().lower()
                if "цвет" not in name:
                    continue
                for vk in ("value", "values", "valueName"):
                    v = ch.get(vk)
                    if isinstance(v, list):
                        vv = ", ".join([str(x).strip() for x in v if str(x).strip()])
                        if vv:
                            return vv
                    elif v is not None:
                        s = str(v).strip()
                        if s:
                            return s
        return None

    def get_cards_by_nm_ids(self, nm_ids: list[int]) -> list[dict]:
        """Получить карточки через content-api по nmId.

        WB часто не матчится по vendorCode/supplierArticle, зато в stocks есть nmId.
        Пробуем несколько вариантов ключей фильтра (API может быть разной версии).
        """
        nm_ids_clean: list[int] = []
        for x in nm_ids or []:
            try:
                nm_ids_clean.append(int(x))
            except Exception:
                continue
        nm_ids_clean = list(dict.fromkeys([x for x in nm_ids_clean if x > 0]))
        if not nm_ids_clean:
            return []

        all_cards: list[dict] = []
        for i in range(0, len(nm_ids_clean), 100):
            part = nm_ids_clean[i:i + 100]
            filter_variants = [
                {"nmIDs": part},
                {"nmIds": part},
                {"nmId": part[0]} if len(part) == 1 else None,
            ]

            got_any = False
            last_diag_payload = None

            for base_flt in [v for v in filter_variants if v is not None]:
                # Попробуем сначала withPhoto=-1 (как показала диагностика), но если не находит — без него.
                for with_photo in (True, False):
                    flt = dict(base_flt)
                    if with_photo:
                        flt["withPhoto"] = -1

                    cursor: dict = {"limit": 100}
                    page = 0
                    while True:
                        page += 1
                        payload = {"settings": {"cursor": cursor, "filter": flt}}
                        last_diag_payload = payload
                        data = self._content_post("/content/v2/get/cards/list", payload, timeout=60)
                        if data is None:
                            break

                        cards = (data.get("cards") or data.get("data") or [])
                        if isinstance(cards, list) and cards:
                            all_cards.extend(cards)
                            got_any = True

                        next_cursor = data.get("cursor")
                        if not isinstance(next_cursor, dict):
                            break
                        # окончание пагинации
                        if not cards:
                            break
                        if next_cursor == cursor:
                            break
                        cursor = next_cursor

                        # ограничим кол-во страниц на всякий случай
                        if page >= 200:
                            break

                        time.sleep(0.15)

                    if got_any:
                        break
                if got_any:
                    break

            if not got_any:
                logger.warning(
                    f"WB content-api get_cards_by_nm_ids: карточки не найдены. nmIds(part)={len(part)} sample={part[:5]} payload={last_diag_payload}"
                )

            time.sleep(0.2)

        # дедуп по nmID/vendorCode
        uniq = {}
        for c in all_cards:
            if not isinstance(c, dict):
                continue
            key = c.get('nmID') or c.get('nmId') or c.get('vendorCode') or id(c)
            uniq[str(key)] = c
        return list(uniq.values())


# ======================
# Нормализация и группировка
# ======================

def normalize_art(art_str):
    """Нормализует строку: приводит к нижнему регистру, удаляет лишние пробелы, очищает от невидимых символов"""
    if not art_str:
        return ""
    s = str(art_str)
    s = ''.join(c for c in s if c.isprintable())
    s = s.strip().lower()
    return s


def group_wb_remains_data(stock_data, template_id_to_cabinet_arts, template_id_to_name):
    """
    Группирует данные остатков WB по шаблонным артикулам.

    :param stock_data: dict {article: {"in_stock": ..., "in_way_from_client": ..., "in_way_to_client": ...}}
    :param template_id_to_cabinet_arts: dict {template_id: [cabinet_art1, cabinet_art2, ...]}
    :param template_id_to_name: dict {template_id: "Шаблонное название"}
    :return: grouped (по template_id), unmatched (артикулы без привязки)
    """
    stock_data_clean = {}
    for art, data in stock_data.items():
        clean_art = normalize_art(art)
        if clean_art:
            stock_data_clean[clean_art] = data

    cabinet_art_to_template_id = {}
    for template_id, arts in template_id_to_cabinet_arts.items():
        for art in arts:
            clean_art = normalize_art(art)
            if clean_art:
                cabinet_art_to_template_id[clean_art] = template_id

    grouped = {}
    unmatched = {}

    for clean_art, data in stock_data_clean.items():
        template_id = cabinet_art_to_template_id.get(clean_art)

        if template_id is not None:
            if template_id not in grouped:
                grouped[template_id] = {
                    'name': template_id_to_name.get(template_id, f"ID {template_id}"),
                    'in_stock': 0,
                    'in_way_from_client': 0,
                    'in_way_to_client': 0
                }
            grouped[template_id]['in_stock'] += data['in_stock']
            grouped[template_id]['in_way_from_client'] += data['in_way_from_client']
            grouped[template_id]['in_way_to_client'] += data['in_way_to_client']
        else:
            unmatched[clean_art] = {
                'name': f"НЕОПОЗНАННЫЙ: {clean_art}",
                'in_stock': data['in_stock'],
                'in_way_from_client': data['in_way_from_client'],
                'in_way_to_client': data['in_way_to_client']
            }

    return grouped, unmatched


# ======================
# Обработчики
# ======================

async def start_wb_remains(update: Update, context: CallbackContext) -> int:
    """Начало — выбор кабинета Wildberries"""
    context.user_data['current_flow'] = 'wb_remains'

    keyboard = [
        [InlineKeyboardButton("🏪 WB_1 Nimba", callback_data='wb_cabinet_1')],
        [InlineKeyboardButton("🏬 WB_2 Galioni", callback_data='wb_cabinet_2')],
        [InlineKeyboardButton("🏢 WB_3 AGNIA", callback_data='wb_cabinet_3')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        "🏢 Выберите кабинет Wildberries для выгрузки остатков:",
        reply_markup=reply_markup
    )

    return WB_REMAINS_CABINET_CHOICE


async def handle_wb_cabinet_choice(update: Update, context: CallbackContext) -> int:
    """Обработка выбора кабинета WB — генерация отчёта"""
    query = update.callback_query
    await query.answer()

    cabinet_data = query.data
    if cabinet_data == 'wb_cabinet_1':
        cabinet_id = 1
        cabinet_name = "WB_1 Nimba"
        sheet_name = "Отдельно ВБ Nimba"
    elif cabinet_data == 'wb_cabinet_2':
        cabinet_id = 2
        cabinet_name = "WB_2 Galioni"
        sheet_name = "Отдельно ВБ Galioni"
    elif cabinet_data == 'wb_cabinet_3':
        cabinet_id = 3
        cabinet_name = "WB_3 AGNIA"
        sheet_name = "Отдельно ВБ AGNIA"
    else:
        await query.message.reply_text("❌ Неизвестный кабинет.")
        return ConversationHandler.END

    context.user_data['wb_cabinet_id'] = cabinet_id

    loading_msg1 = await query.message.edit_text(f"⏳ Получаю остатки с Wildberries API ({cabinet_name})...")
    context.user_data['wb_remains_loading_msg1_id'] = loading_msg1.message_id

    try:
        wb_api = WildberriesAPI(cabinet_id=cabinet_id)

        # === Диагностика content-api / stocks (как в Ozon: максимум фактов в лог) ===
        try:
            hc = wb_api.content_health_check()
            logger.warning(f"WB content-api health-check(before stocks): {hc}")
        except Exception as e:
            logger.warning(f"WB content-api health-check(before stocks) error: {e}")

        loading_msg2 = await query.message.reply_text("📊 Запрашиваю остатки по товарам...")
        context.user_data['wb_remains_loading_msg2_id'] = loading_msg2.message_id
        stocks = wb_api.get_fbo_stocks_v1()

        # Подробная диагностика первых строк stocks
        try:
            sample_rows = (stocks or [])[:2]
            for i, it in enumerate(sample_rows, start=1):
                if isinstance(it, dict):
                    logger.warning(
                        f"WB stocks sample #{i}: keys={list(it.keys())} "
                        f"supplierArticle={it.get('supplierArticle')} nmId={it.get('nmId') or it.get('nmID')} "
                        f"imtId={it.get('imtId') or it.get('imtID')} barcode={it.get('barcode')} chrtId={it.get('chrtId') or it.get('chrtID')}"
                    )
                    # чтобы не засорять лог слишком сильно — полный dict только один раз
                    if i == 1:
                        logger.warning(f"WB stocks sample #1 full: {it}")
        except Exception as e:
            logger.warning(f"WB stocks sample log error: {e}")

        if not stocks:
            await query.message.reply_text(
                "ℹ️ Остатки не найдены. Возможные причины:\n"
                "• У вас нет товаров на складах Wildberries (FBO)\n"
                "• Токен не имеет доступа к остаткам",
                reply_markup=ReplyKeyboardRemove()
            )
            return ConversationHandler.END

        # === 1. Сырые данные ===
        raw_data = []
        stock_dict = {}
        category_by_article = {}

        size_by_article: dict[str, str] = {}
        color_by_article: dict[str, str] = {}
        # Собираем vendor_code(=supplierArticle) список для content-api

        vendor_codes: list[str] = []

        for item in stocks:
            vendor_code = item.get("supplierArticle")
            if not vendor_code:
                continue

            if len(vendor_codes) < 5000:  # простая защита
                vendor_codes.append(str(vendor_code).strip())

            article = clean_article(vendor_code)
            if not article:
                continue

            # Размер чаще всего в stocks.techSize
            tech_size = item.get("techSize")
            if tech_size is not None and str(tech_size).strip():
                size_by_article[article] = normalize_wb_size(tech_size)

            # Иногда WB отдаёт цвет прямо в stocks
            for ck in ("color", "Color", "цвет", "Цвет"):
                cv = item.get(ck)
                if cv is not None and str(cv).strip():
                    color_by_article[article] = str(cv).strip()
                    break

            category = item.get("subject") or item.get("category") or "—"
            if article not in category_by_article and category:
                category_by_article[article] = str(category).strip() if str(category).strip() else "—"

            if article not in stock_dict:
                stock_dict[article] = {
                    'in_stock': 0,
                    'in_way_to_client': 0,
                    'in_way_from_client': 0
                }

            stock_dict[article]['in_stock'] += item.get('quantity', 0)
            stock_dict[article]['in_way_to_client'] += item.get('inWayToClient', 0)
            stock_dict[article]['in_way_from_client'] += item.get('inWayFromClient', 0)

        # === 1.1. Состав материала через nmId (card API) ===
        # Кеши в рамках одного запуска
        charcs_cache: dict[int, list[dict]] = {}
        composition_charc_id_by_subject: dict[int, int] = {}
        composition_by_vendor_code: dict[str, str] = {}
        composition_by_nm_id: dict[int, str] = {}

        # Соберём nmId по supplierArticle
        nm_id_by_article: dict[str, int] = {}
        for it in stocks or []:
            art = clean_article(it.get("supplierArticle"))
            if not art:
                continue
            nm = it.get("nmId") or it.get("nmID")
            try:
                if nm is not None:
                    nm_id_by_article[art] = int(nm)
            except Exception:
                pass

        # Диагностически попробуем вытащить состав по nmId для первых N товаров
        try:
            t_nm = time.time()
            found_nm = 0
            checked = 0
            for art, nm in list(nm_id_by_article.items())[:60]:
                checked += 1
                if nm in composition_by_nm_id:
                    continue
                payload = wb_api.get_card_by_nm_id(nm)
                comp = wb_api.extract_composition_from_card_api(payload or {})
                if comp:
                    composition_by_nm_id[nm] = comp
                    found_nm += 1
                time.sleep(0.03)
            logger.warning(f"WB nmId composition: checked={checked} found={found_nm} за {time.time()-t_nm:.2f}s")
        except Exception as e:
            logger.warning(f"WB nmId composition error: {e}")

        # Запрашиваем карточки батчами и заполняем цвет (content-api)
        try:
            # Лучше матчиться по nmId
            nm_ids = list({nm for nm in nm_id_by_article.values() if nm and int(nm) > 0})
            t_cards = time.time()
            cards = wb_api.get_cards_by_nm_ids(nm_ids)
            logger.info(f"WB: get_cards_by_nm_ids карточек={len(cards or [])} за {time.time()-t_cards:.2f}s")

            by_nm: dict[int, dict] = {}
            for c in (cards or []):
                try:
                    nm = int(c.get("nmID") or c.get("nmId"))
                except Exception:
                    continue
                by_nm[nm] = c

            # Заполним цвета по артикулу через nmId
            for art, nm in nm_id_by_article.items():
                if art in color_by_article and str(color_by_article.get(art)).strip() and color_by_article[art] != '—':
                    continue
                card = by_nm.get(nm)
                if not card:
                    continue
                cval = wb_api.extract_color_from_content_card(card)
                if cval:
                    color_by_article[art] = cval
        except Exception as e:
            logger.warning(f"Не удалось получить цвет WB через content-api: {e}")

        for article, counts in stock_dict.items():
            total = (
                    counts['in_stock'] +
                    counts['in_way_to_client'] +
                    counts['in_way_from_client']
            )
            raw_data.append({
                'Категория': category_by_article.get(article, '—'),
                'Артикул': article,
                'Доступно на складах': counts['in_stock'],
                'Возвращаются от покупателей': counts['in_way_from_client'],
                'В пути до покупателей': counts['in_way_to_client'],
                'Итого на МП': total
            })

        df_raw = pd.DataFrame(raw_data).sort_values(by='Артикул').reset_index(drop=True)
        headers_raw = ["Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей", "В пути до покупателей", "Итого на МП"]

        # === 2. Группировка по шаблону Nimba/Galioni ===
        template_id_to_name, template_id_to_cabinet_arts = get_cabinet_articles_by_template_id(sheet_name)

        linked_template_ids = set(template_id_to_cabinet_arts.keys())

        # Получаем main_ids_ordered — ID в порядке появления в Excel (без дубликатов)
        template_path = os.path.join(root_dir, "База данных артикулов для выкупов и начислений.xlsx")
        if not os.path.exists(template_path):
            template_path = "База данных артикулов для выкупов и начислений.xlsx"
        df_order = pd.read_excel(template_path, sheet_name=sheet_name)
        main_ids_ordered = []
        seen = set()
        for _, row in df_order.iterrows():
            if not pd.isna(row.get('ID')):
                tid = int(row['ID'])
                if tid not in seen:
                    main_ids_ordered.append(tid)
                    seen.add(tid)

        cabinet_arts_set = set()
        for arts in template_id_to_cabinet_arts.values():
            for art in arts:
                cabinet_arts_set.add(normalize_art(art))

        template_rows_to_color = []
        for idx, id_val in enumerate(main_ids_ordered, start=3):
            if id_val in linked_template_ids:
                template_rows_to_color.append(idx)

        wb_stock_data = {}
        for art, counts in stock_dict.items():
            wb_stock_data[art] = {
                "in_stock": counts['in_stock'],
                "in_way_from_client": counts['in_way_from_client'],
                "in_way_to_client": counts['in_way_to_client']
            }

        grouped, unmatched = group_wb_remains_data(wb_stock_data, template_id_to_cabinet_arts, template_id_to_name)

        template_data = []
        for id_val in main_ids_ordered:
            if id_val in grouped:
                d = grouped[id_val]
                total = d['in_stock'] + d['in_way_from_client'] + d['in_way_to_client']
                template_data.append({
                    'Артикул': d['name'],
                    'Доступно на складах': d['in_stock'],
                    'Возвращаются от покупателей': d['in_way_from_client'],
                    'В пути до покупателей': d['in_way_to_client'],
                    'Итого на МП': total
                })
            else:
                name = template_id_to_name.get(id_val, f"ID {id_val}")
                template_data.append({
                    'Артикул': name,
                    'Доступно на складах': 0,
                    'Возвращаются от покупателей': 0,
                    'В пути до покупателей': 0,
                    'Итого на МП': 0
                })

        for art, d in unmatched.items():
            total = d['in_stock'] + d['in_way_from_client'] + d['in_way_to_client']
            template_data.append({
                'Артикул': f"НЕОПОЗНАННЫЙ: {art}",
                'Доступно на складах': d['in_stock'],
                'Возвращаются от покупателей': d['in_way_from_client'],
                'В пути до покупателей': d['in_way_to_client'],
                'Итого на МП': total
            })

        df_template = pd.DataFrame(template_data)
        headers_template = ["Артикул", "Доступно на складах", "Возвращаются от покупателей", "В пути до покупателей", "Итого на МП"]

        thresholds = resolve_stock_thresholds(context, query.message.chat_id)
        raw_rows_to_color = []
        for idx, art in enumerate(df_raw["Артикул"], start=3):
            if normalize_art(art) in cabinet_arts_set:
                raw_rows_to_color.append(idx)

        # === Сводка ===
        total_in_stock = sum(d['in_stock'] for d in stock_dict.values())
        total_in_way_from = sum(d['in_way_from_client'] for d in stock_dict.values())
        total_in_way_to = sum(d['in_way_to_client'] for d in stock_dict.values())
        total_mp = total_in_stock + total_in_way_from + total_in_way_to

        def fmt_num(x):
            return f"{x:,}".replace(",", " ")

        summary_text = (
            f"📊 <b>Сводка по остаткам Wildberries (FBO)</b>\n"
            f"Кабинет: <b>{cabinet_name}</b>\n\n"
            f"📦 <b>Доступно на складах:</b> {fmt_num(total_in_stock)} шт\n"
            f"↩️ <b>Возвращаются от покупателей:</b> {fmt_num(total_in_way_from)} шт\n"
            f"🚚 <b>В пути до покупателей:</b> {fmt_num(total_in_way_to)} шт\n"
            f"✅ <b>Итого на МП:</b> {fmt_num(total_mp)} шт"
        )

        # === Создаём Excel с двумя листами ===
        report_path = f"WB_Remains_Report_Cabinet{cabinet_id}.xlsx"
        create_excel_with_two_sheets(
            df_raw,
            headers_raw,
            df_template,
            headers_template,
            report_path,
            thresholds=thresholds,
            template_rows_to_color=template_rows_to_color,
            raw_rows_to_color=raw_rows_to_color
        )

        # === Отправляем ===
        await query.message.reply_document(
            document=open(report_path, 'rb'),
            caption=f"📊 Отчёт по остаткам Wildberries: {cabinet_name}",
            reply_markup=ReplyKeyboardRemove()
        )
        await query.message.reply_text(summary_text, parse_mode="HTML")

        # === Очистка ===
        if os.path.exists(report_path):
            os.remove(report_path)

        # Удаляем служебные сообщения
        chat_id = query.message.chat_id
        try:
            msg1_id = context.user_data.get('wb_remains_loading_msg1_id')
            if msg1_id:
                await context.bot.delete_message(chat_id=chat_id, message_id=msg1_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить первое сообщение о загрузке WB: {e}")

        try:
            msg2_id = context.user_data.get('wb_remains_loading_msg2_id')
            if msg2_id:
                await context.bot.delete_message(chat_id=chat_id, message_id=msg2_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить второе сообщение о загрузке WB: {e}")

    except Exception as e:
        logger.error(f"Ошибка при получении остатков WB (кабинет {cabinet_id}): {str(e)}", exc_info=True)
        await query.message.reply_text(f"❌ Ошибка: {str(e)}", reply_markup=ReplyKeyboardRemove())
        # Удаляем служебные сообщения даже при ошибке
        chat_id = query.message.chat_id
        try:
            msg1_id = context.user_data.get('wb_remains_loading_msg1_id')
            if msg1_id:
                await context.bot.delete_message(chat_id=chat_id, message_id=msg1_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить первое сообщение о загрузке WB при ошибке: {e}")

        try:
            msg2_id = context.user_data.get('wb_remains_loading_msg2_id')
            if msg2_id:
                await context.bot.delete_message(chat_id=chat_id, message_id=msg2_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить второе сообщение о загрузке WB при ошибке: {e}")

    return ConversationHandler.END


def create_excel_with_two_sheets(
        df_raw,
        headers_raw,
        df_template,
        headers_template,
        filename,
        thresholds=None,
        template_rows_to_color=None,
        raw_rows_to_color=None
):
    """Создаёт Excel с двумя листами"""
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet(title="Остатки шаблон Nimba")
    _write_sheet(ws1, df_template, headers_template, has_name=False)
    if template_rows_to_color and thresholds:
        apply_fill_to_cells(ws1, template_rows_to_color, [5], thresholds)

    ws2 = wb.create_sheet(title="Остатки исходные артикулы")
    _write_sheet(ws2, df_raw, headers_raw, has_name=False)
    if raw_rows_to_color and thresholds:
        total_col = headers_raw.index("Итого на МП") + 1
        apply_fill_to_cells(ws2, raw_rows_to_color, [total_col], thresholds)

    wb.save(filename)


def _write_sheet(ws, df, headers, has_name):
    """Вспомогательная функция для записи одного листа"""
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    ws.merge_cells('A1:A2')

    data_start_row = 3
    sum_row = 2

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), data_start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border

    num_rows = len(df)
    if num_rows > 0:
        start_col_index = 2
        for col in range(start_col_index, len(headers) + 1):
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_start_row + num_rows - 1})"
            cell = ws.cell(row=sum_row, column=col, value=formula)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


# ======================
# Заглушки для совместимости
# ======================

async def handle_wb_remains_files(update: Update, context: CallbackContext):
    await update.message.reply_text("Файлы не требуются.")
    return ConversationHandler.END
