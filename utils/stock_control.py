from openpyxl.styles import PatternFill

from utils.settings_manager import get_stock_thresholds

FILL_GREEN = PatternFill(fill_type="solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill(fill_type="solid", fgColor="FFEB9C")
FILL_RED = PatternFill(fill_type="solid", fgColor="FFC7CE")


def resolve_stock_thresholds(context, chat_id):
    thresholds = None
    if context is not None and hasattr(context, "user_data") and context.user_data is not None:
        thresholds = context.user_data.get("stock_thresholds")
    if not thresholds:
        thresholds = get_stock_thresholds(chat_id)
        if thresholds and context is not None and hasattr(context, "user_data") and context.user_data is not None:
            context.user_data["stock_thresholds"] = thresholds
    return thresholds


def get_fill_for_value(value, thresholds):
    if not thresholds:
        return None
    try:
        val = int(float(value))
    except (ValueError, TypeError):
        return None

    red_level = thresholds.get("red")
    yellow_level = thresholds.get("yellow")
    if red_level is None or yellow_level is None:
        return None

    if val < red_level:
        return FILL_RED
    if val <= yellow_level:
        return FILL_YELLOW
    return FILL_GREEN


def apply_fill_to_cells(ws, rows, cols, thresholds):
    if not thresholds:
        return
    for row in rows:
        for col in cols:
            cell = ws.cell(row=row, column=col)
            fill = get_fill_for_value(cell.value, thresholds)
            if fill:
                cell.fill = fill
