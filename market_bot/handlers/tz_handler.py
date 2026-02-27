import os
import re
import time
import zipfile
import shutil
import logging
import tempfile
from dataclasses import dataclass
from datetime import datetime
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from telegram import (
    Update,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    ReplyKeyboardRemove,
)
from telegram.ext import CallbackContext

from states import SELECTING_ACTION, TZ_CABINET_SELECT, TZ_WAIT_FILE
from utils.menu import get_main_menu
from utils.template_loader import get_cabinet_articles_by_template_id
from utils.settings_manager import get_stock_thresholds

from handlers.all_mp_remains_handler import fetch_ozon_remains_raw, fetch_wb_remains_raw
from handlers.ozon_remains_handler import OzonAPI, clean_offer_id
from handlers.wb_remains_handler import WildberriesAPI, clean_article, normalize_wb_size
from utils.ozon_attributes import extract_attribute_values_from_product_attributes

logger = logging.getLogger(__name__)

TELEGRAM_SAFE_MESSAGE_LEN = 3800


def _split_message(text: str, limit: int = TELEGRAM_SAFE_MESSAGE_LEN) -> list[str]:
    """Разбивает длинный текст на части для Telegram (лимит ~4096)."""
    text = str(text or "").strip()
    if not text:
        return []

    parts: list[str] = []
    buf: list[str] = []
    cur = 0

    for line in text.splitlines():
        chunk = (line + "\n")
        if cur + len(chunk) > limit and buf:
            parts.append("".join(buf).rstrip())
            buf = []
            cur = 0
        # если одна строка слишком длинная — режем её
        while len(chunk) > limit:
            head = chunk[:limit]
            parts.append(head.rstrip())
            chunk = chunk[limit:]
        buf.append(chunk)
        cur += len(chunk)

    if buf:
        parts.append("".join(buf).rstrip())

    return [p for p in parts if p.strip()]


async def _send_long_text(update: Update, text: str, filename: str | None = None) -> None:
    """Отправляет длинный текст сообщениями; при необходимости — файлом."""
    parts = _split_message(text)
    if not parts:
        return

    # если частей слишком много — лучше одним txt
    if len(parts) > 8 and filename:
        try:
            tmp_dir = tempfile.mkdtemp(prefix="tz_msg_")
            path = os.path.join(tmp_dir, filename)
            with open(path, "w", encoding="utf-8") as f:
                f.write(text)
            await update.message.reply_document(document=open(path, "rb"), caption="📄 Отчёт по позициям, не попавшим в ТЗ")
            shutil.rmtree(tmp_dir, ignore_errors=True)
            return
        except Exception:
            # запасной вариант: шлём кусками
            pass

    for i, p in enumerate(parts, start=1):
        prefix = "" if len(parts) == 1 else f"(часть {i}/{len(parts)})\n"
        await update.message.reply_text(prefix + p)


def _chunk_list(lst: list, n: int):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


@dataclass(frozen=True)
class Cabinet:
    mp: str  # 'ozon' | 'wb'
    cabinet_id: int

    @property
    def key(self) -> str:
        return f"{self.mp}:{self.cabinet_id}"

    @property
    def label(self) -> str:
        if self.mp == "ozon":
            return {1: "Ozon_1 Nimba", 2: "Ozon_2 Galioni", 3: "Ozon_3 AGNIA"}.get(self.cabinet_id, f"Ozon_{self.cabinet_id}")
        if self.mp == "wb":
            return {1: "WB_1 Nimba", 2: "WB_2 Galioni", 3: "WB_3 AGNIA"}.get(self.cabinet_id, f"WB_{self.cabinet_id}")
        return self.key


ALL_CABINETS: list[Cabinet] = [
    Cabinet("ozon", 1),
    Cabinet("ozon", 2),
    Cabinet("ozon", 3),
    Cabinet("wb", 1),
    Cabinet("wb", 2),
    Cabinet("wb", 3),
]


SELLER_BY_CABINET: dict[str, str] = {
    "ozon:2": "ИП Гребнев А. А.",
    "wb:1": "ИП Гребнев А. А.",

    "ozon:1": "ИП Снигирева Г. М.",
    "wb:2": "ИП Снигирева Г. М.",

    "ozon:3": "ООО \"АГНИЯ\"",
    "wb:3": "ООО \"АГНИЯ\"",
}


def _safe_filename(s: str) -> str:
    s = str(s or "").strip()
    s = re.sub(r"[\\/*?:\"<>|]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip() or "file"


def _normalize_name(s: str) -> str:
    return str(s or "").strip().lower()


def _normalize_article_key(s: str) -> str:
    return str(s or "").strip().lower()


def _format_selected(selected: set[str]) -> str:
    if not selected:
        return "(ничего не выбрано)"
    labs = []
    for c in ALL_CABINETS:
        if c.key in selected:
            labs.append(c.label)
    return ", ".join(labs) if labs else "(ничего не выбрано)"


def _build_cabinet_select_keyboard(selected: set[str]) -> InlineKeyboardMarkup:
    rows = []
    for c in ALL_CABINETS:
        mark = "✅" if c.key in selected else "☑️"
        rows.append([InlineKeyboardButton(f"{mark} {c.label}", callback_data=f"tz_toggle:{c.key}")])

    rows.append([
        InlineKeyboardButton("Готово", callback_data="tz_done"),
        InlineKeyboardButton("Сбросить", callback_data="tz_reset"),
    ])
    rows.append([InlineKeyboardButton("⬅️ Назад", callback_data="tz_back")])
    return InlineKeyboardMarkup(rows)


async def start_tz_generation(update: Update, context: CallbackContext) -> int:
    context.user_data.pop("tz_selected", None)
    selected: set[str] = set()

    await update.message.reply_text(
        "📝 Формирование ТЗ\n\n"
        "Выберите один или несколько кабинетов, по которым нужно сделать ТЗ.",
        reply_markup=_build_cabinet_select_keyboard(selected),
    )
    context.user_data["tz_selected"] = selected
    return TZ_CABINET_SELECT


async def handle_tz_cabinet_select(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()

    selected: set[str] = set(context.user_data.get("tz_selected") or set())

    data = query.data or ""
    if data.startswith("tz_toggle:"):
        key = data.split(":", 1)[1]
        if key in selected:
            selected.remove(key)
        else:
            selected.add(key)
        context.user_data["tz_selected"] = selected

        await query.message.edit_text(
            "📝 Формирование TЗ\n\n"
            f"Выбрано: {_format_selected(selected)}\n\n"
            "Выберите кабинеты:",
            reply_markup=_build_cabinet_select_keyboard(selected),
        )
        return TZ_CABINET_SELECT

    if data == "tz_reset":
        selected.clear()
        context.user_data["tz_selected"] = selected
        await query.message.edit_text(
            "📝 Формирование ТЗ\n\nВыбор сброшен. Выберите кабинеты:",
            reply_markup=_build_cabinet_select_keyboard(selected),
        )
        return TZ_CABINET_SELECT

    if data == "tz_back":
        await query.message.reply_text("Возвращаю в главное меню.", reply_markup=get_main_menu())
        return SELECTING_ACTION

    if data == "tz_done":
        if not selected:
            await query.message.reply_text("⚠️ Сначала выберите хотя бы один кабинет.")
            return TZ_CABINET_SELECT

        await query.message.reply_text(
            "📎 Теперь загрузите Excel-файл со столбцами:\n"
            "• Наименование (шаблонный артикул из базы)\n"
            "• Количество\n\n"
            "После загрузки я сформирую ТЗ для выбранных кабинетов.",
            reply_markup=ReplyKeyboardRemove(),
        )
        return TZ_WAIT_FILE

    await query.message.reply_text("⚠️ Неизвестная команда.")
    return TZ_CABINET_SELECT


def _read_user_excel(path: str) -> list[dict]:
    df = pd.read_excel(path)
    cols = {str(c).strip().lower(): c for c in df.columns}
    if "наименование" not in cols or "количество" not in cols:
        raise ValueError("В файле должны быть столбцы 'Наименование' и 'Количество'.")

    name_col = cols["наименование"]
    qty_col = cols["количество"]

    rows: list[dict] = []
    for _, r in df.iterrows():
        name = r.get(name_col)
        qty = r.get(qty_col)
        if pd.isna(name) or str(name).strip() == "":
            continue
        try:
            qty_int = int(float(str(qty).strip()))
        except Exception:
            continue
        if qty_int <= 0:
            continue
        rows.append({"name": str(name).strip(), "qty": qty_int})

    if not rows:
        raise ValueError("Файл пустой или не содержит корректных строк.")

    agg: dict[str, int] = {}
    orig: dict[str, str] = {}
    for it in rows:
        k = _normalize_name(it["name"])
        agg[k] = agg.get(k, 0) + int(it["qty"])
        orig.setdefault(k, it["name"])

    return [{"name": orig[k], "qty": agg[k], "key": k} for k in agg]


async def handle_tz_file(update: Update, context: CallbackContext) -> int:
    doc = update.message.document
    if not doc:
        await update.message.reply_text("Пришлите Excel-файл .xlsx")
        return TZ_WAIT_FILE

    if not (doc.file_name or "").lower().endswith(".xlsx"):
        await update.message.reply_text("Поддерживается только .xlsx")
        return TZ_WAIT_FILE

    tmp_dir = tempfile.mkdtemp(prefix="tz_")
    local_path = os.path.join(tmp_dir, doc.file_name)
    f = await doc.get_file()
    await f.download_to_drive(custom_path=local_path)

    context.user_data["tz_tmp_dir"] = tmp_dir

    try:
        rows = _read_user_excel(local_path)
    except Exception as e:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        context.user_data.pop("tz_tmp_dir", None)
        await update.message.reply_text(f"❌ Не удалось прочитать файл: {e}")
        return SELECTING_ACTION

    selected: set[str] = set(context.user_data.get("tz_selected") or set())

    status = await update.message.reply_text("⏳ Собираю остатки и формирую ТЗ. Это может занять несколько минут...")

    try:
        zip_path, not_added_msg = await _generate_tz_zip(context, update.effective_chat.id, selected, rows)
    except Exception as e:
        logger.error(f"Ошибка формирования ТЗ: {e}", exc_info=True)
        await status.edit_text(f"❌ Ошибка формирования ТЗ: {e}")
        _cleanup_tz_tmp(context)
        return SELECTING_ACTION

    await status.edit_text("✅ ТЗ сформировано. Отправляю архив...")
    await update.message.reply_document(document=open(zip_path, "rb"), caption="📦 ТЗ (архив)")

    if not_added_msg:
        await _send_not_added_report(update, not_added_msg, filename="tz_not_added.txt")

    _cleanup_tz_tmp(context)
    return SELECTING_ACTION


def _cleanup_tz_tmp(context: CallbackContext) -> None:
    tmp_dir = context.user_data.pop("tz_tmp_dir", None)
    if tmp_dir and os.path.exists(tmp_dir):
        shutil.rmtree(tmp_dir, ignore_errors=True)
    context.user_data.pop("tz_selected", None)


def _needs_restock(qty: int, thresholds: dict | None) -> bool:
    """Требование для ТЗ: попадать должны все артикулы, у которых остаток <= yellow (включительно)."""
    if not thresholds:
        return False
    try:
        yellow = int(thresholds.get('yellow', 0))
    except Exception:
        yellow = 0
    if yellow <= 0:
        return False
    try:
        return int(qty) <= yellow
    except Exception:
        return False


def _stocks_level(qty: int, thresholds: dict | None) -> str:
    if not thresholds:
        return "unknown"
    red = int(thresholds.get("red", 0))
    yellow = int(thresholds.get("yellow", 0))
    # критический уровень: <= red
    if qty <= red:
        return "red"
    # предупредительный уровень: > red и < yellow
    if qty < yellow:
        return "yellow"
    return "green"


def _is_danger_level(level: str) -> bool:
    return level in ("red", "yellow")


def _distribute_qty_equalize_final(items: list[dict], need: int, allowed_keys: set[tuple[str, str]] | None = None) -> dict[tuple[str, str], int]:
    """Распределение по принципу выравнивания ИТОГОВЫХ остатков: (stock + ship) примерно равны.

    На каждом шаге добавляем 1 шт тому, у кого сейчас минимальный final.
    allowed_keys: если задан, распределяем только по этим (cabinet, article).
    """
    if need <= 0 or not items:
        return {}

    by_key = {(it['cabinet_key'], it['article']): int(it.get('stock') or 0) for it in items}
    if allowed_keys is not None:
        by_key = {k: v for k, v in by_key.items() if k in allowed_keys}
    if not by_key:
        return {}

    ship = {k: 0 for k in by_key}

    keys = list(by_key.keys())
    for _ in range(int(need)):
        # выбираем минимальный final
        best = None
        best_final = None
        for k in keys:
            final = by_key[k] + ship[k]
            if best is None or final < best_final:
                best = k
                best_final = final
        if best is None:
            break
        ship[best] += 1

    return {k: v for k, v in ship.items() if v > 0}


def _distribute_qty_equalize_across(items: list[dict], need: int) -> dict[tuple[str, str], int]:
    """Back-compat alias: теперь распределяем по выравниванию final (stock+ship)."""
    return _distribute_qty_equalize_final(items, need)


def _ozon_extract_offer_id_to_attrs(cabinet_id: int) -> dict[str, dict]:
    """
    Тянем для Ozon offer_id -> {category, barcode, color, size, composition}.

    Используем:
    - /v1/description-category/tree (кеш в cache/ozon)
    - /v1/description-category/attribute (кеш в cache/ozon)
    - /v4/product/info/attributes
    """
    ozon = OzonAPI(cabinet_id=cabinet_id)

    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    cache_dir = os.path.join(root_dir, "cache", "ozon")
    os.makedirs(cache_dir, exist_ok=True)

    def _cache_read_json(path: str) -> dict | None:
        import json
        try:
            if not os.path.exists(path):
                return None
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None

    def _cache_write_json(path: str, data: dict) -> None:
        import json
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False)
        except Exception:
            pass

    # 1) offer_ids + meta (dcid/type/product_id)
    product_list = ozon.get_product_list(limit=1000)
    items = (product_list or {}).get('result', {}).get('items', [])
    offer_ids: list[str] = []
    for it in items or []:
        oid = clean_offer_id(it.get('offer_id'))
        if oid:
            offer_ids.append(oid)

    offer_id_to_product_id: dict[str, int] = {}
    offer_id_to_dcid: dict[str, int] = {}
    offer_id_to_type_id: dict[str, int] = {}

    # Используем локальный чанкер, чтобы не зависеть от импорта из ozon_remains_handler.
    for chunk in _chunk_list(offer_ids, 1000):
        resp = ozon.get_product_info_list(offer_ids=chunk)
        its = []
        if 'result' in (resp or {}) and 'items' in (resp or {}).get('result', {}):
            its = resp['result']['items']
        elif 'items' in (resp or {}):
            its = resp['items']
        elif isinstance((resp or {}).get('result'), list):
            its = resp['result']
        for info in its or []:
            oid = clean_offer_id(info.get('offer_id'))
            if not oid:
                continue
            pid = info.get('id') or info.get('product_id')
            if pid is not None:
                try:
                    offer_id_to_product_id[oid] = int(pid)
                except Exception:
                    pass
            dcid = info.get('description_category_id')
            if dcid is not None:
                try:
                    offer_id_to_dcid[oid] = int(dcid)
                except Exception:
                    pass
            tpid = info.get('type_id')
            if tpid is not None:
                try:
                    offer_id_to_type_id[oid] = int(tpid)
                except Exception:
                    pass

    # 2) дерево категорий: строим (dcid,type_id)->type_name
    type_name_by_pair: dict[tuple[int, int], str] = {}
    tree_cache_path = os.path.join(cache_dir, 'description_category_tree_DEFAULT.json')
    tree = _cache_read_json(tree_cache_path)
    if not tree:
        tree = ozon.get_description_category_tree(language='DEFAULT')
        if tree:
            _cache_write_json(tree_cache_path, tree)

    def _walk(nodes, parent_dcid=None):
        for n in nodes or []:
            dcid = n.get('description_category_id', parent_dcid)
            tpid = n.get('type_id')
            tname = n.get('type_name')
            if dcid is not None and tpid is not None and tname:
                try:
                    type_name_by_pair[(int(dcid), int(tpid))] = str(tname).strip()
                except Exception:
                    pass
            _walk(n.get('children') or [], dcid)

    if tree and tree.get('result'):
        _walk(tree.get('result'), None)

    # 3) meta attributes per (dcid,type_id) to resolve attr ids
    meta_cache: dict[tuple[int, int], list[dict]] = {}

    def _ozon_post(path: str, payload: dict) -> dict | None:
        import requests
        try:
            resp = requests.post(f"{ozon.base_url}{path}", json=payload, headers=ozon.headers, timeout=60)
            if resp.status_code != 200:
                return None
            return resp.json() or {}
        except Exception:
            return None

    def _get_meta(dcid: int, tpid: int) -> list[dict]:
        key = (int(dcid), int(tpid))
        if key in meta_cache:
            return meta_cache[key]
        cache_path = os.path.join(cache_dir, f"attrs_{key[0]}_{key[1]}_DEFAULT.json")
        cached = _cache_read_json(cache_path)
        if cached and isinstance(cached.get('result'), list):
            meta_cache[key] = cached['result']
            return meta_cache[key]
        resp = _ozon_post('/v1/description-category/attribute', {
            'description_category_id': key[0],
            'type_id': key[1],
            'language': 'DEFAULT'
        })
        meta_cache[key] = (resp or {}).get('result') or []
        if resp:
            _cache_write_json(cache_path, resp)
        return meta_cache[key]

    def _pick_attr_id(attrs: list[dict], kind: str) -> int | None:
        kind = kind.lower().strip()
        # composition: предпочтение имен со словом 'состав'
        if kind == 'composition':
            best_id = None
            best_score = -1
            for a in attrs or []:
                name = str(a.get('name') or '').strip().lower()
                if not name:
                    continue
                t = str(a.get('type') or '').lower()
                if t not in ('string', 'text'):
                    continue
                if bool(a.get('is_collection')):
                    continue
                score = 0
                if name == 'состав':
                    score += 120
                if 'состав' in name:
                    score += 40
                group = str(a.get('group_name') or '').strip().lower()
                if 'состав' in group:
                    score += 10
                if score > best_score:
                    try:
                        best_id = int(a.get('id'))
                        best_score = score
                    except Exception:
                        pass
            return best_id

        # color/size: по ключевым словам
        keywords = {
            'color': ('цвет', 'color', 'colour'),
            'size': ('размер', 'size'),
        }.get(kind)
        if not keywords:
            return None
        best_id = None
        best_score = -1
        for a in attrs or []:
            name = str(a.get('name') or '').strip().lower()
            if not name:
                continue
            score = 0
            if any(k in name for k in keywords):
                score += 10
            if kind == 'size' and not bool(a.get('is_collection')):
                score += 2
            if score > best_score:
                try:
                    best_id = int(a.get('id'))
                    best_score = score
                except Exception:
                    pass
        return best_id

    def _fallback_extract_by_name(info_item: dict, name_keywords: tuple[str, ...]) -> str | None:
        """Если size/color не нашли по attribute_id, пытаемся вытащить по названию характеристики."""
        try:
            for attr in (info_item.get('attributes') or []):
                nm = str(attr.get('name') or '').strip().lower()
                if not nm:
                    continue
                if not any(k in nm for k in name_keywords):
                    continue
                vals = []
                for v in (attr.get('values') or []):
                    vv = v.get('value')
                    if vv is not None and str(vv).strip():
                        vals.append(str(vv).strip())
                if vals:
                    # для размера берём первое
                    return ", ".join(vals)
        except Exception:
            return None
        return None

    # 4) вытаскиваем атрибуты по product_ids
    offer_id_to_attrs: dict[str, dict] = {}
    product_ids = list({pid for pid in offer_id_to_product_id.values() if pid is not None})

    import requests
    url = f"{ozon.base_url}/v4/product/info/attributes"
    for chunk in _chunk_list(product_ids, 250):
        payload = {
            'filter': {'product_id': chunk, 'visibility': 'ALL'},
            'limit': len(chunk)
        }
        resp = requests.post(url, json=payload, headers=ozon.headers, timeout=60)
        if resp.status_code != 200:
            logger.warning(f"TZ: Ozon cabinet={cabinet_id} /v4/product/info/attributes status={resp.status_code} text={resp.text[:500]}")
            continue
        data = resp.json() or {}
        for info_item in (data.get('result') or []):
            oid = clean_offer_id(info_item.get('offer_id'))
            if not oid:
                continue

            dcid = offer_id_to_dcid.get(oid)
            tpid = offer_id_to_type_id.get(oid)

            # category
            cat = '—'
            if dcid and tpid:
                cat = type_name_by_pair.get((int(dcid), int(tpid))) or cat
            if cat == '—':
                # запасной вариант: иногда dcid можно распарсить здесь
                try:
                    dcid2 = info_item.get('description_category_id')
                    tpid2 = info_item.get('type_id')
                    if dcid2 is not None and tpid2 is not None:
                        cat = type_name_by_pair.get((int(dcid2), int(tpid2))) or cat
                except Exception:
                    pass

            # barcode
            barcodes = []
            bcs = info_item.get('barcodes')
            if isinstance(bcs, list):
                barcodes = [str(x).strip() for x in bcs if str(x).strip()]
            if not barcodes:
                bc = info_item.get('barcode')
                if bc is not None and str(bc).strip():
                    barcodes = [str(bc).strip()]

            # resolve attr ids by meta
            color = '—'
            size = '—'
            composition = '—'
            if dcid and tpid:
                attrs = _get_meta(dcid, tpid)
                cid = _pick_attr_id(attrs, 'color')
                sid = _pick_attr_id(attrs, 'size')
                compid = _pick_attr_id(attrs, 'composition')
                if cid:
                    vals, _ = extract_attribute_values_from_product_attributes(info_item, int(cid))
                    if vals:
                        color = ", ".join([v for v in vals if str(v).strip()]) or '—'
                if sid:
                    vals, _ = extract_attribute_values_from_product_attributes(info_item, int(sid))
                    if vals:
                        size = ", ".join([v for v in vals if str(v).strip()]) or '—'

                if compid:
                    vals, _ = extract_attribute_values_from_product_attributes(info_item, int(compid))
                    vals = [str(v).strip() for v in (vals or []) if str(v).strip()]
                    if vals:
                        composition = vals[0]

                if (not size or size == '—'):
                    fb = _fallback_extract_by_name(info_item, ("размер", "size"))
                    if fb:
                        size = fb

            # Нормализация: если размера по факту нет/0/ONE — считаем, что он "единый"
            size_norm = str(size or '').strip()
            if not size_norm or size_norm in {'—', '0'} or size_norm.upper() in {'ONE', 'ONE SIZE', 'ONESIZE'}:
                size = 'единый'

            offer_id_to_attrs[oid] = {
                'category': cat or '—',
                'barcode': barcodes[0] if barcodes else '—',
                'color': color or '—',
                'size': size or 'единый',
                'composition': composition or '—',
            }

    return offer_id_to_attrs


# --- WB helpers для состава через object/charcs + card.wb.ru ---

def _wb_find_composition_charc_id(wb: WildberriesAPI, subject_id: int) -> int | None:
    """Ищем charcID для характеристики состава по subjectId."""
    try:
        charcs = wb.get_object_charcs(int(subject_id))
    except Exception as e:
        logger.warning(f"WB TZ: get_object_charcs subjectId={subject_id} error={e}")
        return None

    best_id = None
    best_score = -1
    for ch in charcs or []:
        if not isinstance(ch, dict):
            continue
        name = str(ch.get('name') or '').strip().lower()
        if not name:
            continue
        score = 0
        if name == 'состав':
            score += 200
        if 'состав' in name:
            score += 80
        if 'материал' in name:
            score += 20
        if score > best_score:
            try:
                best_id = int(ch.get('charcID') or ch.get('charcId') or ch.get('id'))
                best_score = score
            except Exception:
                pass

    return best_id


def _wb_extract_composition_from_card_api_by_charc_id(payload: dict, charc_id: int) -> str | None:
    """Достаём значение конкретной характеристики по её charcID из card.wb.ru."""
    if not isinstance(payload, dict) or not charc_id:
        return None

    data = payload.get('data') or payload.get('Data') or {}
    products = data.get('products') or data.get('Products') or []
    if not products or not isinstance(products, list) or not isinstance(products[0], dict):
        return None

    p0 = products[0]

    def _iter_name_value_id(items):
        for it in items or []:
            if not isinstance(it, dict):
                continue
            cid = it.get('charcID') or it.get('charcId') or it.get('id')
            nm = it.get('name') or it.get('Name')
            val = it.get('value') or it.get('Value')
            yield cid, nm, val

    # часто нужное лежит в properties
    for key in ('properties', 'Properties'):
        for cid, nm, val in _iter_name_value_id(p0.get(key) or []):
            try:
                if cid is not None and int(cid) == int(charc_id):
                    if isinstance(val, list):
                        txt = ', '.join([str(x).strip() for x in val if str(x).strip()])
                    else:
                        txt = str(val).strip() if val is not None else ''
                    return txt or None
            except Exception:
                continue

    # запасной вариант: characteristics
    for key in ('characteristics', 'Characteristics'):
        for cid, nm, val in _iter_name_value_id(p0.get(key) or []):
            try:
                if cid is not None and int(cid) == int(charc_id):
                    if isinstance(val, list):
                        txt = ', '.join([str(x).strip() for x in val if str(x).strip()])
                    else:
                        txt = str(val).strip() if val is not None else ''
                    return txt or None
            except Exception:
                continue

    # последний шанс — по имени
    for key in ('properties', 'Properties', 'characteristics', 'Characteristics', 'options', 'Options'):
        for cid, nm, val in _iter_name_value_id(p0.get(key) or []):
            try:
                nm_l = str(nm or '').strip().lower()
                if 'состав' in nm_l:
                    if isinstance(val, list):
                        txt = ', '.join([str(x).strip() for x in val if str(x).strip()])
                    else:
                        txt = str(val).strip() if val is not None else ''
                    if txt:
                        return txt
            except Exception:
                continue

    return None


def _wb_get_cards_by_vendor_codes_fuzzy(cabinet_id: int, vendor_codes: list[str]) -> list[dict]:
    """Пытаемся получить карточки WB через content-api по vendorCodes, даже если нет остатков."""
    wb = WildberriesAPI(cabinet_id=cabinet_id)
    # Метод get_cards_by_vendor_codes уже есть в wb_remains_handler и содержит часть запасной логики.
    try:
        cards = wb.get_cards_by_vendor_codes(vendor_codes)
        if cards:
            return cards
    except Exception:
        cards = []

    # запасной вариант: textSearch по нескольким запросам (контент API иногда не ищет по vendorCodes)
    found: list[dict] = []
    try:
        for q in vendor_codes[:20]:
            payload = {
                "settings": {
                    "cursor": {"limit": 100},
                    "filter": {"textSearch": str(q)}
                }
            }
            data = wb._content_post("/content/v2/get/cards/list", payload, timeout=60)  # type: ignore[attr-defined]
            for c in (data or {}).get("cards") or []:
                if isinstance(c, dict):
                    found.append(c)
    except Exception:
        pass

    # убираем дубликаты по nmID
    uniq = {}
    for c in found:
        try:
            nm = int(c.get('nmID') or c.get('nmId'))
        except Exception:
            nm = None
        if nm:
            uniq[nm] = c
    return list(uniq.values())


def _wb_extract_attrs_for_articles(cabinet_id: int, stocks_raw_items: list[dict], needed_articles: set[str] | None = None) -> dict[str, dict]:
    """WB article key -> {category, color, size, composition, barcode}."""
    wb = WildberriesAPI(cabinet_id=cabinet_id)
    single_size = normalize_wb_size(None)
    na = "-"

    def _size_key(v) -> str:
        s = normalize_wb_size(v)
        if s == single_size:
            return s
        src = str(v if v is not None else s).replace("\\", "/")
        m = re.search(r"(\d{2,3})\s*[-/]\s*(\d{2,3})", src)
        if m:
            return f"{m.group(1)}-{m.group(2)}"
        return s

    def _infer_color_from_article(a: str) -> str:
        low = _normalize_key_fuzzy(a)
        color_map = {
            "\u0431\u0435\u043b": "\u0431\u0435\u043b\u044b\u0439",
            "\u0433\u043e\u043b\u0443\u0431": "\u0433\u043e\u043b\u0443\u0431\u043e\u0439",
            "\u0440\u043e\u0437\u043e\u0432": "\u0440\u043e\u0437\u043e\u0432\u044b\u0439",
            "\u0447\u0435\u0440\u043d": "\u0447\u0435\u0440\u043d\u044b\u0439",
            "\u0441\u0438\u043d": "\u0441\u0438\u043d\u0438\u0439",
            "\u0437\u0435\u043b": "\u0437\u0435\u043b\u0435\u043d\u044b\u0439",
            "\u043a\u0440\u0430\u0441": "\u043a\u0440\u0430\u0441\u043d\u044b\u0439",
            "\u0431\u0435\u0436": "\u0431\u0435\u0436\u0435\u0432\u044b\u0439",
            "\u0441\u0435\u0440": "\u0441\u0435\u0440\u044b\u0439",
        }
        for token, color_name in color_map.items():
            if token in low:
                return color_name
        return na

    nm_ids: list[int] = []
    nm_id_by_article: dict[str, int] = {}
    stock_barcode_by_article: dict[str, str] = {}
    category_by_article: dict[str, str] = {}
    size_by_article: dict[str, str] = {}
    base_by_article: dict[str, str] = {}
    barcode_by_base_size: dict[tuple[str, str], str] = {}
    all_articles: set[str] = set()

    for it in stocks_raw_items or []:
        raw_art = clean_article(it.get('supplierArticle'))
        if not raw_art:
            continue

        # В некоторых кабинетах supplierArticle уже приходит с размером
        # (например, "... 112-134 см"). Разделяем на базу + размер и нормализуем.
        parsed_base, parsed_size = _split_wb_article_and_size(raw_art)
        base_art = parsed_base if parsed_size != single_size else raw_art

        size_val = _size_key(it.get('techSize'))
        if size_val == single_size and parsed_size != single_size:
            size_val = _size_key(parsed_size)

        full_art = base_art if size_val == single_size else f"{base_art} {size_val}"

        all_articles.add(base_art)
        all_articles.add(full_art)
        all_articles.add(raw_art)
        base_by_article[base_art] = base_art
        base_by_article[full_art] = base_art
        base_by_article[raw_art] = base_art
        size_by_article[base_art] = size_val
        size_by_article[full_art] = size_val
        size_by_article[raw_art] = size_val

        nm = it.get('nmId') or it.get('nmID')
        try:
            if nm is not None:
                nm_i = int(nm)
                if nm_i > 0:
                    nm_ids.append(nm_i)
                    nm_id_by_article[base_art] = nm_i
                    nm_id_by_article[full_art] = nm_i
                    nm_id_by_article[raw_art] = nm_i
        except Exception:
            pass

        bc = it.get('barcode')
        if bc is not None and str(bc).strip():
            bc_str = str(bc).strip()
            stock_barcode_by_article[full_art] = bc_str
            stock_barcode_by_article.setdefault(base_art, bc_str)
            stock_barcode_by_article.setdefault(raw_art, bc_str)
            barcode_by_base_size[(base_art, size_val)] = bc_str

        cat = it.get('subject') or it.get('category')
        if cat is not None and str(cat).strip():
            cat_str = str(cat).strip()
            category_by_article[full_art] = cat_str
            category_by_article.setdefault(base_art, cat_str)
            category_by_article.setdefault(raw_art, cat_str)

    known_articles = set(all_articles)
    missing_needed: list[str] = []
    needed_alias_to_base: dict[str, str] = {}
    needed_alias_size: dict[str, str] = {}
    if needed_articles:
        for a0 in needed_articles:
            a = clean_article(a0)
            if not a:
                continue

            base, explicit_size = _split_wb_article_and_size(a)
            explicit_size = _size_key(explicit_size)
            if explicit_size == single_size:
                explicit_size = ''

            resolved = None
            need_base_norm = _normalize_wb_base_for_match(base or a)
            ranked: list[tuple[int, str]] = []
            for cand in known_articles:
                c_base, c_size = _split_wb_article_and_size(cand)
                if _normalize_wb_base_for_match(c_base) != need_base_norm:
                    continue
                score = 0
                if explicit_size:
                    if c_size == explicit_size:
                        score = 3
                    elif c_size == single_size:
                        score = 1
                else:
                    score = 2
                if score > 0:
                    ranked.append((score, cand))
            if ranked:
                ranked.sort(key=lambda x: (x[0], len(x[1])), reverse=True)
                chosen = ranked[0][1]
                resolved = base_by_article.get(chosen, chosen)

            for probe in (base, a):
                if resolved:
                    break
                if probe in base_by_article and probe in known_articles:
                    resolved = base_by_article[probe]
                    break
                nk_probe = _normalize_key_fuzzy(probe)
                candidates = [x for x in known_articles if _normalize_key_fuzzy(x) == nk_probe]
                if not candidates:
                    candidates = [x for x in known_articles if nk_probe.startswith(_normalize_key_fuzzy(x))]
                if candidates:
                    resolved = sorted(candidates, key=len, reverse=True)[0]
                    break
            if not resolved:
                # Сохраняем исходный ключ, если стабильную базу найти не удалось.
                # Это помогает для WB-карточек, где vendorCode уже содержит суффикс размера/единиц.
                resolved = a
            base = resolved

            needed_alias_to_base[a] = base
            if explicit_size:
                needed_alias_size[a] = explicit_size

            if base not in known_articles and base not in missing_needed:
                missing_needed.append(base)
            all_articles.add(a)
            all_articles.add(base)

    cards_by_vendor: dict[str, dict] = {}
    if missing_needed:
        cards = _wb_get_cards_by_vendor_codes_fuzzy(cabinet_id, missing_needed)
        for c in cards or []:
            vc = str(c.get('vendorCode') or '').strip()
            if vc:
                cards_by_vendor[vc] = c
        for vc, c in cards_by_vendor.items():
            try:
                nm = int(c.get('nmID') or c.get('nmId'))
            except Exception:
                nm = None
            if nm:
                nm_id_by_article.setdefault(vc, nm)
                nm_ids.append(nm)

    nm_ids = list(dict.fromkeys(nm_ids))
    by_nm, by_vendor_from_nm = _wb_build_card_indexes(cabinet_id, nm_ids)
    by_vendor = dict(by_vendor_from_nm)
    by_vendor.update(cards_by_vendor)
    by_vendor_norm: dict[str, dict] = {}
    by_vendor_size_norm: dict[tuple[str, str], dict] = {}

    def _upsert_vendor_indexes(vc: str, card: dict) -> None:
        if not isinstance(card, dict):
            return
        vc_base, _vc_size = _split_wb_article_and_size(vc)
        for probe in {vc, vc_base}:
            nk = _normalize_wb_base_for_match(probe)
            if nk and nk not in by_vendor_norm:
                by_vendor_norm[nk] = card
        base_norm = _normalize_wb_base_for_match(vc_base or vc)
        for sz in (card.get('sizes') or []):
            if not isinstance(sz, dict):
                continue
            sz_norm = _size_key(sz.get('techSize') or sz.get('wbSize') or card.get('techSize'))
            if not sz_norm or sz_norm == single_size:
                continue
            key = (base_norm, sz_norm)
            if key not in by_vendor_size_norm:
                by_vendor_size_norm[key] = card

    for vc, c in (by_vendor or {}).items():
        _upsert_vendor_indexes(vc, c)

    # В get_fbo_stocks_v1 этих vendorCode может не быть. В запасном варианте
    # подтягиваем все карточки из content API.
    try:
        all_cards = wb.get_all_cards(limit=50) or []
        for c in all_cards:
            if not isinstance(c, dict):
                continue
            vc = clean_article(c.get('vendorCode') or c.get('supplierArticle') or '')
            if not vc:
                continue
            by_vendor.setdefault(vc, c)
            _upsert_vendor_indexes(vc, c)
    except Exception:
        pass
    composition_charc_id_by_subject: dict[int, int] = {}

    attrs: dict[str, dict] = {}
    keys = all_articles | set(category_by_article.keys()) | set(nm_id_by_article.keys()) | set(stock_barcode_by_article.keys()) | set(size_by_article.keys())
    for art in keys:
        base_art = base_by_article.get(art, art)
        nm = nm_id_by_article.get(art) or nm_id_by_article.get(base_art)
        card = by_nm.get(nm) if nm else (by_vendor.get(base_art) or by_vendor.get(art))
        if not card:
            nk = _normalize_wb_base_for_match(base_art)
            if nk:
                card = by_vendor_norm.get(nk)
        if not card:
            art_base, art_size = _split_wb_article_and_size(art)
            if art_base and art_size and art_size != single_size:
                card = by_vendor_size_norm.get((_normalize_wb_base_for_match(art_base), art_size))

        color = na
        try:
            cval = wb.extract_color_from_content_card(card) if card else None
            if cval:
                color = cval
        except Exception:
            pass
        if color == na:
            color = _infer_color_from_article(base_art)

        composition = na
        try:
            if not nm and isinstance(card, dict):
                try:
                    nm = int(card.get('nmID') or card.get('nmId'))
                except Exception:
                    nm = None

            if nm:
                comp = _wb_extract_composition_from_content_card(card)
                if comp:
                    composition = comp
                else:
                    payload = wb.get_card_by_nm_id(nm)
                    comp2 = wb.extract_composition_from_card_api(payload or {})
                    if comp2:
                        composition = comp2
                    else:
                        subject_id = None
                        try:
                            subject_id = (card or {}).get('subjectID') or (card or {}).get('subjectId')
                        except Exception:
                            subject_id = None
                        if subject_id is not None:
                            try:
                                sid_int = int(subject_id)
                            except Exception:
                                sid_int = None
                            if sid_int:
                                if sid_int not in composition_charc_id_by_subject:
                                    cid = _wb_find_composition_charc_id(wb, sid_int)
                                    if cid:
                                        composition_charc_id_by_subject[sid_int] = int(cid)
                                cid = composition_charc_id_by_subject.get(sid_int)
                                if cid:
                                    comp3 = _wb_extract_composition_from_card_api_by_charc_id(payload or {}, int(cid))
                                    if comp3:
                                        composition = comp3

                        if not composition or composition == na:
                            logger.warning(
                                f"WB TZ: composition not found cabinet={cabinet_id} nmId={nm} art={art} "
                                f"hasCard={bool(card)} subjectId={subject_id}"
                            )
        except Exception as e:
            logger.warning(f"WB TZ: composition error cabinet={cabinet_id} nmId={nm} art={art} error={e}")

        cat_val = category_by_article.get(art) or category_by_article.get(base_art)
        if (not cat_val or cat_val == na) and isinstance(card, dict):
            cat_val = str(card.get('subjectName') or card.get('subject') or card.get('category') or na).strip() or na

        art_base_for_size, art_size_from_key = _split_wb_article_and_size(art)
        size_val = size_by_article.get(art) or size_by_article.get(base_art)
        if art_size_from_key and art_size_from_key != single_size:
            size_val = art_size_from_key
        if (not size_val or size_val == single_size) and isinstance(card, dict):
            try:
                s0 = (card.get('sizes') or [{}])[0]
                ts = s0.get('techSize') or card.get('techSize')
                if ts:
                    size_val = _size_key(ts)
            except Exception:
                pass

        bc_val = stock_barcode_by_article.get(art) or stock_barcode_by_article.get(base_art)
        if (not bc_val or bc_val == na) and isinstance(card, dict):
            try:
                wanted_size = _size_key(art_size_from_key if art_size_from_key and art_size_from_key != single_size else (size_by_article.get(art) or size_by_article.get(base_art)))
                chosen_size_obj = None
                sizes_arr = card.get('sizes') or []
                if isinstance(sizes_arr, list):
                    if wanted_size and wanted_size != single_size:
                        for sz in sizes_arr:
                            if not isinstance(sz, dict):
                                continue
                            sz_norm = _size_key(sz.get('techSize') or sz.get('wbSize') or card.get('techSize'))
                            if sz_norm == wanted_size:
                                chosen_size_obj = sz
                                break
                    if chosen_size_obj is None and sizes_arr:
                        chosen_size_obj = sizes_arr[0]
                if isinstance(chosen_size_obj, dict):
                    bcs = chosen_size_obj.get('skus') or chosen_size_obj.get('barcodes') or []
                    if isinstance(bcs, list) and bcs:
                        bc_val = str(bcs[0]).strip()
            except Exception:
                pass

        attrs[art] = {
            'category': cat_val or na,
            'barcode': bc_val or na,
            'color': color or na,
            'size': size_val or single_size,
            'composition': composition or na,
        }

    for alias, base in needed_alias_to_base.items():
        if alias in attrs:
            continue
        base_attrs = attrs.get(base)
        if not base_attrs:
            continue
        out = dict(base_attrs)
        forced_size = needed_alias_size.get(alias)
        if forced_size:
            out['size'] = forced_size
            bc = barcode_by_base_size.get((base, forced_size))
            if bc:
                out['barcode'] = bc
        attrs[alias] = out

    return attrs

def _wb_build_card_indexes(cabinet_id: int, nm_ids: list[int]) -> tuple[dict[int, dict], dict[str, dict]]:
    """Запрашивает WB content-api карточки по nmId и строит индексы для быстрого доступа."""
    wb = WildberriesAPI(cabinet_id=cabinet_id)
    cards = wb.get_cards_by_nm_ids(nm_ids)
    by_nm: dict[int, dict] = {}
    by_vendor: dict[str, dict] = {}
    for c in cards or []:
        if not isinstance(c, dict):
            continue
        nm = None
        try:
            nm = int(c.get('nmID') or c.get('nmId'))
        except Exception:
            nm = None
        if nm is not None:
            by_nm[nm] = c
        vc = str(c.get('vendorCode') or '').strip()
        if vc:
            by_vendor[vc] = c
    return by_nm, by_vendor


def _wb_extract_composition_from_content_card(card: dict | None) -> str | None:
    """Пытается вытащить состав из карточки WB content-api.

    Структура может отличаться, поэтому проверяем несколько мест.
    """
    if not isinstance(card, dict):
        return None

    def _iter_name_value(items):
        for it in items or []:
            if not isinstance(it, dict):
                continue
            name = str(it.get('name') or it.get('title') or '').strip().lower()
            if not name:
                continue
            value = it.get('value')
            yield name, value

    # 1) characteristics
    for name, value in _iter_name_value(card.get('characteristics')):
        if 'состав' in name:
            if isinstance(value, list):
                txt = ", ".join([str(x).strip() for x in value if str(x).strip()])
            else:
                txt = str(value).strip() if value is not None else ''
            if txt:
                return txt

    # 2) options/properties
    for key in ('options', 'properties'):
        for name, value in _iter_name_value(card.get(key)):
            if 'состав' in name:
                if isinstance(value, list):
                    txt = ", ".join([str(x).strip() for x in value if str(x).strip()])
                else:
                    txt = str(value).strip() if value is not None else ''
                if txt:
                    return txt

    # 3) sizes[].chars
    for size in (card.get('sizes') or []):
        if not isinstance(size, dict):
            continue
        for name, value in _iter_name_value(size.get('chars') or size.get('characteristics')):
            if 'состав' in name:
                if isinstance(value, list):
                    txt = ", ".join([str(x).strip() for x in value if str(x).strip()])
                else:
                    txt = str(value).strip() if value is not None else ''
                if txt:
                    return txt

    return None


def _normalize_key_fuzzy(s: str) -> str:
    """Нормализация ключей артикула/offer_id для устойчивого мэтчинга."""
    s = str(s or '')
    s = ''.join(c for c in s if c.isprintable())
    s = s.strip().lower()
    s = s.replace("\u0451", "\u0435")
    s = s.replace('\\', '/')
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\s+", " ", s)
    s = s.replace(" /", "/").replace("/ ", "/")
    s = re.sub(r"/+", "/", s)
    # удаляем знаки препинания, которые часто расходятся между базой/МП
    s = s.replace('.', '')
    s = s.replace("-", "").replace("\u2013", "").replace("\u2014", "")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _split_wb_article_and_size(article: str) -> tuple[str, str]:
    """Разделить WB-артикул на (базовый артикул, размер)."""
    s = str(article or '').strip()
    single_size = normalize_wb_size(None)
    if not s:
        return '', single_size

    # Formats:
    # - "артикул 92-110"
    # - "артикул/92/110"
    src = s.replace("\\", "/")
    m = re.search(r"(\d{2,3})\s*[-/]\s*(\d{2,3})", src)
    if not m:
        return s, single_size

    size = f"{m.group(1)}-{m.group(2)}"
    base = src[:m.start()].strip(" /-")
    base = base.strip(" /-")
    return (base or s), size


def _normalize_wb_base_for_match(article_base: str) -> str:
    s = str(article_base or "").strip().lower().replace("\\", "/")
    s = s.replace("/", " ")
    s = (
        s.replace("\u0433\u043e\u043b\u0443\u0431\u0430\u044f", "\u0433\u043e\u043b\u0443\u0431")
        .replace("\u0433\u043e\u043b\u0443\u0431\u043e\u0439", "\u0433\u043e\u043b\u0443\u0431")
        .replace("\u0431\u0435\u043b\u0430\u044f", "\u0431\u0435\u043b")
        .replace("\u0431\u0435\u043b\u044b\u0439", "\u0431\u0435\u043b")
        .replace("\u0440\u043e\u0437\u043e\u0432\u0430\u044f", "\u0440\u043e\u0437\u043e\u0432")
        .replace("\u0440\u043e\u0437\u043e\u0432\u044b\u0439", "\u0440\u043e\u0437\u043e\u0432")
    )
    s = (
        s.replace("\u0434\u0435\u0442\u0441\u043a\u0430\u044f", "\u0434\u0435\u0442\u0441\u043a")
        .replace("\u0434\u0435\u0442\u0441\u043a\u0438\u0439", "\u0434\u0435\u0442\u0441\u043a")
        .replace("\u0434\u0435\u0442\u0441\u043a\u043e\u0435", "\u0434\u0435\u0442\u0441\u043a")
    )
    s = re.sub(r"\s+", " ", s).strip(" /-")
    return _normalize_key_fuzzy(s)


def _canonical_wb_article(article: str) -> str:
    """Return canonical WB article form: '<base> <size>' or '<base>'."""
    base, size = _split_wb_article_and_size(article)
    if not base:
        return str(article or "").strip()
    if size and size != normalize_wb_size(None):
        return f"{base} {size}".strip()
    return base


async def _generate_tz_zip(context: CallbackContext, chat_id: int, selected: set[str], rows: list[dict]) -> tuple[str, str]:
    thresholds = get_stock_thresholds(chat_id)
    yellow_limit = None
    try:
        yellow_limit = int((thresholds or {}).get('yellow', 0))
    except Exception:
        yellow_limit = None

    # 1) остатки по кабинетам
    cabinet_stocks: dict[str, dict[str, int]] = {}
    # нормализованные индексы остатков (для устойчивого поиска по артикулу)
    cabinet_stocks_norm: dict[str, dict[str, int]] = {}
    # для WB дополнительно держим сырые stocks для атрибутов
    wb_stocks_raw_by_cabinet: dict[int, list[dict]] = {}

    for key in selected:
        mp, cid_str = key.split(":", 1)
        cid = int(cid_str)
        t0 = time.time()
        if mp == 'ozon':
            _, raw = await fetch_ozon_remains_raw(cid)
            st: dict[str, int] = {}
            st_norm: dict[str, int] = {}
            for it in raw or []:
                art_raw = it.get('Артикул')
                art = str(art_raw or '').strip()
                if not art:
                    continue
                try:
                    val = int(it.get('Итого на МП') or 0)
                except Exception:
                    val = 0
                st[art] = val
                nk = _normalize_key_fuzzy(art)
                # если коллизия — берём максимум (обычно это один и тот же товар с разным регистром)
                if nk:
                    st_norm[nk] = max(val, st_norm.get(nk, 0))
            cabinet_stocks[key] = st
            cabinet_stocks_norm[key] = st_norm
        else:
            # WB
            _, raw = await fetch_wb_remains_raw(cid)
            st = {}
            st_norm: dict[str, int] = {}
            for it in raw or []:
                art = str(it.get('Артикул') or '').strip()
                if not art:
                    continue
                try:
                    val = int(it.get('Итого на МП') or 0)
                except Exception:
                    val = 0
                st[art] = val
                nk = _normalize_key_fuzzy(art)
                if nk:
                    st_norm[nk] = max(val, st_norm.get(nk, 0))
            cabinet_stocks[key] = st
            cabinet_stocks_norm[key] = st_norm

            # получаем сырые stocks через api (для techSize/barcode/nmId/category)
            try:
                wb = WildberriesAPI(cabinet_id=cid)
                wb_stocks_raw_by_cabinet[cid] = wb.get_fbo_stocks_v1() or []
            except Exception:
                wb_stocks_raw_by_cabinet[cid] = []

        logger.info(f"TZ: остатки {key} артикула={len(cabinet_stocks[key])} за {time.time()-t0:.2f}s")

    # 2) связи template_name -> real articles в разрезе кабинетов
    links_by_cabinet: dict[str, dict[int, list[str]]] = {}
    name_by_cabinet: dict[str, dict[int, str]] = {}

    for key in selected:
        mp, cid_str = key.split(":", 1)
        cid = int(cid_str)
        if mp == 'ozon':
            sheet = {1: 'Отдельно Озон Nimba', 2: 'Отдельно Озон Galioni', 3: 'Отдельно Озон AGNIA'}[cid]
        else:
            sheet = {1: 'Отдельно ВБ Nimba', 2: 'Отдельно ВБ Galioni', 3: 'Отдельно ВБ AGNIA'}[cid]
        id_to_name, id_to_arts = get_cabinet_articles_by_template_id(sheet)
        name_by_cabinet[key] = id_to_name
        links_by_cabinet[key] = id_to_arts

    # индекс template_name->template_id
    template_name_to_id: dict[str, int] = {}
    for key in selected:
        for tid, nm in (name_by_cabinet.get(key) or {}).items():
            template_name_to_id.setdefault(_normalize_name(nm), int(tid))

    # 3) Собираем нужные позиции
    planned_lines: list[dict] = []

    # сюда соберём случаи, когда ничего не попало в ТЗ из-за достаточных остатков
    # key: template_key -> info
    not_added: dict[str, dict] = {}

    for row in rows:
        tname_key = row['key']
        qty_need = int(row['qty'])
        tid = template_name_to_id.get(tname_key)
        if tid is None:
            logger.warning(f"TZ: не найден template_id для '{row['name']}'")
            continue

        # собрать все связанные real_articles по всем выбранным кабинетам
        items: list[dict] = []
        added_item_keys: set[tuple[str, str]] = set()
        for cab_key in selected:
            id_to_arts = links_by_cabinet.get(cab_key) or {}
            arts = [str(a).strip() for a in (id_to_arts.get(tid) or []) if str(a).strip()]
            if not arts:
                continue
            st = cabinet_stocks.get(cab_key) or {}
            stn = cabinet_stocks_norm.get(cab_key) or {}
            for a in arts:
                # поиск остатков: exact -> fuzzy(lower/punct)
                # Для WB дополнительно учитываем размерные строки вида "артикул + размер".
                mp, _cid = cab_key.split(":", 1)
                if mp == "wb":
                    nk_a = _normalize_key_fuzzy(a)
                    matches: list[tuple[str, int]] = []
                    for art_key, stock_val in st.items():
                        art_key_s = str(art_key).strip()
                        nk_item = _normalize_key_fuzzy(art_key_s)
                        if (
                            art_key_s == a
                            or art_key_s.startswith(f"{a} ")
                            or nk_item == nk_a
                            or nk_item.startswith(nk_a)
                        ):
                            try:
                                stock_val_i = int(stock_val or 0)
                            except Exception:
                                stock_val_i = 0
                            matches.append((art_key_s, stock_val_i))

                # Если для базового артикула есть размерные строки,
                # оставляем только их и пропускаем базовый дубликат.
                    size_matches = [x for x in matches if x[0] != a and x[0].startswith(f"{a} ")]
                    selected_matches = size_matches or matches

                    if selected_matches:
                        for art_key_s, stock_val_i in selected_matches:
                            item_key = (cab_key, art_key_s)
                            if item_key in added_item_keys:
                                continue
                            added_item_keys.add(item_key)
                            lvl = _stocks_level(stock_val_i, thresholds)
                            items.append({
                                'cabinet_key': cab_key,
                                'article': art_key_s,
                                'stock': stock_val_i,
                                'level': lvl,
                            })
                    else:
                        stock_val = st.get(a)
                        if stock_val is None:
                            stock_val = stn.get(_normalize_key_fuzzy(a), 0)
                        try:
                            stock_val_i = int(stock_val or 0)
                        except Exception:
                            stock_val_i = 0
                        item_key = (cab_key, a)
                        if item_key in added_item_keys:
                            continue
                        added_item_keys.add(item_key)
                        lvl = _stocks_level(stock_val_i, thresholds)
                        items.append({
                            'cabinet_key': cab_key,
                            'article': a,
                            'stock': stock_val_i,
                            'level': lvl,
                        })
                else:
                    stock_val = st.get(a)
                    if stock_val is None:
                        stock_val = stn.get(_normalize_key_fuzzy(a), 0)
                    try:
                        stock_val_i = int(stock_val or 0)
                    except Exception:
                        stock_val_i = 0
                    item_key = (cab_key, a)
                    if item_key in added_item_keys:
                        continue
                    added_item_keys.add(item_key)
                    lvl = _stocks_level(stock_val_i, thresholds)
                    items.append({
                        'cabinet_key': cab_key,
                        'article': a,
                        'stock': stock_val_i,
                        'level': lvl,
                    })

        if not items:
            continue

        # кандидаты для ТЗ: все товары с остатком <= желтой границы
        danger_items = [it for it in items if _needs_restock(int(it.get('stock') or 0), thresholds)]

        if not danger_items:
            # ничего не добавили: все связанные артикулы имеют stock > yellow
            # соберём подробности по кабинетам
            by_cab: dict[str, list[tuple[str, int]]] = {}
            for it in items:
                by_cab.setdefault(str(it.get('cabinet_key')), []).append((str(it.get('article')), int(it.get('stock') or 0)))
            not_added[tname_key] = {
                'name': row['name'],
                'qty': qty_need,
                'by_cab': by_cab,
            }
            continue

        allowed = {(it['cabinet_key'], it['article']) for it in danger_items if _needs_restock(int(it.get('stock') or 0), thresholds)}
        ship = _distribute_qty_equalize_final(items, qty_need, allowed_keys=allowed)

        # если по какой-то причине распределение не дало строк (например, qty_need=0) — тоже учтём
        if not ship:
            by_cab: dict[str, list[tuple[str, int]]] = {}
            for it in items:
                by_cab.setdefault(str(it.get('cabinet_key')), []).append((str(it.get('article')), int(it.get('stock') or 0)))
            not_added[tname_key] = {
                'name': row['name'],
                'qty': qty_need,
                'by_cab': by_cab,
            }
            continue

        for (cab_key, art), ship_qty in ship.items():
            cur_stock = int(next((x['stock'] for x in items if x['cabinet_key'] == cab_key and x['article'] == art), 0))
            if thresholds and not _needs_restock(cur_stock, thresholds):
                logger.warning(
                    f"TZ: WARNING артикул попал в план, но stock>=yellow: cab={cab_key} art='{art}' stock={cur_stock} thresholds={thresholds}"
                )
            planned_lines.append({
                'template_name': row['name'],
                'template_key': tname_key,
                'cabinet': cab_key,
                'article': art,
                'ship_qty': int(ship_qty),
                'current_stock': cur_stock,
                'stock_level': str(next((x['level'] for x in items if x['cabinet_key'] == cab_key and x['article'] == art), "unknown")),
            })

    if not planned_lines:
        raise ValueError("Не удалось сформировать ТЗ: по выбранным кабинетам нет позиций с остатками <= жёлтой границы.")

    # 4) подтянуть атрибуты по кабинетам для всех запланированных строк
    needed_by_cabinet: dict[str, set[str]] = {}
    for ln in planned_lines:
        needed_by_cabinet.setdefault(ln['cabinet'], set()).add(str(ln['article']))

    ozon_attrs_by_cabinet: dict[int, dict[str, dict]] = {}
    wb_attrs_by_cabinet: dict[int, dict[str, dict]] = {}

    # вспомогательные индексы norm->attrs, чтобы матчить «кривые» ключи
    ozon_attrs_norm_by_cabinet: dict[int, dict[str, dict]] = {}
    wb_attrs_norm_by_cabinet: dict[int, dict[str, dict]] = {}

    for cab_key in needed_by_cabinet.keys():
        mp, cid_str = cab_key.split(':', 1)
        cid = int(cid_str)
        if mp == 'ozon':
            t0 = time.time()
            offer_to_attrs = _ozon_extract_offer_id_to_attrs(cid)

            # индекс по нормализованному ключу
            norm_map: dict[str, dict] = {}
            norm_map_no_ozn: dict[str, dict] = {}
            for k, v in (offer_to_attrs or {}).items():
                nk = _normalize_key_fuzzy(k)
                if nk and nk not in norm_map:
                    norm_map[nk] = v
                # иногда в базе артикулы могут храниться без префикса OZN
                if nk.startswith('ozn'):
                    nk2 = nk[3:]
                    if nk2 and nk2 not in norm_map_no_ozn:
                        norm_map_no_ozn[nk2] = v
            ozon_attrs_norm_by_cabinet[cid] = norm_map

            needed = needed_by_cabinet[cab_key]
            out = {}
            for k in needed:
                if k in offer_to_attrs:
                    out[k] = offer_to_attrs[k]
                    continue
                nk = _normalize_key_fuzzy(k)
                v = norm_map.get(nk)
                if not v and nk.startswith('ozn'):
                    v = norm_map_no_ozn.get(nk[3:])
                if not v:
                    # ещё один шанс: если в базе БЕЗ OZN, а в API С OZN
                    v = norm_map.get('ozn' + nk)
                if v:
                    out[k] = v
            ozon_attrs_by_cabinet[cid] = out

            missed = [x for x in needed if x not in out]
            if missed:
                logger.warning(
                    f"TZ: Ozon attrs cabinet={cid} missed={len(missed)}/{len(needed)} sample={missed[:10]}"
                )
            logger.info(f"TZ: Ozon attrs cabinet={cid} got={len(out)} за {time.time()-t0:.2f}s")
        else:
            t0 = time.time()
            raw_items = wb_stocks_raw_by_cabinet.get(cid) or []
            attrs = _wb_extract_attrs_for_articles(cid, raw_items, needed_articles=needed_by_cabinet[cab_key])

            # нормализованный индекс
            norm_map: dict[str, dict] = {}
            for k, v in (attrs or {}).items():
                nk = _normalize_key_fuzzy(k)
                if nk and nk not in norm_map:
                    norm_map[nk] = v
            wb_attrs_norm_by_cabinet[cid] = norm_map

            needed_orig = needed_by_cabinet[cab_key]
            out = {}
            for k in needed_orig:
                # attrs ключи могут не совпадать регистром/слэшами
                v = attrs.get(k)
                if not v:
                    nk = _normalize_key_fuzzy(k)
                    v = norm_map.get(nk)
                if not v:
                    base_k, size_k = _split_wb_article_and_size(k)
                    canonical_k = f"{base_k} {size_k}" if base_k and size_k != normalize_wb_size(None) else base_k
                    if canonical_k:
                        v = attrs.get(canonical_k)
                        if not v:
                            v = norm_map.get(_normalize_key_fuzzy(canonical_k))
                    if not v and size_k != normalize_wb_size(None):
                        alt_k = f"{base_k} {size_k.replace('-', '/')}"
                        v = attrs.get(alt_k)
                        if not v:
                            v = norm_map.get(_normalize_key_fuzzy(alt_k))
                    if base_k and base_k != k:
                        if not v:
                            v = attrs.get(base_k)
                        if not v:
                            v = norm_map.get(_normalize_key_fuzzy(base_k))
                        if v:
                            merged = dict(v)
                            if size_k and size_k != normalize_wb_size(None):
                                merged['size'] = size_k
                            v = merged
                    # Последний запасной вариант: сопоставить по нормализованной базе + размеру
                    # среди всех ключей attrs.
                    if not v and base_k:
                        nk_base = _normalize_wb_base_for_match(base_k)
                        for ak, av in (attrs or {}).items():
                            a_base, a_size = _split_wb_article_and_size(str(ak))
                            if _normalize_wb_base_for_match(a_base) != nk_base:
                                continue
                            if size_k == normalize_wb_size(None) or a_size == size_k:
                                merged = dict(av or {})
                                if size_k and size_k != normalize_wb_size(None):
                                    merged['size'] = size_k
                                v = merged
                                break
                if v:
                    out[k] = v
            wb_attrs_by_cabinet[cid] = out

            missed = [x for x in needed_orig if x not in out]
            if missed:
                logger.warning(
                    f"TZ: WB attrs cabinet={cid} missed={len(missed)}/{len(needed_orig)} sample={missed[:10]}"
                )

            logger.info(f"TZ: WB attrs cabinet={cid} got={len(out)} за {time.time()-t0:.2f}s")

    # 5) сформировать файлы по кабинетам
    tmp_dir = context.user_data.get('tz_tmp_dir')
    assert tmp_dir

    template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Шаблон ТЗ.xlsx')
    template_path = os.path.normpath(template_path)
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Не найден шаблон ТЗ: {template_path}")

    today = datetime.now().strftime('%Y-%m-%d')
    out_files: list[str] = []

    for cab_key in sorted(set(ln['cabinet'] for ln in planned_lines)):
        mp, cid_str = cab_key.split(':', 1)
        cid = int(cid_str)
        seller = SELLER_BY_CABINET.get(cab_key, '—')
        shop = next((c.label for c in ALL_CABINETS if c.key == cab_key), cab_key)

        lines = [ln for ln in planned_lines if ln['cabinet'] == cab_key]

        # дополнить атрибутами
        for ln in lines:
            art = str(ln['article']).strip()
            if mp == 'ozon':
                attrs_map = ozon_attrs_by_cabinet.get(cid) or {}
                attrs = attrs_map.get(art)
                if not attrs:
                    # запасной вариант: нечеткое сопоставление
                    attrs = (ozon_attrs_norm_by_cabinet.get(cid) or {}).get(_normalize_key_fuzzy(art))
                attrs = attrs or {}
            else:
                attrs_map = wb_attrs_by_cabinet.get(cid) or {}
                attrs = attrs_map.get(art)
                if not attrs:
                    attrs = (wb_attrs_norm_by_cabinet.get(cid) or {}).get(_normalize_key_fuzzy(art))
                base_art, size_from_article = _split_wb_article_and_size(art)
                if not attrs and base_art:
                    canonical_art = f"{base_art} {size_from_article}" if size_from_article != normalize_wb_size(None) else base_art
                    attrs = attrs_map.get(canonical_art)
                    if not attrs:
                        attrs = (wb_attrs_norm_by_cabinet.get(cid) or {}).get(_normalize_key_fuzzy(canonical_art))
                    if not attrs and size_from_article != normalize_wb_size(None):
                        alt_art = f"{base_art} {size_from_article.replace('-', '/')}"
                        attrs = attrs_map.get(alt_art)
                        if not attrs:
                            attrs = (wb_attrs_norm_by_cabinet.get(cid) or {}).get(_normalize_key_fuzzy(alt_art))
                if not attrs and base_art:
                    nk_base = _normalize_wb_base_for_match(base_art)
                    for ak, av in (attrs_map or {}).items():
                        a_base, a_size = _split_wb_article_and_size(str(ak))
                        if _normalize_wb_base_for_match(a_base) != nk_base:
                            continue
                        if size_from_article == normalize_wb_size(None) or a_size == size_from_article:
                            attrs = dict(av or {})
                            if size_from_article and size_from_article != normalize_wb_size(None):
                                attrs['size'] = size_from_article
                            break
                base_attrs = attrs_map.get(base_art)
                if not base_attrs:
                    base_attrs = (wb_attrs_norm_by_cabinet.get(cid) or {}).get(_normalize_key_fuzzy(base_art))
                if not base_attrs and base_art:
                    nk_base = _normalize_wb_base_for_match(base_art)
                    for ak, av in (attrs_map or {}).items():
                        a_base, _ = _split_wb_article_and_size(str(ak))
                        if _normalize_wb_base_for_match(a_base) == nk_base:
                            base_attrs = av
                            break
                if base_attrs:
                    merged = dict(base_attrs)
                    if attrs:
                        merged.update(attrs)
                    attrs = merged
                    if size_from_article != normalize_wb_size(None):
                        attrs['size'] = size_from_article
                        # Проверяем, что баркод тоже взят из той же размерной строки,
                        # если это возможно.
                        nk_base = _normalize_wb_base_for_match(base_art)
                        for ak, av in (attrs_map or {}).items():
                            a_base, a_size = _split_wb_article_and_size(str(ak))
                            if _normalize_wb_base_for_match(a_base) != nk_base:
                                continue
                            if a_size == size_from_article and isinstance(av, dict):
                                bc = str(av.get('barcode') or '').strip()
                                if bc and bc != '—' and bc != '-':
                                    attrs['barcode'] = bc
                                if av.get('category'):
                                    attrs['category'] = av.get('category')
                                if av.get('composition'):
                                    attrs['composition'] = av.get('composition')
                                break
                attrs = attrs or {}

            if not attrs:
                logger.warning(f"TZ: attrs empty cab={cab_key} article='{art}'")

            ln['category'] = attrs.get('category', '—') or '—'
            ln['barcode'] = attrs.get('barcode', '—') or '—'
            ln['color'] = attrs.get('color', '—') or '—'
            ln['size'] = attrs.get('size', normalize_wb_size(None)) or normalize_wb_size(None)
            ln['composition'] = attrs.get('composition', '—') or '—'

        out_path = os.path.join(tmp_dir, _safe_filename(f"ТЗ ({seller}, {shop}, {today}).xlsx"))
        _fill_tz_excel(template_path, out_path, lines, seller, thresholds=thresholds)
        out_files.append(out_path)

    zip_path = os.path.join(tmp_dir, _safe_filename(f"ТЗ_{today}.zip"))
    with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for p in out_files:
            zf.write(p, arcname=os.path.basename(p))

    # формируем сообщение пользователю о не добавленных позициях
    not_added_msg = ""
    if not_added:
        ytxt = str(yellow_limit) if yellow_limit is not None else "(не задано)"

        # Группируем по кабинетам: cabinet -> list[(article, stock)]
        by_cabinet: dict[str, list[tuple[str, int]]] = {}
        for _, info in not_added.items():
            for cab, arts in (info.get('by_cab') or {}).items():
                for art, stock in arts or []:
                    by_cabinet.setdefault(str(cab), []).append((str(art), int(stock)))

        total_templates = len(not_added)
        lines_msg = [
            "⚠️ Не все позиции из файла попали в ТЗ.",
            "Причина: у перечисленных ниже исходных артикулов достаточно остатков (они выше жёлтой границы), поэтому поставка по ним не требуется.",
            f"Текущая жёлтая граница: {ytxt} шт (в ТЗ попадают артикулы с остатком ≤ {ytxt}).",
            f"Затронуто шаблонных позиций из файла: {total_templates}",
            "",
            "Исходные артикулы, которые НЕ попали в ТЗ (остатки на МП) по кабинетам:",
            "",
        ]

        # стабильно выводим кабинеты в порядке ALL_CABINETS
        ordered_cabs = [c.key for c in ALL_CABINETS if c.key in by_cabinet]
        for cab in ordered_cabs + [c for c in by_cabinet.keys() if c not in ordered_cabs]:
            cab_label = next((c.label for c in ALL_CABINETS if c.key == cab), cab)
            rows_list = by_cabinet.get(cab) or []

            # убираем дубликаты (один артикул мог встречаться в нескольких шаблонных строках)
            uniq: dict[str, int] = {}
            for art, stock in rows_list:
                a = str(art).strip()
                if not a:
                    continue
                uniq[a] = max(int(stock), uniq.get(a, -10**9))

            if not uniq:
                continue

            lines_msg.append(f"📌 {cab_label}")
            # сортировка: больше остатков — выше, чтобы сразу было видно "почему не добавили"
            for art, stock in sorted(uniq.items(), key=lambda x: (-x[1], x[0].lower())):
                lines_msg.append(f"  • {art}: {stock} шт")
            lines_msg.append("")

        not_added_msg = "\n".join(lines_msg).strip()

    return zip_path, not_added_msg


def _find_header_row(ws, required_headers: list[str], max_rows: int = 50) -> int | None:
    req = [str(x).strip().lower() for x in required_headers]
    for r in range(1, max_rows + 1):
        values = [str(ws.cell(row=r, column=c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if all(h in values for h in req):
            return r
    return None


def _fill_tz_excel(template_path: str, out_path: str, lines: list[dict], seller: str, thresholds: dict | None = None) -> None:
    wb = load_workbook(template_path)
    ws = wb.active

    header_row = _find_header_row(
        ws,
        [
            "Баркод",
            "Количество",
            "Наименование",
            "Артикул продавца",
            "Цвет на бирке",
            "Размер на бирке",
            "Состав на бирке",
            "Продавец",
            "Техническое задание",
        ],
        max_rows=120,
    )
    if header_row is None:
        raise ValueError("Не удалось найти строку заголовков в шаблоне ТЗ.")

    header_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(row=header_row, column=c).value or "").strip()
        if v:
            header_map[v.strip().lower()] = c

    start_row = header_row + 1

    tz_col = header_map.get("техническое задание")
    if not tz_col:
        raise ValueError("В шаблоне нет колонки 'Техническое задание'.")

    template_tz_value = None
    template_row_for_style = start_row
    for r in range(start_row, start_row + 200):
        v = ws.cell(row=r, column=tz_col).value
        if v is not None and str(v).strip() != "":
            template_tz_value = v
            template_row_for_style = r
            break

    if template_tz_value is None:
        template_tz_value = ws.cell(row=start_row, column=tz_col).value

    def _copy_row_style(src_row: int, dst_row: int):
        for c in range(1, ws.max_column + 1):
            src = ws.cell(row=src_row, column=c)
            dst = ws.cell(row=dst_row, column=c)
            # Копируем стиль безопасно (без обращения к приватному _style)
            dst.number_format = src.number_format
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.alignment = copy(src.alignment)
            dst.protection = copy(src.protection)

    def _enable_wrap_text(row: int) -> None:
        """Включает перенос текста для всех ячеек строки."""
        from openpyxl.styles import Alignment

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=c)
            a = cell.alignment or Alignment()
            cell.alignment = Alignment(
                horizontal=a.horizontal,
                vertical=a.vertical,
                text_rotation=a.text_rotation,
                wrap_text=True,
                shrink_to_fit=a.shrink_to_fit,
                indent=a.indent,
                relativeIndent=getattr(a, 'relativeIndent', 0),
                justifyLastLine=getattr(a, 'justifyLastLine', False),
                readingOrder=getattr(a, 'readingOrder', 0),
                textRotation=getattr(a, 'textRotation', a.text_rotation),
            )

    def _col_width_chars(col_idx: int) -> int:
        """Возвращает ширину колонки примерно в 'символах' (как в Excel)."""
        from openpyxl.utils import get_column_letter

        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        w = None
        if dim is not None:
            w = getattr(dim, 'width', None)
        # в шаблоне ширина может быть None => Excel покажет дефолт ~8.43
        try:
            w = float(w) if w is not None else 8.43
        except Exception:
            w = 8.43
        # уменьшим немного на поля
        return max(5, int(w))

    def _estimate_cell_lines(text: str, col_chars: int) -> int:
        """Оценивает, сколько строк займет текст в ячейке при wrap_text.

        Это приближение: Excel считает пиксели/шрифты точнее, но для авто-высоты хватает.
        """
        if text is None:
            return 1
        s = str(text)
        if not s.strip():
            return 1
        # учитываем явные переводы строк
        chunks = s.splitlines() or [s]
        lines = 0
        for ch in chunks:
            ch = ch.rstrip("\r")
            if ch == "":
                lines += 1
                continue
            # грубая оценка: количество оберток по ширине
            lines += max(1, (len(ch) + col_chars - 1) // col_chars)
        return max(1, lines)

    def _autofit_row_height(row: int, base_points: float = 15.0, max_points: float = 600.0) -> None:
        """Подбирает высоту строки по самому 'высокому' тексту в строке.

        base_points ~ высота одной строки (обычно 15pt для 11pt шрифта).
        """
        max_lines = 1
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=c)
            val = cell.value
            if val is None:
                continue
            # если перенос не включен, не раздуваем строку сильно
            if not (cell.alignment and cell.alignment.wrap_text):
                continue
            col_chars = _col_width_chars(c)
            max_lines = max(max_lines, _estimate_cell_lines(val, col_chars))

        height = min(max_points, base_points * max_lines)
        # если всего 1 строка — оставляем как есть (следуем шаблону)
        if max_lines > 1:
            ws.row_dimensions[row].height = height

    # упорядочим строки по категории/артикулу, чтобы было стабильнее
    lines_sorted = sorted(lines, key=lambda x: (str(x.get('category') or ''), str(x.get('article') or '')))

    for i, line in enumerate(lines_sorted):
        r = start_row + i
        if r != template_row_for_style:
            _copy_row_style(template_row_for_style, r)

        # Количество
        qty_col = header_map.get("количество")
        if qty_col:
            ws.cell(row=r, column=qty_col).value = int(line.get("ship_qty") or 0)

        # Артикул продавца
        art_col = header_map.get("артикул продавца")
        if art_col:
            ws.cell(row=r, column=art_col).value = str(line.get("article") or "").strip()

        # Наименование = категория
        name_col = header_map.get("наименование")
        if name_col:
            ws.cell(row=r, column=name_col).value = str(line.get("category") or "—").strip() or "—"

        # Баркод
        bc_col = header_map.get('баркод')
        if bc_col:
            ws.cell(row=r, column=bc_col).value = str(line.get('barcode') or '—').strip() or '—'

        # Цвет
        col_col = header_map.get('цвет на бирке')
        if col_col:
            ws.cell(row=r, column=col_col).value = str(line.get('color') or '—').strip() or '—'

        # Размер
        size_col = header_map.get('размер на бирке')
        if size_col:
            val = str(line.get('size') or '').strip()
            ws.cell(row=r, column=size_col).value = val or 'единый'

        # Состав
        comp_col = header_map.get('состав на бирке')
        if comp_col:
            ws.cell(row=r, column=comp_col).value = str(line.get('composition') or '—').strip() or '—'

        # Продавец
        seller_col = header_map.get("продавец")
        if seller_col:
            ws.cell(row=r, column=seller_col).value = seller

        # Техническое задание
        ws.cell(row=r, column=tz_col).value = template_tz_value

        # В ТЗ НЕ красим ячейки "Количество".
        # (Ранее оно подсвечивалось по текущим остаткам через get_fill_for_value.)

        # Включаем перенос текста для всей строки
        _enable_wrap_text(r)
        # Подбираем высоту строки под самый длинный текст (обычно 'Техническое задание')
        _autofit_row_height(r)

    # --- Итоговая сумма по количеству ---
    from openpyxl.styles import Font

    qty_col = header_map.get("количество")
    if qty_col and lines_sorted:
        last_row = start_row + len(lines_sorted) - 1
        sum_row = last_row + 2  # через 1 пустую строку

        _copy_row_style(template_row_for_style, sum_row)

        sum_cell = ws.cell(row=sum_row, column=qty_col)
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(qty_col)
        sum_cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{last_row})"

        # жирный + 12 (копируем текущий шрифт и меняем только нужное)
        base_font = sum_cell.font or Font()
        sum_cell.font = Font(
            name=base_font.name,
            sz=12,
            bold=True,
            italic=base_font.italic,
            underline=base_font.underline,
            strike=base_font.strike,
            color=base_font.color,
            vertAlign=base_font.vertAlign,
            outline=base_font.outline,
            shadow=base_font.shadow,
            charset=base_font.charset,
            family=base_font.family,
            scheme=base_font.scheme,
        )

        _enable_wrap_text(sum_row)
        _autofit_row_height(sum_row)

    wb.save(out_path)


async def _send_not_added_report(update: Update, full_text: str, filename: str = "tz_not_added.txt") -> None:
    """Отправляет полный отчёт и в чат, и отдельным txt-файлом.

    - В чат: целиком (с разбиением на части)
    - Файлом: целиком

    Это удобно: в чате можно быстро прочитать, а файл можно сохранить.
    """
    if not full_text or not str(full_text).strip():
        return

    # 1) В чат целиком (режем по 4к)
    await _send_long_text(update, full_text, filename=None)

    # 2) И дополнительно файлом целиком
    try:
        tmp_dir = tempfile.mkdtemp(prefix="tz_msg_")
        path = os.path.join(tmp_dir, filename)
        with open(path, "w", encoding="utf-8") as f:
            f.write(full_text)
        await update.message.reply_document(
            document=open(path, "rb"),
            caption="📄 Полный список позиций, не попавших в ТЗ",
        )
        shutil.rmtree(tmp_dir, ignore_errors=True)
    except Exception:
        # если файл по какой-то причине не отправился — не падаем
        logger.warning("TZ: не удалось отправить txt-файл отчёта not_added", exc_info=True)
