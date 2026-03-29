import sys
import os
import pandas as pd
import logging
import time
import requests
from telegram import Update, ReplyKeyboardRemove, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import CallbackContext, ConversationHandler
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from utils.template_loader import get_cabinet_articles_by_template_id
from utils.stock_control import resolve_stock_thresholds, apply_fill_to_cells
from utils.ozon_attributes import extract_attribute_values_from_product_attributes

# Настройка путей
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)
utils_dir = os.path.join(root_dir, 'utils')

if root_dir not in sys.path:
    sys.path.append(root_dir)
if utils_dir not in sys.path:
    sys.path.append(utils_dir)

logger = logging.getLogger(__name__)

from states import OZON_REMAINS_CABINET_CHOICE



# ======================
# Ozon API Класс
# ======================
class OzonAPI:
    def __init__(self, cabinet_id=1):
        from dotenv import load_dotenv
        load_dotenv()

        if cabinet_id == 1:
            self.client_id = os.getenv('OZON_CLIENT_ID_1')
            self.api_key = os.getenv('OZON_API_KEY_1')
        elif cabinet_id == 2:
            self.client_id = os.getenv('OZON_CLIENT_ID_2')
            self.api_key = os.getenv('OZON_API_KEY_2')
        elif cabinet_id == 3:
            self.client_id = os.getenv('OZON_CLIENT_ID_3')
            self.api_key = os.getenv('OZON_API_KEY_3')
        else:
            raise ValueError("Поддерживаются только cabinet_id 1, 2 или 3")

        if not self.client_id or not self.api_key:
            raise ValueError(f"❌ OZON_CLIENT_ID или OZON_API_KEY не заданы в .env для кабинета {cabinet_id}")

        self.base_url = "https://api-seller.ozon.ru"
        self.headers = {
            'Client-Id': self.client_id,
            'Api-Key': self.api_key,
            'Content-Type': 'application/json'
        }

    def get_product_list(self, limit=1000, last_id=""):
        url = f"{self.base_url}/v3/product/list"
        payload = {"filter": {"visibility": "ALL"}, "last_id": last_id, "limit": limit}
        try:
            response = requests.post(url, json=payload, headers=self.headers)
            return response.json() if response.status_code == 200 else None
        except Exception as e:
            logger.error(f"Ошибка при получении списка товаров: {e}")
            return None

    def get_product_info_list(self, offer_ids=None, product_ids=None, skus=None):
        url = f"{self.base_url}/v3/product/info/list"
        payload = {
            "offer_id": offer_ids or [],
            "product_id": product_ids or [],
            "sku": skus or []
        }
        try:
            response = requests.post(url, json=payload, headers=self.headers)
            return response.json() if response.status_code == 200 else None
        except Exception as e:
            logger.error(f"Ошибка при получении информации о товарах: {e}")
            return None

    def get_analytics_stocks(self, sku_list):
        url = f"{self.base_url}/v1/analytics/stocks"
        sku_list_clean = []
        for sku in sku_list:
            try:
                sku_list_clean.append(int(float(sku)))
            except (ValueError, TypeError):
                continue

        if not sku_list_clean:
            return []

        payload = {
            "skus": sku_list_clean,
            "turnover_grades": [
                "TURNOVER_GRADE_NONE", "DEFICIT", "POPULAR", "ACTUAL", "SURPLUS",
                "NO_SALES", "WAS_NO_SALES", "RESTRICTED_NO_SALES", "COLLECTING_DATA",
                "WAITING_FOR_SUPPLY", "WAS_DEFICIT", "WAS_POPULAR", "WAS_ACTUAL", "WAS_SURPLUS"
            ]
        }

        try:
            response = requests.post(url, json=payload, headers=self.headers, timeout=60)
            if response.status_code != 200:
                logger.warning(f"Ozon /v1/analytics/stocks -> {response.status_code}: {response.text[:500]}")
                return []
            data = response.json() or {}
            if isinstance(data.get('items'), list):
                return data.get('items') or []
            result = data.get("result")
            if isinstance(result, dict) and isinstance(result.get("items"), list):
                return result.get("items") or []
            if isinstance(result, list):
                return result
            return []
        except Exception as e:
            logger.error(f"Ошибка при получении аналитики остатков: {e}")
            return []

    def get_description_category_tree(self, language: str = "DEFAULT"):
        """/v1/description-category/tree"""
        url = f"{self.base_url}/v1/description-category/tree"
        payload = {"language": language}
        try:
            response = requests.post(url, json=payload, headers=self.headers)
            return response.json() if response.status_code == 200 else None
        except Exception as e:
            logger.error(f"Ошибка при получении дерева категорий: {e}")
            return None

    def get_description_category_attributes(self, description_category_id: int, type_id: int, language: str = "DEFAULT"):
        """/v1/description-category/attribute"""
        url = f"{self.base_url}/v1/description-category/attribute"
        payload = {
            "description_category_id": int(description_category_id),
            "type_id": int(type_id),
            "language": language,
        }
        try:
            response = requests.post(url, json=payload, headers=self.headers)
            return response.json() if response.status_code == 200 else None
        except Exception as e:
            logger.error(f"Ошибка при получении характеристик категории: {e}")
            return None

    def get_product_info_attributes(self, product_ids: list[int]):
        """/v4/product/info/attributes"""
        url = f"{self.base_url}/v4/product/info/attributes"
        payload = {
            "filter": {
                "product_id": [int(x) for x in (product_ids or [])],
                "visibility": "ALL",
            },
            "limit": len(product_ids or []),
        }
        try:
            response = requests.post(url, json=payload, headers=self.headers)
            return response.json() if response.status_code == 200 else None
        except Exception as e:
            logger.error(f"Ошибка при получении attributes товаров: {e}")
            return None


    def get_analytics_stocks_clustered(self, skus: list[int], cluster_ids: list[int] | None = None) -> list[dict]:
        """/v1/analytics/stocks с разрезом по кластерам/складам.

        Важно: skus <= 100.
        """
        url = f"{self.base_url}/v1/analytics/stocks"
        skus_clean: list[int] = []
        for s in skus or []:
            try:
                skus_clean.append(int(s))
            except Exception:
                continue
        if not skus_clean:
            return []

        payload: dict = {
            "skus": [str(x) for x in skus_clean],
            "turnover_grades": [
                "TURNOVER_GRADE_NONE", "DEFICIT", "POPULAR", "ACTUAL", "SURPLUS",
                "NO_SALES", "WAS_NO_SALES", "RESTRICTED_NO_SALES", "COLLECTING_DATA",
                "WAITING_FOR_SUPPLY", "WAS_DEFICIT", "WAS_POPULAR", "WAS_ACTUAL", "WAS_SURPLUS"
            ],
        }
        if cluster_ids:
            payload["cluster_ids"] = [str(int(x)) for x in cluster_ids if x is not None]

        try:
            resp = requests.post(url, json=payload, headers=self.headers, timeout=90)
            if resp.status_code != 200:
                logger.warning(f"Ozon /v1/analytics/stocks clustered -> {resp.status_code}: {resp.text}")
                return []
            data = resp.json() or {}
            return data.get("items") or []
        except Exception as e:
            logger.warning(f"Ozon /v1/analytics/stocks clustered exception: {e}")
            return []


def _build_type_name_map_from_tree(tree_result: list) -> dict[tuple[int, int], str]:
    """Строит маппинг (description_category_id, type_id) -> type_name.

    В дереве Ozon встречаются узлы без description_category_id (тогда берём родительский).
    """

    result: dict[tuple[int, int], str] = {}

    def _walk(nodes, current_dcid=None):
        for node in nodes or []:
            dcid = node.get("description_category_id", current_dcid)
            tpid = node.get("type_id")
            tname = node.get("type_name")
            if dcid is not None and tpid is not None and tname:
                try:
                    result[(int(dcid), int(tpid))] = str(tname).strip()
                except Exception:
                    pass

            _walk(node.get("children") or [], dcid)

    _walk(tree_result, None)
    return result


def _resolve_ozon_narrow_category(
    item_info: dict,
    type_name_map: dict[tuple[int, int], str],
) -> str:
    """Возвращает максимально узкую категорию для товара.

    Приоритет: type_name по (description_category_id,type_id). Иначе фолбэк на старую логику.
    """

    dcid = item_info.get("description_category_id")
    tpid = item_info.get("type_id")
    if dcid is not None and tpid is not None:
        try:
            name = type_name_map.get((int(dcid), int(tpid)))
            if name:
                return name
        except Exception:
            pass

    return _extract_ozon_category(item_info)


def clean_offer_id(offer_id_raw):
    """Только очищает от невидимых символов, НЕ меняет регистр"""
    try:
        if not offer_id_raw:
            return None
        s = str(offer_id_raw)
        s = ''.join(c for c in s if c.isprintable())
        s = s.strip()
        return s if s else None
    except Exception:
        return None


def chunk_list(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


# ======================
# Нормализация и группировка
# ======================

def normalize_art(art_str):
    """Нормализует строку: приводит к нижнему регистру, удаляет лишние пробелы, очищает от невидимых символов"""
    if not art_str:
        return ""
    s = str(art_str)
    s = ''.join(c for c in s if c.isprintable())
    s = s.strip().lower()
    return s


def group_ozon_remains_data(stock_data, template_id_to_cabinet_arts, template_id_to_name):
    """
    Группирует данные остатков по шаблонным артикулам.

    :param stock_data: dict {offer_id: {"available": ..., "returning": ..., "prepare": ...}}
    :param template_id_to_cabinet_arts: dict {template_id: [cabinet_art1, cabinet_art2, ...]}
    :param template_id_to_name: dict {template_id: "Шаблонное название"}
    :return: grouped (по template_id), unmatched (артикулы без привязки)
    """
    stock_data_clean = {}
    for art, data in stock_data.items():
        clean_art = normalize_art(art)
        if clean_art:
            stock_data_clean[clean_art] = data

    cabinet_art_to_template_id = {}
    for template_id, arts in template_id_to_cabinet_arts.items():
        for art in arts:
            clean_art = normalize_art(art)
            if clean_art:
                cabinet_art_to_template_id[clean_art] = template_id

    grouped = {}
    unmatched = {}

    for clean_art, data in stock_data_clean.items():
        template_id = cabinet_art_to_template_id.get(clean_art)

        if template_id is not None:
            if template_id not in grouped:
                grouped[template_id] = {
                    'name': template_id_to_name.get(template_id, f"ID {template_id}"),
                    'available': 0,
                    'returning': 0,
                    'prepare': 0
                }
            grouped[template_id]['available'] += data['available']
            grouped[template_id]['returning'] += data['returning']
            grouped[template_id]['prepare'] += data['prepare']
        else:
            unmatched[clean_art] = {
                'name': f"НЕОПОЗНАННЫЙ: {clean_art}",
                'available': data['available'],
                'returning': data['returning'],
                'prepare': data['prepare']
            }

    return grouped, unmatched


def _extract_ozon_category(item_info):
    for key in ("category", "category_name", "category_id"):
        value = item_info.get(key)
        if value is not None and str(value).strip() != "":
            return str(value).strip()
    return "—"


# ======================
# Обработчики
# ======================

async def start_ozon_remains(update: Update, context: CallbackContext) -> int:
    """Начало — выбор кабинета Ozon"""
    context.user_data['current_flow'] = 'remains'

    keyboard = [
        [InlineKeyboardButton("🏪 Озон_1 Nimba", callback_data='cabinet_1')],
        [InlineKeyboardButton("🏬 Озон_2 Galioni", callback_data='cabinet_2')],
        [InlineKeyboardButton("🏢 Озон_3 AGNIA", callback_data='cabinet_3')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        "🏢 Выберите кабинет Ozon для выгрузки остатков:",
        reply_markup=reply_markup
    )

    return OZON_REMAINS_CABINET_CHOICE


async def handle_cabinet_choice(update: Update, context: CallbackContext) -> int:
    """Обработка выбора кабинета Ozon — сразу генерируем оба отчёта"""
    query = update.callback_query
    await query.answer()

    # Явный alias на локально объявленную функцию генерации чанков,
    # чтобы PyCharm корректно резолвил символ внутри этой функции.
    chunk_list_fn = chunk_list

    cabinet_data = query.data
    cabinet_map = {
        'cabinet_1': 1,
        'cabinet_2': 2,
        'cabinet_3': 3
    }
    if cabinet_data not in cabinet_map:
        await query.message.reply_text("❌ Неизвестный кабинет.")
        return ConversationHandler.END

    cabinet_id = cabinet_map[cabinet_data]
    context.user_data['ozon_cabinet_id'] = cabinet_id

    loading_message = await query.message.edit_text(f"⏳ Получаю остатки с Ozon API (Озон {cabinet_id})...")
    context.user_data['ozon_remains_loading_message_id'] = loading_message.message_id

    try:
        ozon = OzonAPI(cabinet_id=cabinet_id)

        # Узкая категория: подтягиваем дерево категорий 1 раз и строим словарь (dcid,type_id)->type_name
        type_name_map: dict[tuple[int, int], str] = {}
        try:
            tree = ozon.get_description_category_tree(language="DEFAULT")
            if tree and tree.get("result"):
                type_name_map = _build_type_name_map_from_tree(tree.get("result"))
        except Exception as e:
            logger.warning(f"Не удалось построить словарь type_name по дереву категорий Ozon: {e}")

        # --- Получение данных ---
        product_list = ozon.get_product_list(limit=1000)
        if not product_list:
            raise Exception("Не удалось получить список товаров")

        items = product_list.get('result', {}).get('items', [])
        if not items:
            raise Exception("Товары не найдены")

        offer_ids = []
        for item in items:
            offer_id = clean_offer_id(item.get('offer_id'))
            if offer_id:
                offer_ids.append(offer_id)

        # === Подтягиваем info/list один раз: sku + name + category ===
        all_skus = []
        offer_id_to_name = {}
        offer_id_to_category = {}
        offer_id_to_sku: dict[str, int] = {}

        for chunk in chunk_list_fn(offer_ids, 1000):
            product_info_response = ozon.get_product_info_list(offer_ids=chunk)
            if not product_info_response:
                continue

            items_in_response = []
            if 'result' in product_info_response and 'items' in product_info_response['result']:
                items_in_response = product_info_response['result']['items']
            elif 'items' in product_info_response:
                items_in_response = product_info_response['items']
            elif isinstance(product_info_response.get('result'), list):
                items_in_response = product_info_response['result']
            else:
                continue

            for item_info in items_in_response:
                oid = clean_offer_id(item_info.get('offer_id'))
                sku = item_info.get('sku')
                name = item_info.get('name', '—')
                category = _resolve_ozon_narrow_category(item_info, type_name_map)
                if oid and sku:
                    try:
                        sku_int = int(sku)
                    except Exception:
                        continue
                    all_skus.append(sku_int)
                    offer_id_to_sku[oid] = sku_int
                    offer_id_to_name[oid] = name
                    offer_id_to_category[oid] = category

            time.sleep(0.5)

        if not all_skus:
            raise Exception("Не удалось получить SKU")

        # === NEW: Остатки по кластерам через /v1/analytics/stocks (кластерный разрез) ===
        cluster_sheet_df = None
        cluster_sheet_headers = None
        try:
            # --- правило ручной категории "Туники детские" ---
            # Шаблонные артикулы, для которых все привязанные исходные offer_id должны считаться категорией "Туники детские".
            CHILD_TUNICS_TEMPLATES = {
                'детская туника белая',
                'детская туника розовая',
                'детская туника голубая',
                'детская туника фиолетовая',
            }

            # Пытаемся построить множество исходных артикулах (offer_id), которые привязаны к этим шаблонам
            child_tunics_offer_ids_norm: set[str] = set()
            try:
                sheet_map_tmp = {
                    1: "Отдельно Озон Nimba",
                    2: "Отдельно Озон Galioni",
                    3: "Отдельно Озон AGNIA",
                }
                sheet_name_tmp = sheet_map_tmp.get(cabinet_id)
                if sheet_name_tmp:
                    template_id_to_name_tmp, template_id_to_cabinet_arts_tmp = get_cabinet_articles_by_template_id(sheet_name_tmp)
                    for tid, tname in (template_id_to_name_tmp or {}).items():
                        if normalize_art(tname) in CHILD_TUNICS_TEMPLATES:
                            for art in template_id_to_cabinet_arts_tmp.get(tid) or []:
                                norm = normalize_art(art)
                                if norm:
                                    child_tunics_offer_ids_norm.add(norm)
                logger.info(f"Ozon кабинет {cabinet_id}: child tunics linked offer_ids={len(child_tunics_offer_ids_norm)}")
            except Exception as e:
                logger.warning(f"Ozon кабинет {cabinet_id}: не удалось построить список детских туник из базы: {e}")

            # 1) берем список кластеров (для информации)
            url_clusters = f"{ozon.base_url}/v1/cluster/list"
            resp_clusters = requests.post(
                url_clusters,
                json={"cluster_type": "CLUSTER_TYPE_OZON"},
                headers=ozon.headers,
                timeout=60,
            )
            cluster_ids: list[int] = []
            if resp_clusters.status_code == 200:
                data_clusters = resp_clusters.json() or {}
                for cl in (data_clusters.get("clusters") or []):
                    cid = cl.get("id")
                    try:
                        if cid is not None:
                            cluster_ids.append(int(cid))
                    except Exception:
                        continue
            logger.info(f"Ozon кабинет {cabinet_id}: clusters list ids={len(cluster_ids)}")

            t0 = time.time()
            rows_raw: list[dict] = []
            uniques_clusters: set[str] = set()

            # ВАЖНО: не передаём cluster_ids, чтобы получить полный разрез
            for sku_chunk in chunk_list_fn(list(set(all_skus)), 100):
                items_clustered = ozon.get_analytics_stocks_clustered(skus=sku_chunk, cluster_ids=None)
                for it in items_clustered or []:
                    oid = clean_offer_id(it.get("offer_id"))
                    if not oid:
                        continue
                    cname = str(it.get("cluster_name") or it.get("clusterName") or "—").strip() or "—"
                    qty = it.get("available_stock_count")
                    try:
                        qty_int = int(qty) if qty is not None else 0
                    except Exception:
                        qty_int = 0
                    rows_raw.append({"Артикул": oid, "Кластер": cname, "Кол-во": qty_int})
                    uniques_clusters.add(cname)

            logger.info(
                f"Ozon кабинет {cabinet_id}: /v1/analytics/stocks clustered rows={len(rows_raw)} clusters={len(uniques_clusters)} за {time.time() - t0:.2f}s"
            )

            # --- строим маппинг артикул -> категория (как в ТЗ), с ручным оверрайдом на "Туники детские" ---
            # В offer_id_to_category (ключи в исходном регистре) могут быть расхождения, поэтому делаем нормализованный словарь.
            offer_id_to_category_norm: dict[str, str] = {}
            for k, v in (offer_id_to_category or {}).items():
                nk = normalize_art(k)
                if nk:
                    offer_id_to_category_norm[nk] = v or '—'

            # применяем ручную категорию
            for norm_oid in child_tunics_offer_ids_norm:
                offer_id_to_category_norm[norm_oid] = "Туники детские"

            def _category_for_offer_id(oid: str) -> str:
                n = normalize_art(oid)
                return offer_id_to_category_norm.get(n, '—')

            if rows_raw:
                df_long = pd.DataFrame(rows_raw)

                # Подмешиваем категорию к каждой строке
                df_long.insert(0, "Категория", df_long["Артикул"].map(_category_for_offer_id))

                # Агрегируем сразу по категориям/кластерам
                df_cat = df_long.pivot_table(
                    index="Категория",
                    columns="Кластер",
                    values="Кол-во",
                    aggfunc="sum",
                    fill_value=0,
                )

                df_cat.reset_index(inplace=True)
                cluster_names = [c for c in df_cat.columns if c != "Категория"]
                cluster_names_sorted = sorted(cluster_names, key=lambda x: str(x).lower())
                df_cat = df_cat[["Категория", *cluster_names_sorted]]
                df_cat["ВСЕГО"] = df_cat[cluster_names_sorted].sum(axis=1)

                cluster_sheet_df = df_cat.sort_values(
                    by="Категория",
                    key=lambda x: x.astype(str).str.lower(),
                ).reset_index(drop=True)
                cluster_sheet_headers = ["Категория", *cluster_names_sorted, "ВСЕГО"]

                # # Пустая строка + ИТОГО - УБИРАЕМ ПО ПРОСЬБЕ
                # empty_row = {k: "" for k in cluster_sheet_headers}
                # totals = {"Категория": "ИТОГО"}
                # for cn in cluster_names_sorted:
                #     totals[cn] = int(cluster_sheet_df[cn].sum())
                # totals["ВСЕГО"] = int(cluster_sheet_df["ВСЕГО"].sum())
                #
                # cluster_sheet_df = pd.concat(
                #     [
                #         cluster_sheet_df,
                #         pd.DataFrame([empty_row]),
                #         pd.DataFrame([totals]),
                #     ],
                #     ignore_index=True,
                # )
            else:
                # если вообще нет данных clustered, всё равно кладём пустую табличку
                cluster_sheet_df = pd.DataFrame([], columns=["Категория", "ВСЕГО"])
                cluster_sheet_headers = ["Категория", "ВСЕГО"]

        except Exception as e:
            logger.warning(
                f"Ozon кабинет {cabinet_id}: не удалось сформировать лист по кластерам через /v1/analytics/stocks: {e}"
            )

        stock_dict = {}

        for sku_chunk in chunk_list_fn(all_skus, 100):
            items = ozon.get_analytics_stocks(sku_chunk)
            for item in items:
                offer_id = clean_offer_id(item.get('offer_id'))
                if not offer_id:
                    continue

                if offer_id in stock_dict:
                    stock_dict[offer_id]['available_stock_count'] += item.get('available_stock_count', 0)
                    stock_dict[offer_id]['return_from_customer_stock_count'] += item.get(
                        'return_from_customer_stock_count', 0)
                    stock_dict[offer_id]['valid_stock_count'] += item.get('valid_stock_count', 0)
                else:
                    stock_dict[offer_id] = {
                        'name': item.get('name', offer_id_to_name.get(offer_id, '—')),
                        'category': offer_id_to_category.get(offer_id, '—'),
                        'available_stock_count': item.get('available_stock_count', 0),
                        'return_from_customer_stock_count': item.get('return_from_customer_stock_count', 0),
                        'valid_stock_count': item.get('valid_stock_count', 0),
                        'ads': item.get('ads', 0),
                        'idc': item.get('idc', 0)
                    }
            time.sleep(0.5)

        missing_offer_ids = list(set(offer_ids) - set(stock_dict.keys()))
        if missing_offer_ids:
            for chunk in chunk_list_fn(missing_offer_ids, 100):
                info_response = ozon.get_product_info_list(offer_ids=chunk)
                if not info_response:
                    continue

                items_in_response = []
                if 'result' in info_response and 'items' in info_response['result']:
                    items_in_response = info_response['result']['items']
                elif 'items' in info_response:
                    items_in_response = info_response['items']
                elif isinstance(info_response.get('result'), list):
                    items_in_response = info_response['result']
                else:
                    continue

                for item in items_in_response:
                    offer_id = clean_offer_id(item.get('offer_id'))
                    if not offer_id:
                        continue

                    stocks = item.get('stocks', {})
                    name = item.get('name', '—')
                    # Узкая категория для отсутствующих остатков
                    category = _resolve_ozon_narrow_category(item, type_name_map)
                    stock_dict[offer_id] = {
                        'name': name,
                        'category': category,
                        'available_stock_count': stocks.get('present', 0),
                        'return_from_customer_stock_count': 0,
                        'valid_stock_count': stocks.get('reserved', 0),
                        'ads': 0,
                        'idc': 0
                    }

                time.sleep(0.5)

        # === 1. Отчёт по исходным артикулам ===
        # --- Цвет и размер через /v4/product/info/attributes ---
        offer_id_to_color: dict[str, str] = {}
        offer_id_to_size: dict[str, str] = {}
        offer_id_to_turnover: dict[str, float | int | None] = {}

        # Собираем product_id и мету (dcid/type_id) для offer_id
        offer_id_to_product_id: dict[str, int] = {}
        offer_id_to_dcid: dict[str, int] = {}
        offer_id_to_type_id: dict[str, int] = {}

        # Для этого ещё раз получим info/list (у нас offer_ids уже есть выше по коду)
        for chunk in chunk_list_fn(offer_ids, 1000):
            info = ozon.get_product_info_list(offer_ids=chunk)
            if not info:
                continue
            items_in_response = []
            if 'result' in info and 'items' in info['result']:
                items_in_response = info['result']['items']
            elif 'items' in info:
                items_in_response = info['items']
            elif isinstance(info.get('result'), list):
                items_in_response = info['result']
            for it in items_in_response or []:
                oid = clean_offer_id(it.get('offer_id'))
                if not oid:
                    continue
                pid = it.get('id') or it.get('product_id')
                if pid is not None:
                    try:
                        offer_id_to_product_id[oid] = int(pid)
                    except Exception:
                        pass
                dcid = it.get('description_category_id')
                if dcid is not None:
                    try:
                        offer_id_to_dcid[oid] = int(dcid)
                    except Exception:
                        pass
                tpid = it.get('type_id')
                if tpid is not None:
                    try:
                        offer_id_to_type_id[oid] = int(tpid)
                    except Exception:
                        pass

        # Кеш meta атрибутов по (dcid,type_id)
        meta_cache: dict[tuple[int, int], list[dict]] = {}
        resolved_attr_ids: dict[tuple[int, int], dict[str, int | None]] = {}

        def _pick_attr_id(attrs: list[dict], kind: str) -> int | None:
            # kind: 'color' | 'size'
            if not attrs:
                return None
            kind = kind.lower().strip()

            if kind == 'color':
                preferred = {
                    'цвет',
                    'цвет товара',
                    'основной цвет',
                    'цвет (основной)',
                    'основной цвет товара',
                }
                keywords = ('цвет', 'color', 'colour')
            else:
                preferred = {
                    'размер',
                    'размер (ru)',
                    'размер ru',
                    'российский размер',
                    'размер производителя',
                    'размер одежды',
                }
                keywords = ('размер', 'size')

            best_id = None
            best_score = -1
            for a in attrs:
                name = str(a.get('name') or '').strip().lower()
                if not name:
                    continue
                score = 0
                if name in preferred:
                    score += 100
                if any(k in name for k in keywords):
                    score += 20
                # лёгкий плюс за не-collection (для размера чаще одно значение)
                if kind != 'color' and not bool(a.get('is_collection')):
                    score += 5
                if score > best_score:
                    try:
                        best_id = int(a.get('id'))
                        best_score = score
                    except Exception:
                        continue
            return best_id

        def _get_attr_ids_for_pair(dcid: int, tpid: int) -> dict[str, int | None]:
            key = (int(dcid), int(tpid))
            if key in resolved_attr_ids:
                return resolved_attr_ids[key]
            if key not in meta_cache:
                resp = ozon.get_description_category_attributes(key[0], key[1], language='DEFAULT')
                meta_cache[key] = (resp or {}).get('result') or []
            attrs = meta_cache.get(key) or []
            ids = {
                'color': _pick_attr_id(attrs, 'color'),
                'size': _pick_attr_id(attrs, 'size'),
            }
            resolved_attr_ids[key] = ids
            return ids

        product_ids = list({pid for pid in offer_id_to_product_id.values() if pid is not None})
        # батчим, чтобы не было таймаутов
        for chunk in chunk_list_fn(product_ids, 250):
            attrs_resp = ozon.get_product_info_attributes(chunk)
            for info_item in (attrs_resp or {}).get('result') or []:
                oid = clean_offer_id(info_item.get('offer_id'))
                if not oid:
                    continue
                dcid = offer_id_to_dcid.get(oid)
                tpid = offer_id_to_type_id.get(oid)
                if not dcid or not tpid:
                    continue
                ids = _get_attr_ids_for_pair(dcid, tpid)

                color_attr_id = ids.get('color')
                if color_attr_id:
                    str_vals, _ = extract_attribute_values_from_product_attributes(info_item, int(color_attr_id))
                    if str_vals:
                        offer_id_to_color[oid] = ", ".join([v for v in str_vals if str(v).strip()])

                size_attr_id = ids.get('size')
                if size_attr_id:
                    str_vals, _ = extract_attribute_values_from_product_attributes(info_item, int(size_attr_id))
                    if str_vals:
                        offer_id_to_size[oid] = ", ".join([v for v in str_vals if str(v).strip()])

        raw_data = []
        for offer_id, data in stock_dict.items():
            category = data.get('category', '—')
            available = data['available_stock_count']
            returning = data['return_from_customer_stock_count']
            prepare = data['valid_stock_count']
            total = available + returning + prepare

            ads = data.get('ads', '')
            idc = data.get('idc', '')

            raw_data.append({
                'Категория': category,
                'Артикул': offer_id,
                'Доступно на складах': available,
                'Возвращаются от покупателей': returning,
                'Подготовка к продаже': prepare,
                'Итого на МП': total,
                'Среднесуточные продажи (28 дн)': ads,
                'Запас (дни)': idc,
            })

        df_raw = pd.DataFrame(raw_data).sort_values(by='Категория', key=lambda x: x.str.lower()).reset_index(
            drop=True)
        headers_raw = [
            "Категория", "Артикул", "Доступно на складах", "Возвращаются от покупателей",
            "Подготовка к продаже", "Итого на МП", "Среднесуточные продажи (28 дн)", "Запас (дни)"
        ]

        # === 2. Отчёт по шаблону Nimba/Galioni ===
        sheet_map = {
            1: "Отдельно Озон Nimba",
            2: "Отдельно Озон Galioni",
            3: "Отдельно Озон AGNIA"
        }
        sheet_name = sheet_map.get(cabinet_id)
        if not sheet_name:
            raise ValueError(f"Неподдерживаемый кабинет Ozon: {cabinet_id}")

        template_id_to_name, template_id_to_cabinet_arts = get_cabinet_articles_by_template_id(sheet_name)

        linked_template_ids = set(template_id_to_cabinet_arts.keys())

        # Получаем main_ids_ordered — ID в порядке появления в Excel (без дубликатов)
        template_path = os.path.join(root_dir, "База данных артикулов для выкупов и начислений.xlsx")
        if not os.path.exists(template_path):
            template_path = "База данных артикулов для выкупов и начислений.xlsx"
        df_order = pd.read_excel(template_path, sheet_name=sheet_name)
        main_ids_ordered = []
        seen = set()
        for _, row in df_order.iterrows():
            if not pd.isna(row.get('ID')):
                tid = int(row['ID'])
                if tid not in seen:
                    main_ids_ordered.append(tid)
                    seen.add(tid)

        cabinet_arts_set = set()
        for arts in template_id_to_cabinet_arts.values():
            for art in arts:
                cabinet_arts_set.add(normalize_art(art))

        template_rows_to_color = []
        for idx, id_val in enumerate(main_ids_ordered, start=3):
            if id_val in linked_template_ids:
                template_rows_to_color.append(idx)

        # Подготовка stock_data
        stock_data = {}
        for offer_id, data in stock_dict.items():
            stock_data[offer_id] = {
                "available": data['available_stock_count'],
                "returning": data['return_from_customer_stock_count'],
                "prepare": data['valid_stock_count']
            }

        # Группировка по шаблонам
        grouped, unmatched = group_ozon_remains_data(
            stock_data,
            template_id_to_cabinet_arts,
            template_id_to_name
        )

        template_data = []
        for id_val in main_ids_ordered:
            if id_val in grouped:
                d = grouped[id_val]
                total = d['available'] + d['returning'] + d['prepare']
                template_data.append({
                    'Артикул': d['name'],
                    'Доступно на складах': d['available'],
                    'Возвращаются от покупателей': d['returning'],
                    'Подготовка к продаже': d['prepare'],
                    'Итого на МП': total
                })
            else:
                name = template_id_to_name.get(id_val, f"ID {id_val}")
                template_data.append({
                    'Артикул': name,
                    'Доступно на складах': 0,
                    'Возвращаются от покупателей': 0,
                    'Подготовка к продаже': 0,
                    'Итого на МП': 0
                })

        for art, d in unmatched.items():
            total = d['available'] + d['returning'] + d['prepare']
            template_data.append({
                'Артикул': f"НЕОПОЗНАННЫЙ: {art}",
                'Доступно на складах': d['available'],
                'Возвращаются от покупателей': d['returning'],
                'Подготовка к продаже': d['prepare'],
                'Итого на МП': total
            })

        df_template = pd.DataFrame(template_data)
        headers_template = ["Артикул", "Доступно на складах", "Возвращаются от покупателей", "Подготовка к продаже",
                            "Итого на МП"]

        thresholds = resolve_stock_thresholds(context, query.message.chat_id)
        raw_rows_to_color = []
        for idx, art in enumerate(df_raw["Артикул"], start=3):
            if normalize_art(art) in cabinet_arts_set:
                raw_rows_to_color.append(idx)

        # === Сводка по всем остаткам ===
        total_available = sum(data['available_stock_count'] for data in stock_dict.values())
        total_returning = sum(data['return_from_customer_stock_count'] for data in stock_dict.values())
        total_prepare = sum(data['valid_stock_count'] for data in stock_dict.values())
        total_mp = total_available + total_returning + total_prepare

        def fmt_num(x):
            return f"{x:,}".replace(",", " ")

        summary_text = (
            f"📊 <b>Сводка по остаткам Ozon</b>\n"
            f"Кабинет: <b>Озон {cabinet_id}</b>\n\n"
            f"📦 <b>Доступно на складах:</b> {fmt_num(total_available)} шт\n"
            f"↩️ <b>Возвращаются от покупателей:</b> {fmt_num(total_returning)} шт\n"
            f"🔄 <b>Подготовка к продаже:</b> {fmt_num(total_prepare)} шт\n"
            f"✅ <b>Итого на МП:</b> {fmt_num(total_mp)} шт"
        )

        # ✅ Создаём Excel с двумя листами
        report_path = "Ozon_Remains_Report.xlsx"
        create_excel_with_two_sheets(
            df_raw,
            headers_raw,
            df_template,
            headers_template,
            report_path,
            thresholds=thresholds,
            template_rows_to_color=template_rows_to_color,
            raw_rows_to_color=raw_rows_to_color,
            df_clusters=cluster_sheet_df,
            headers_clusters=cluster_sheet_headers,
        )

        # 📤 Отправляем файл
        await query.message.reply_document(
            document=open(report_path, 'rb'),
            caption="📊 Отчёт по остаткам Ozon: исходные артикулы, шаблон + остатки по кластерам",
            reply_markup=ReplyKeyboardRemove()
        )

        # 💬 Отправляем сводку
        await query.message.reply_text(summary_text, parse_mode="HTML")

        # 🧹 Очистка
        if os.path.exists(report_path):
            os.remove(report_path)

        # Удаляем сообщение о загрузке
        chat_id = query.message.chat_id
        try:
            loading_msg_id = context.user_data.get('ozon_remains_loading_message_id')
            if loading_msg_id:
                await context.bot.delete_message(chat_id=chat_id, message_id=loading_msg_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить сообщение о загрузке остатков: {e}")

    except Exception as e:
        logger.error(f"Ошибка при получении данных: {str(e)}", exc_info=True)
        await query.message.reply_text(
            f"❌ Ошибка: {str(e)}",
            reply_markup=ReplyKeyboardRemove()
        )
        # Удаляем сообщение о загрузке даже при ошибке
        chat_id = query.message.chat_id
        try:
            loading_msg_id = context.user_data.get('ozon_remains_loading_message_id')
            if loading_msg_id:
                await context.bot.delete_message(chat_id=chat_id, message_id=loading_msg_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить сообщение о загрузке остатков при ошибке: {e}")

    return ConversationHandler.END


def create_excel_with_two_sheets(
        df_raw,
        headers_raw,
        df_template,
        headers_template,
        filename,
        thresholds=None,
        template_rows_to_color=None,
        raw_rows_to_color=None,
        df_clusters=None,
        headers_clusters=None,
):
    """Создаёт Excel с листами: 'Остатки шаблон Nimba', 'Остатки исходные артикулы' (+ опционально 'Остатки по кластерам')."""
    wb = Workbook()
    wb.remove(wb.active)  # удаляем дефолтный лист

    # Сначала — шаблон Nimba/Galioni
    ws1 = wb.create_sheet(title="Остатки шаблон Nimba")
    _write_sheet(ws1, df_template, headers_template, has_name=False)
    if template_rows_to_color and thresholds:
        apply_fill_to_cells(ws1, template_rows_to_color, [5], thresholds)

    # Затем — исходные артикулы
    ws2 = wb.create_sheet(title="Остатки исходные артикулы")
    _write_sheet(ws2, df_raw, headers_raw, has_name=True)
    if raw_rows_to_color and thresholds:
        apply_fill_to_cells(ws2, raw_rows_to_color, [6], thresholds)

    # NEW: Остатки по кластерам/складам
    if df_clusters is not None and headers_clusters:
        ws3 = wb.create_sheet(title="Остатки по кластерам")
        _write_sheet(ws3, df_clusters, headers_clusters, has_name=False)

    wb.save(filename)


def _write_sheet(ws, df, headers, has_name):
    """Вспомогательная функция для записи одного листа с форматированием"""
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Объединение ячеек в заголовке
    ws.merge_cells('A1:A2')
    if has_name:
        ws.merge_cells('B1:B2')

    data_start_row = 3
    sum_row = 2

    # Данные
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), data_start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border

    # Суммы
    num_rows = len(df)
    if num_rows > 0:
        start_col_index = 3 if has_name else 2
        for col in range(start_col_index, len(headers) + 1):
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_start_row + num_rows - 1})"
            cell = ws.cell(row=sum_row, column=col, value=formula)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border

    # Автоподбор ширины
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
